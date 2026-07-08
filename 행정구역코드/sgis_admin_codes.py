# -*- coding: utf-8 -*-
"""
SGIS 행정구역 코드표 생성기
 시도 → 시군구 → 행정동 코드/명칭을 SGIS Open API(stage)로 전부 수집해
 엑셀(.xlsx)로 저장한다.
 구조: 연번 | 시도코드 | 시도명 | 시군구코드 | 시군구명 | 행정동코드 | 행정동명
 시트: '전체' 1개 + 시도별 17개
"""
import time
import urllib.request
import urllib.parse
import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# SGIS 인증키는 환경변수로 주입(코드에 하드코딩 금지 — 저장소 방침).
#   실행 예) Windows PowerShell:
#     $env:SGIS_KEY="발급받은키"; $env:SGIS_SECRET="발급받은시크릿"; python sgis_admin_codes.py
#   실행 예) mac/linux:
#     SGIS_KEY=... SGIS_SECRET=... python3 sgis_admin_codes.py
CONSUMER_KEY = os.environ.get("SGIS_KEY", "")
CONSUMER_SECRET = os.environ.get("SGIS_SECRET", "")
BASE = "https://sgisapi.mods.go.kr/OpenAPI3"  # 국가데이터처 이관 후 주소(구 sgisapi.kostat.go.kr)

OUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUT_XLSX = os.path.join(OUT_DIR, "행정구역코드_전국.xlsx")
SIDO_DIR = os.path.join(OUT_DIR, "시도별파일")  # 시도 하나당 엑셀 1개

_token = None


def _get(url):
    for attempt in range(5):
        try:
            with urllib.request.urlopen(url, timeout=30) as r:
                return json.loads(r.read().decode("utf-8"))
        except Exception as e:
            print(f"   재시도({attempt+1}) {e}")
            time.sleep(1.5)
    raise RuntimeError("요청 실패: " + url)


def auth():
    global _token
    url = (f"{BASE}/auth/authentication.json"
           f"?consumer_key={CONSUMER_KEY}&consumer_secret={CONSUMER_SECRET}")
    data = _get(url)
    if data.get("errCd") != 0:
        raise RuntimeError("인증 실패: " + json.dumps(data, ensure_ascii=False))
    _token = data["result"]["accessToken"]
    print("[인증 성공] 토큰:", _token)


def stage(cd=None):
    """cd 하위 행정구역 목록. cd 없으면 시도 목록."""
    params = {"accessToken": _token}
    if cd is not None:
        params["cd"] = cd
    url = f"{BASE}/addr/stage.json?" + urllib.parse.urlencode(params)
    data = _get(url)
    # 토큰 만료 시 재인증
    if data.get("errCd") not in (0, None):
        msg = str(data.get("errMsg", ""))
        if "인증" in msg or "token" in msg.lower() or data.get("errCd") in (-401, 401):
            print("   토큰 만료 → 재인증")
            auth()
            params["accessToken"] = _token
            url = f"{BASE}/addr/stage.json?" + urllib.parse.urlencode(params)
            data = _get(url)
    res = data.get("result")
    return res if isinstance(res, list) else []


def collect():
    """전국 행정동 단위 행 리스트 반환."""
    rows = []
    sido_list = stage()
    print(f"시도 {len(sido_list)}개 수집 시작")
    for sido in sido_list:
        sido_cd, sido_nm = sido["cd"], sido["addr_name"]
        sgg_list = stage(sido_cd)
        print(f"  [{sido_cd} {sido_nm}] 시군구 {len(sgg_list)}개")
        for sgg in sgg_list:
            sgg_cd, sgg_nm = sgg["cd"], sgg["addr_name"]
            dong_list = stage(sgg_cd)
            if not dong_list:
                # 하위 없음 → 이 자체가 최하위(세종 등). 행정동 칸에 그대로.
                rows.append({
                    "시도코드": sido_cd, "시도명": sido_nm,
                    "시군구코드": "", "시군구명": "",
                    "행정동코드": sgg_cd, "행정동명": sgg_nm,
                })
                continue
            for dong in dong_list:
                rows.append({
                    "시도코드": sido_cd, "시도명": sido_nm,
                    "시군구코드": sgg_cd, "시군구명": sgg_nm,
                    "행정동코드": dong["cd"], "행정동명": dong["addr_name"],
                })
            time.sleep(0.03)
    print(f"총 {len(rows)}개 행정동 수집 완료")
    return rows


# ---------- 엑셀 스타일 ----------
HEADERS = ["연번", "시도코드", "시도명", "시군구코드", "시군구명", "행정동코드", "행정동명"]
WIDTHS = [6, 10, 16, 12, 20, 12, 16]
HEAD_FILL = PatternFill("solid", fgColor="2F5597")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CODE_COLS = {2, 4, 6}  # 시도/시군구/행정동 코드 → 텍스트 서식(앞자리 0 보존)


def style_sheet(ws, rows):
    # 헤더
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, c, h)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    # 데이터
    for i, row in enumerate(rows, 1):
        vals = [i, row["시도코드"], row["시도명"], row["시군구코드"],
                row["시군구명"], row["행정동코드"], row["행정동명"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(i + 1, c, v)
            cell.border = BORDER
            if c == 1 or c in CODE_COLS:
                cell.alignment = Alignment(horizontal="center")
            if c in CODE_COLS:
                cell.number_format = "@"
    # 폭/틀고정/필터
    for c, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(rows)+1}"


def build_xlsx(rows):
    wb = Workbook()
    # 전체 시트
    ws_all = wb.active
    ws_all.title = "전체"
    style_sheet(ws_all, rows)

    # 시도별 시트
    seen = {}
    order = []
    for r in rows:
        key = (r["시도코드"], r["시도명"])
        seen.setdefault(key, []).append(r)
        if key not in order:
            order.append(key)
    for (cd, nm) in order:
        # 시트명 31자 제한 + 금지문자 제거
        title = f"{cd}_{nm}"[:31].replace("/", "").replace("\\", "")
        ws = wb.create_sheet(title)
        style_sheet(ws, seen[(cd, nm)])

    wb.save(OUT_XLSX)
    print(f"[저장 완료] {OUT_XLSX}  (행정동 {len(rows)}개, 시트 {len(order)+1}개)")

    # 시도별 개별 파일
    os.makedirs(SIDO_DIR, exist_ok=True)
    for (cd, nm) in order:
        wb1 = Workbook()
        ws1 = wb1.active
        ws1.title = nm[:31].replace("/", "").replace("\\", "")
        style_sheet(ws1, seen[(cd, nm)])  # 연번은 파일마다 1부터 다시 매겨짐
        fname = f"{cd}_{nm}.xlsx".replace("/", "").replace("\\", "")
        wb1.save(os.path.join(SIDO_DIR, fname))
    print(f"[시도별 파일 {len(order)}개 저장] {SIDO_DIR}")


if __name__ == "__main__":
    if not CONSUMER_KEY or not CONSUMER_SECRET:
        raise SystemExit(
            "환경변수 SGIS_KEY / SGIS_SECRET 를 먼저 설정하세요.\n"
            "  PowerShell: $env:SGIS_KEY=\"키\"; $env:SGIS_SECRET=\"시크릿\"; python sgis_admin_codes.py"
        )
    auth()
    rows = collect()
    build_xlsx(rows)
