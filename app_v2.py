# -*- coding: utf-8 -*-
"""
app_v2.py — 복합쇠퇴진단 자동화 앱 (Streamlit) · UX 리디자인 버전
==================================================================
기존 app.py 는 그대로 두고, 검증된 엔진 모듈(decline_engine / legal_engine /
export / loader / golden_io / custom_indicators / sheet_builder)을 그대로 재사용하되
화면 구성만 "단계형 워크스페이스"로 다시 설계한 버전.

흐름:  ① 자료 신청 → ② 데이터 입력 → ③ 설정(가중치·등급) → ④ 시트 검토 → ⑤ 진단 산출
- 한 번에 한 단계만 보여줘 초보도 길을 잃지 않음(상단 스테퍼로 자유 이동).
- 무거운 산출물(원시 데이터·결과)은 session_state 에 보관해 단계 이동에도 유지.
- 결과 화면은 요약 타일 + 행정동 순위 차트 + 표 + 다운로드로 강화.

실행:  streamlit run app_v2.py
"""
import io
import os

import pandas as pd
import streamlit as st
# altair·openpyxl은 결과 화면(step4)에서만 필요 → 지연 import로 초기 로딩 단축.

import code_audit as CA
import config as C
import custom_indicators as CI
import decline_engine as E
import dong_names
import export
import golden_io
import legal_engine as LG
import loader as L
import recipe_engine as RE
import sgis_collect as SC
import sgis_request as SR
import sheet_builder
import template_export as TE

# ──────────────────────────────────────────────────────────────────────────
# 페이지 · 스타일
# ──────────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="쇠퇴진단 자동화 시스템", page_icon="🧭", layout="wide")

st.markdown(
    """
    <style>
    :root { --brand:#2E7D6B; --brand-d:#1F5A4C; --ink:#17202A; --muted:#65727E; --line:#E1E7EF; }
    .stApp { background:#F5F7FA; }
    .block-container { padding-top:1.2rem; padding-bottom:4rem; max-width:1320px; }
    [data-testid="stSidebar"] { background:#FFFFFF; border-right:1px solid var(--line); }

    /* 헤더 */
    .app-hero { display:flex; align-items:center; gap:14px; margin-bottom:6px; }
    .hero-badge { width:42px; height:42px; border-radius:11px; flex:0 0 42px;
        background:linear-gradient(135deg,var(--brand) 0%,var(--brand-d) 100%);
        color:#fff; font-size:1.4rem; display:flex; align-items:center; justify-content:center; }
    .hero-kicker { color:var(--brand); font-size:.72rem; font-weight:800; letter-spacing:.12em; }
    .hero-title { color:var(--ink); font-size:1.5rem; font-weight:800; line-height:1.1; }
    .hero-sub { color:var(--muted); font-size:.86rem; margin-top:1px; }

    /* 스텝 진행바 (버튼 위에 얹는 라인) */
    .step-track { height:4px; background:var(--line); border-radius:2px; margin:2px 0 14px 0; position:relative; }
    .step-fill { height:4px; background:var(--brand); border-radius:2px; transition:width .25s; }

    /* 스텝 네비 버튼 */
    div[data-testid="column"] div.stButton>button { border-radius:9px; font-weight:700; padding:.55rem .2rem; }
    div.stButton>button[kind="primary"] { background:var(--brand); border-color:var(--brand); }
    div.stButton>button[kind="primary"]:hover { background:var(--brand-d); border-color:var(--brand-d); }

    /* 섹션 */
    .sec-title { color:var(--ink); font-size:1.12rem; font-weight:800; margin:14px 0 2px; }
    .sec-sub { color:var(--muted); font-size:.86rem; margin-bottom:10px; }

    /* 타일/카드 */
    .tile { border:1px solid var(--line); border-radius:12px; background:#fff; padding:14px 16px;
        box-shadow:0 1px 2px rgba(16,24,40,.04); min-height:84px; }
    .tile-accent { border-left:4px solid var(--brand); }
    .tile-danger { border-left:4px solid #C0392B; }
    .tile-label { color:var(--muted); font-size:.76rem; font-weight:700; margin-bottom:3px; }
    .tile-value { color:var(--ink); font-size:1.5rem; font-weight:800; line-height:1.1; }
    .tile-value.danger { color:#C0392B; }
    .tile-note { color:#8A95A1; font-size:.74rem; margin-top:3px; }

    .panel { border:1px solid var(--line); border-radius:12px; background:#fff; padding:18px 20px; }

    /* 사이드바 상태 */
    .side-h { color:var(--ink); font-size:.95rem; font-weight:800; margin:4px 0 8px; }
    .side-row { display:flex; justify-content:space-between; font-size:.82rem; padding:5px 0;
        border-bottom:1px dashed var(--line); }
    .side-k { color:var(--muted); } .side-v { color:var(--ink); font-weight:700; }
    .pill-ok { color:#1E7D5A; font-weight:800; } .pill-no { color:#B0B7C0; font-weight:700; }

    div.stDownloadButton>button { border-radius:9px; font-weight:800; }

    /* Streamlit 기본 UI 숨김(우측상단 메뉴·Deploy·하단 푸터·상태위젯) */
    #MainMenu { visibility:hidden; }
    [data-testid="stToolbar"] { visibility:hidden; height:0; }
    [data-testid="stStatusWidget"] { visibility:hidden; }
    [data-testid="stDecoration"] { display:none; }
    header[data-testid="stHeader"] { background:transparent; }
    footer { visibility:hidden; height:0; }
    </style>
    """,
    unsafe_allow_html=True,
)


def sec(title, sub=""):
    st.markdown(f'<div class="sec-title">{title}</div>', unsafe_allow_html=True)
    if sub:
        st.markdown(f'<div class="sec-sub">{sub}</div>', unsafe_allow_html=True)


def tile(label, value, note="", variant="accent", danger_value=False):
    cls = "tile-danger" if variant == "danger" else "tile-accent"
    vcls = "tile-value danger" if danger_value else "tile-value"
    st.markdown(
        f'<div class="tile {cls}"><div class="tile-label">{label}</div>'
        f'<div class="{vcls}">{value}</div><div class="tile-note">{note}</div></div>',
        unsafe_allow_html=True,
    )


# ──────────────────────────────────────────────────────────────────────────
# 파싱/로딩 유틸 (기존 app.py 와 동일 — 캐시 포함)
# ──────────────────────────────────────────────────────────────────────────
@st.cache_data(show_spinner="원시 데이터 추출 중…")
def extract_raw(file_bytes: bytes, mapping_items):
    mapping = dict(mapping_items) if mapping_items else None
    import openpyxl          # 지연 import(파일 로딩 시에만 필요)
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        miss = golden_io.missing_sheets(wb)
        if miss:
            return None, miss
        return golden_io.load_raw_from_workbook(wb, mapping), []
    finally:
        wb.close()


@st.cache_data(show_spinner=False)
def cached_load_folders(folders_tuple, mapping_items):
    mapping = dict(mapping_items) if mapping_items else None
    return L.load_raw_from_folders(list(folders_tuple), mapping=mapping)


@st.cache_data(show_spinner=False)
def cached_summarize_folders(folders_tuple):
    return L.summarize_folders(list(folders_tuple)) if hasattr(L, "summarize_folders") else pd.DataFrame()


class _NamedBytes:
    """업로드 파일을 (이름, bytes)로 감싸 loader가 그대로 먹게 함(캐시 키 안정용)."""
    def __init__(self, name, data):
        self.name = name
        self._d = data

    def getvalue(self):
        return self._d


@st.cache_data(show_spinner="원시 파일 정제 중…")
def cached_load_uploads(files_sig, mapping_items, column_map=None):
    """업로드 파일 (이름, bytes) 튜플 → raw dict. bytes 기준 캐시 → 재파싱 방지."""
    mapping = dict(mapping_items) if mapping_items else None
    cmap = dict(column_map) if isinstance(column_map, tuple) else column_map
    wrapped = [_NamedBytes(n, d) for n, d in files_sig]
    return L.load_raw_from_uploaded_files(wrapped, mapping=mapping, column_map=cmap)


def _files_sig(files):
    """업로드 목록 → ((이름, bytes), …) 해시 안정 튜플."""
    return tuple((getattr(f, "name", "f"), f.getvalue()) for f in (files or []))


def xlsx_bytes(df, sheet_name):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    buf.seek(0)
    return buf.getvalue()


def read_mapping_file(uploaded):
    name = uploaded.name.lower()
    if name.endswith(".xlsx"):
        return pd.read_excel(uploaded, dtype=str).fillna("")
    for enc in ("utf-8-sig", "cp949", "euc-kr", "utf-8"):
        try:
            uploaded.seek(0)
            return pd.read_csv(uploaded, dtype=str, encoding=enc).fillna("")
        except Exception:
            continue
    raise ValueError("매핑 파일을 읽지 못했습니다. xlsx 형식을 권장합니다.")


def parse_dong_mapping(mdf):
    """매핑표 DataFrame → (mapping_items, name_map) 자동 인식(어느 지역이든).
    지원: [행정동코드, 행정동명] 2열 / [집계구코드, 행정동코드, 행정동명] 3열 / [집계구코드, 행정동명] 등.
    컬럼명(집계구·행정동·명) 또는 값(14자리/8자리/한글)으로 추정."""
    cols = list(mdf.columns)

    def col_str(c):
        return mdf[c].astype(str).str.strip()

    def is_code(c, n):
        s = col_str(c)
        return len(s) and (s.str.fullmatch(r"\d{%d}" % n).mean() > 0.5)

    def is_hangul(c):
        s = col_str(c)
        return len(s) and (s.str.contains(r"[가-힣]").mean() > 0.5)

    jgu = next((c for c in cols if "집계구" in str(c)), None) or next((c for c in cols if is_code(c, 14)), None)
    dong = (next((c for c in cols if ("행정동" in str(c) or "읍면동" in str(c)) and "코드" in str(c)), None)
            or next((c for c in cols if is_code(c, 8) and c != jgu), None))
    name = (next((c for c in cols if any(k in str(c) for k in ("행정동명", "동명", "명칭", "이름")) and "코드" not in str(c)), None)
            or next((c for c in cols if is_hangul(c) and c not in (jgu, dong)), None))

    mapping_items, name_map = None, None
    if jgu and dong:
        mapping_items = tuple(zip(col_str(jgu), col_str(dong)))
    if dong and name:
        name_map = dict(zip(col_str(dong), col_str(name)))
    elif jgu and name and not dong:                      # 집계구+이름만 → 앞8자리로
        name_map = {g[:8]: nm for g, nm in zip(col_str(jgu), col_str(name))}
    return mapping_items, name_map


def read_ref_code_file(uploaded):
    name = uploaded.name.lower()
    if name.endswith(".xlsx") or name.endswith(".xls"):
        df = pd.read_excel(uploaded, dtype=str).fillna("")
    else:
        df = None
        for enc in ("utf-8-sig", "cp949", "euc-kr", "utf-8"):
            try:
                uploaded.seek(0)
                df = pd.read_csv(uploaded, dtype=str, encoding=enc).fillna("")
                break
            except Exception:
                continue
        if df is None:
            raise ValueError("참조 코드표를 읽지 못했습니다.")
    cols = list(df.columns)
    code_col = next((c for c in cols if "CODE" in str(c).upper() or "코드" in str(c)), cols[0])
    name_col = next((c for c in cols if c != code_col and ("명" in str(c) or "항목" in str(c) or "name" in str(c).lower())), cols[1] if len(cols) > 1 else cols[0])
    return dict(zip(df[code_col].astype(str).str.strip(), df[name_col].astype(str).str.strip()))


def bucket_years(raw, keys):
    years = set()
    for key in keys:
        if key in raw and len(raw[key]):
            years.update(pd.to_numeric(raw[key]["연도"], errors="coerce").dropna().astype(int).tolist())
    return sorted(years)


BUCKET_GUIDE = pd.DataFrame([
    {"분류": "to_in", "용도": "법적인구/인구변화율", "필요 CODE": "to_in_001", "파일 예": "총인구"},
    {"분류": "in_age", "용도": "복합 인문사회", "필요 CODE": "in_age_001~021, 065~068", "파일 예": "성연령별인구"},
    {"분류": "to_fa", "용도": "법적산업/총사업체수", "필요 CODE": "to_fa_010", "파일 예": "총괄사업체수"},
    {"분류": "cp_bem", "용도": "복합 산업경제", "필요 CODE": "cp_bem_*", "파일 예": "산업분류별 종사자수"},
    {"분류": "ho_yr", "용도": "법적물리/노후건축물", "필요 CODE": "ho_yr_*", "파일 예": "건축년도별주택"},
    {"분류": "ho_ar", "용도": "복합 물리환경/소형주택", "필요 CODE": "ho_ar_*", "파일 예": "연건평별주택"},
])
REQUIRED_FULL = ["to_in", "in_age", "to_fa", "cp_bem", "ho_yr", "ho_ar"]
# 기본 11지표 최종 진단에 실제로 필요한 분류(ho_ar=연면적은 소형주택비율 제거로 '선택'임)
REQUIRED_CORE = ["to_in", "in_age", "to_fa", "cp_bem", "ho_yr"]
OPTIONAL_BUCKETS = ["ho_ar"]


def missing_bucket_message(missing):
    return BUCKET_GUIDE[BUCKET_GUIDE["분류"].isin(missing)].reset_index(drop=True)


# ──────────────────────────────────────────────────────────────────────────
# 세션 상태
# ──────────────────────────────────────────────────────────────────────────
ss = st.session_state
ss.setdefault("step", 1)
ss.setdefault("raw", None)              # 전체 원시 dict
ss.setdefault("selected_years", None)
ss.setdefault("pop_ref_year", None)
ss.setdefault("biz_ref_year", None)
@st.cache_data(show_spinner=False)
def _national_name_map():
    """전국 행정동코드(8자리)→행정동명 (행정구역코드_전국.xlsx에서 1회 캐시).
    파일이 없으면 전주 34동 기본 내장으로 폴백. → 어느 지역이든 이름 자동."""
    try:
        import template_export as _TE
        dong, _ = _TE.load_admin_names(_TE.DEFAULT_ADMIN_PATH)
        if dong:
            return {**dong_names.default_name_map(), **dong}
    except Exception:
        pass
    return dong_names.default_name_map()

ss.setdefault("name_map", _national_name_map())   # 전국 행정동 이름 자동(전국표)·전주 폴백
ss.setdefault("code_label_map", {})
ss.setdefault("sector_df", None)
ss.setdefault("internal_wmap", {})
ss.setdefault("active_map", None)       # 지표별 사용 여부 {지표ID: bool}
ss.setdefault("weights_ver", 0)         # 가중치 표 강제 리셋용 키 버전
ss.setdefault("custom_df", pd.DataFrame())
ss.setdefault("custom_meta", pd.DataFrame())
ss.setdefault("recipes", RE.sample_recipes())   # 학습용 샘플 3종 미리 로드(기본 사용해제)
ss.setdefault("clone_notice", None)     # 기본지표 복제 직후 안내(1회 표시)
ss.setdefault("pivot_level", "dong")    # 복합/피벗 집계 단위
ss.setdefault("cfg", {})                # 계산된 가중치/지표 맵
ss.setdefault("results", None)          # 최종 진단 결과
ss.setdefault("nationwide", False)       # 🚀 전국 디폴트 원스톱 전용 화면 여부

# 기본 off 지표(있으면 처음 진입 시 사용 해제). 소형주택비율은 아예 기본지표에서
# 제거됐으므로(2026-07-06) 현재 비어 있음.
DEFAULT_OFF = set(RE.SAMPLE_NAMES)   # 샘플 계산식은 기본 '사용 해제'(실제 진단은 11지표 그대로)


def goto(n):
    ss.step = n
    st.rerun()


def _sorted_order(comp, mode):
    """정렬 순서(인덱스 리스트). comp=composite('종합' 열 포함). '기본'이면 원래 순서."""
    idx = list(comp.index)
    if mode == "행정동코드 오름차순":
        return sorted(idx, key=str)
    if mode == "행정동코드 내림차순":
        return sorted(idx, key=str, reverse=True)
    if mode == "종합점수 높은순":
        return comp["종합"].sort_values(ascending=False).index.tolist()
    if mode == "종합점수 낮은순":
        return comp["종합"].sort_values(ascending=True).index.tolist()
    return idx


def equalize_weights(active_map, sector_of):
    """각 부문 안에서 '사용 중' 지표에 100을 똑같이 나눠준다(부문별 합 100).
    현재 값과 무관하게 균등분배 — '다 똑같이'가 필요할 때."""
    out = {ind: 0.0 for ind in active_map}
    for sec in C.SECTORS:
        inds = [i for i in active_map if active_map.get(i) and sector_of.get(i) == sec]
        if inds:
            v = round(100.0 / len(inds), 6)
            for i in inds:
                out[i] = v
    return out


def normalize_weights(active_map, wmap, sector_of):
    """각 부문 안에서 '사용 중' 지표의 내부가중치를 비율 유지하며 합 100으로 맞춘다.
    모두 0이면 균등 분배. 반환: {지표ID: 가중치} (사용 안 하는 지표는 0)."""
    out = {ind: 0.0 for ind in active_map}
    for sec in C.SECTORS:
        inds = [i for i in active_map if active_map.get(i) and sector_of.get(i) == sec]
        if not inds:
            continue
        total = sum(max(0.0, float(wmap.get(i, 0.0))) for i in inds)
        if total <= 0:
            for i in inds:
                out[i] = round(100.0 / len(inds), 6)
        else:
            for i in inds:
                out[i] = round(max(0.0, float(wmap.get(i, 0.0))) / total * 100.0, 6)
    return out


# ──────────────────────────────────────────────────────────────────────────
# 헤더 + 스테퍼
# ──────────────────────────────────────────────────────────────────────────
st.markdown(
    """
    <div class="app-hero">
      <div class="hero-badge">🧭</div>
      <div>
        <div class="hero-kicker">SGIS DECLINE ANALYZER</div>
        <div class="hero-title">쇠퇴진단 자동화 시스템</div>
        <div class="hero-sub">SGIS 원시자료 → 법적쇠퇴진단 · 복합쇠퇴지수 · 검토용 Excel</div>
      </div>
    </div>
    """,
    unsafe_allow_html=True,
)

STEPS = [("①", "자료 신청"), ("②", "데이터 입력"), ("③", "설정"), ("④", "시트 검토"), ("⑤", "산출")]
has_data = ss.raw is not None
if not ss.get("nationwide"):     # 전국 원스톱 화면에서는 단계 스테퍼 숨김
    st.markdown(
        f'<div class="step-track"><div class="step-fill" style="width:{(ss.step-1)/4*100:.0f}%"></div></div>',
        unsafe_allow_html=True,
    )
    nav_cols = st.columns(5)
    for i, (col, (num, name)) in enumerate(zip(nav_cols, STEPS), start=1):
        done = "✓ " if (i < ss.step) else ""
        disabled = (i > 2 and not has_data)   # ①자료신청·②데이터입력은 항상 열림, ③~은 데이터 필요
        if col.button(f"{done}{num} {name}", key=f"nav_{i}",
                      type="primary" if ss.step == i else "secondary",
                      use_container_width=True, disabled=disabled):
            goto(i)

# 사이드바: 진행 요약(설정 아님 — 설정은 ②단계 본문)
with st.sidebar:
    st.markdown('<div class="side-h">📋 진행 상태</div>', unsafe_allow_html=True)
    data_pill = '<span class="pill-ok">불러옴</span>' if has_data else '<span class="pill-no">대기</span>'
    yr = ss.selected_years
    yr_txt = f"{min(yr)}–{max(yr)}" if yr else "—"
    res_pill = '<span class="pill-ok">완료</span>' if ss.results else '<span class="pill-no">미산출</span>'
    st.markdown(
        f'<div class="side-row"><span class="side-k">데이터</span><span class="side-v">{data_pill}</span></div>'
        f'<div class="side-row"><span class="side-k">선택 연도</span><span class="side-v">{yr_txt}</span></div>'
        f'<div class="side-row"><span class="side-k">진단 산출</span><span class="side-v">{res_pill}</span></div>',
        unsafe_allow_html=True,
    )
    if ss.results:
        st.markdown(
            f'<div class="side-row"><span class="side-k">법적 쇠퇴지역</span>'
            f'<span class="side-v">{ss.results["n_decl"]}개 행정동</span></div>',
            unsafe_allow_html=True,
        )
    st.markdown("---")
    st.caption("설정(가중치·등급·추가지표)은 **③ 설정** 단계 본문에서 넓게 편집합니다.")
    if st.button("↺ 처음부터 다시", use_container_width=True):
        for k in ("raw", "selected_years", "results", "cfg"):
            ss[k] = None if k != "cfg" else {}
        ss.nationwide = False
        goto(2)
    st.markdown("---")
    if not ss.get("nationwide"):
        if st.button("🚀 전국 디폴트 원스톱", use_container_width=True, type="primary",
                     help="전국 시군구를 디폴트로 처음부터 끝까지 자동 구축하는 전용 화면(신청→자동 다운로드→229개 빌드)."):
            ss.nationwide = True
            st.rerun()
    else:
        if st.button("← 일반 단계별 분석", use_container_width=True):
            ss.nationwide = False
            st.rerun()


def _merge_raw(primary, secondary):
    """primary(API 시계열) 우선, 비어있는 버킷만 secondary(파일 보완)로 채움."""
    cols = ["연도", "집계구", "CODE", "값", "행정동코드"]
    out = {}
    for b in ("to_in", "in_age", "to_fa", "cp_bem", "ho_yr", "ho_ar"):
        p = primary.get(b) if primary else None
        s = secondary.get(b) if secondary else None
        if p is not None and len(p):
            out[b] = p
        elif s is not None and len(s):
            out[b] = s
        else:
            out[b] = pd.DataFrame(columns=cols)
    return out


def _apply_build_items(checked):
    """선택 항목 + 부문별 선택 연도 → 신청 item 목록((도메인,코드,연도,이름))."""
    dom_years = {dk: {int(y) for y in ss.get(f"apply_years_{dk}", SR.YEAR_DOMAINS[dk][2])}
                 for dk in SR.YEAR_DOMAINS}
    items = []
    for name in checked:
        dk = SR.YEAR_DOMAIN.get(name)
        for (d, c, y) in SR.ITEM_CATALOG[name]:
            if name in SR.SNAPSHOT_ITEMS or dk is None:
                items.append((d, c, y, name))
            elif int(y) in dom_years[dk]:
                items.append((d, c, y, name))
    return items


def _apply_build_applicant(ap_userid, ap_company, ap_email, ap_tel, ap_goal):
    """신청자 입력 → SGIS saveRequestData 신청서 필드 dict."""
    eid, edom = (ap_email.split("@", 1) + [""])[:2] if ap_email else ("", "naver.com")
    tel = [t for t in ap_tel.replace(" ", "").split("-") if t] if ap_tel else []
    return {
        "param_userkey": ap_userid, "sgis_census_req_company": ap_company,
        "email_id": eid, "email_addr": edom or "naver.com", "email_addr_select": edom or "naver.com",
        "sgis_census_req_tel_1": tel[0] if len(tel) > 0 else "",
        "sgis_census_req_tel_2": tel[1] if len(tel) > 1 else "",
        "sgis_census_req_tel_3": tel[2] if len(tel) > 2 else "",
        "sgis_census_req_goal": ap_goal or "복합쇠퇴진단",
    }


def _apply_submit_batch(cookie, sgcodes, items, applicant, only_first, progress=None):
    """시군구별로 나눠 신청(서버 부하·타임아웃 예방·지역별 실패 격리). results 반환."""
    results = []
    total = 1 if only_first else len(sgcodes)
    for i, sg in enumerate(sgcodes):
        cart = SR.make_cart(items, [sg], only_first=only_first)
        try:
            stt, resp = SR.submit_cart(cookie, cart, applicant=applicant)
        except Exception as e:
            stt, resp = -1, str(e)
        results.append((sg, stt, resp, len(cart)))
        if progress:
            progress(i + 1, total, sg)
        if only_first:
            break
    return results


def _apply_render_results(results):
    """신청 결과 판정·표시. SGIS saveRequestData 응답: **1=신청됨 · 2=저장됨 → 둘 다 성공**
    (SGIS 페이지 JS가 1·2 모두 '신청내역으로 이동'하는 성공 처리). 그 외=오류. 성공목록 반환."""
    ok = [r for r in results if r[1] == 200 and str(r[2]).strip() in ("1", "2")]
    err = [r for r in results if r not in ok]
    if any("로그인" in str(r[2]) for r in results):
        st.error("쿠키가 만료되었거나 유효하지 않습니다. SGIS에 다시 로그인한 뒤 쿠키를 새로 복사해 붙여넣으십시오.")
    elif ok and not err:
        st.success(f"신청이 접수되었습니다 — 시군구 {len(ok)}곳(응답 1=신청 / 2=저장, 둘 다 성공). "
                   "SGIS 승인 후 **② 데이터 입력의 ‘☁ 승인 자료 자동 다운로드’** 에서 바로 받으십시오.")
    elif ok:
        st.warning(f"일부만 접수되었습니다 — 성공 {len(ok)}곳 / 실패 {len(err)}곳. 아래 결과를 확인하십시오.")
    else:
        bad = err[0] if err else None
        bmsg = str(bad[2]) if bad else ""
        if "time" in bmsg.lower():
            st.error("⏱ 신청 시간초과 — **쿠키 문제가 아닙니다.** 해외 서버(streamlit.app)는 SGIS 정부 서버 접속이 차단됩니다. "
                     "**본인 PC 또는 국내 서버(Cloudtype)에서 실행**해 신청하십시오.")
        else:
            st.error(f"신청 실패 — 서버 응답이 없거나 오류. {bmsg[:200]}")
    with st.expander(f"신청 결과 {len(results)}건", expanded=bool(err)):
        for sg, stt, resp, n in results:
            rs = str(resp).strip()
            mark = "✅" if (stt == 200 and rs in ("1", "2")) else "❌"
            st.caption(f"{mark} {sg} — HTTP {stt}, 항목 {n}건. 응답:{rs[:40]}")
    return ok


def _sgis_apply_block():
    """🏛 SGIS 집계구 자료신청 — 쿠키+지역+항목 선택 → saveRequestData 신청(제출).
    데이터는 승인(약 10분)·다운로드 후 '원시 SGIS CSV 폴더 경로'로 불러온다."""
    st.caption("집계구 자료를 **신청(제출)까지 자동화**합니다. 항목을 일일이 담을 필요 없이 지역과 항목만 선택하면 됩니다. "
               "승인(약 10분)·다운로드 후 받은 폴더를 **② 데이터 입력**의 **‘원시 SGIS CSV 폴더 경로’** 모드로 불러오십시오.")
    with st.expander("❓ 복합쇠퇴 전체 분석에 어떤 데이터가 얼마나 필요한가 (꼭 확인)", expanded=False):
        st.markdown(SR.NEED_EXPLAIN)

    with st.expander("① SGIS 로그인 쿠키 붙여넣기", expanded=not ss.get("apply_cookie")):
        st.markdown("**로그인한 요청 하나를 통째로 복사해 붙여넣으십시오.** "
                    "반드시 **sgis.mods.go.kr 로 가는 요청**을 복사해야 신청에 필요한 값이 모두 포함됩니다.")
        st.caption("① SGIS **로그인** → **자료신청 페이지**(sgis.mods.go.kr) 열기 → ② **F12** → **Network** 탭 → "
                   "③ **F5(새로고침)** → ④ 목록 **맨 위 요청**(이름 `requestData` = 주소창과 같은 **sgis.mods.go.kr** 요청) "
                   "**우클릭 → Copy → Copy as cURL** → ⑤ 아래에 **통째로 붙여넣기**. "
                   "⚠️ 광고·지도·통계 등 **다른 사이트 요청**을 복사하면 값이 누락됩니다 — 반드시 **sgis.mods.go.kr** 요청으로 복사하십시오.")
        cookie_raw = st.text_area("여기에 통째로 붙여넣기 (Copy as cURL)", value=ss.get("apply_curl_raw", ""),
                                  height=90, key="in_apply_cookie", label_visibility="collapsed")
        with st.expander("잘 안 되면 — 값 2개만 직접 넣기"):
            st.caption("F12 → **Application**(크롬)/**저장소**(Firefox) → 왼쪽 **Cookies → sgis.mods.go.kr** → "
                       "`JSESSIONID`·`accessToken` 두 줄의 **Value**를 복사해 아래 칸에 입력하십시오.")
            mk1, mk2 = st.columns(2)
            js = mk1.text_input("JSESSIONID", value=ss.get("ck_js", ""), key="in_ck_js")
            at = mk2.text_input("accessToken", value=ss.get("ck_at", ""), key="in_ck_at")
        # 우선순위: 직접입력 2개 > cURL 통째 → 최종 쿠키를 ss.apply_cookie(다운스트림 공용)에 저장
        if js.strip() and at.strip():
            ss.ck_js, ss.ck_at = js.strip(), at.strip()
            ss.apply_cookie = f"JSESSIONID={js.strip()}; accessToken={at.strip()}"
        elif cookie_raw.strip():
            ss.apply_curl_raw = cookie_raw
            ss.apply_cookie = cookie_raw
        # 실시간 개별 확인 — 맞는 요청을 복사했는지 바로 알려줌
        found = SR.found_cookies(ss.get("apply_cookie", ""))
        s1, s2 = st.columns(2)
        s1.markdown("✅ **JSESSIONID** 확인" if found["JSESSIONID"] else "❌ JSESSIONID 없음")
        s2.markdown("✅ **accessToken** 확인" if found["accessToken"] else "❌ accessToken 없음")
        if ss.get("apply_cookie") and all(found.values()):
            st.success("준비 완료 — 아래에서 지역·항목을 선택해 신청하십시오.", icon="✅")
        elif ss.get("apply_cookie"):
            miss = ", ".join(k for k, v in found.items() if not v)
            st.warning(f"**{miss}** 이(가) 포함되지 않았습니다 — 다른 사이트 요청을 복사한 경우입니다. "
                       "목록에서 **sgis.mods.go.kr** 로 가는 요청(맨 위 `requestData`)을 다시 복사해 붙여넣으십시오.")
        cookie_raw = ss.get("apply_cookie", "")   # 이후 코드가 쓰는 변수 = 최종 확정 쿠키

    with st.expander("② 신청자 정보 (SGIS 신청서에 들어감 — 본인 정보 입력)", expanded=not ss.get("apply_userid")):
        st.caption("SGIS 자료신청서에 채워지는 값입니다. 코드에 저장되지 않으며 현재 세션에서만 사용됩니다.")
        ac1, ac2 = st.columns(2)
        ap_userid = ac1.text_input("SGIS 로그인 아이디", value=ss.get("apply_userid", ""), key="in_apply_userid")
        ap_company = ac2.text_input("소속/회사명", value=ss.get("apply_company", ""), key="in_apply_company")
        ap_email = ac1.text_input("이메일(승인 알림 받을 주소)", value=ss.get("apply_email", ""), key="in_apply_email",
                                  placeholder="hong@naver.com")
        ap_tel = ac2.text_input("연락처", value=ss.get("apply_tel", ""), key="in_apply_tel", placeholder="010-1234-5678")
        ap_goal = st.text_input("활용 목적/과제명", value=ss.get("apply_goal", "복합쇠퇴진단"), key="in_apply_goal")
        # 화면 전환(전국 원스톱 ↔ 일반)에도 신청자 정보가 유지되도록 매 렌더 세션에 저장
        # (기존엔 신청 버튼을 눌러야만 _persist_apply로 저장돼, 버튼 없이 화면을 옮기면 값이 사라졌음)
        ss.apply_userid, ss.apply_company = ap_userid, ap_company
        ss.apply_email, ss.apply_tel, ss.apply_goal = ap_email, ap_tel, ap_goal

    # ── ③ 지역 선택 (시도 고르면 그 시도의 시군구를 불러옴) ──────────────────
    st.markdown("**③ 지역 선택**")
    st.caption("시도를 선택하면 해당 시도의 시군구가 자동으로 표시됩니다(쿠키 필요). 여러 시군구를 클릭해 선택하십시오. "
               "표시되지 않으면 아래 칸에 코드를 직접 입력해도 됩니다.")
    ss.setdefault("sgg_cache", {})
    sido_opts = SR.SIDO_LIST              # 고정 17개 — 네트워크 불필요, 항상 표시
    sido_map = dict(sido_opts)
    sc1, sc2 = st.columns([1, 2])
    sido_code = sc1.selectbox("시도", [c for c, _ in sido_opts],
                              format_func=lambda c: sido_map.get(c, c), key="apply_sido")
    # 선택한 시도의 시군구만 1회 로드(캐시)
    cache = ss.sgg_cache
    if sido_code and sido_code not in cache:
        ck = SR.extract_cookie(cookie_raw or "")
        if not ck:
            sc2.info("① 쿠키를 붙여넣으면 시군구 목록이 표시됩니다.")
        else:
            try:
                with st.spinner(f"{sido_map.get(sido_code)} 시군구 불러오는 중…"):
                    cache[sido_code] = SR.fetch_sigungu_list(ck, sido_code)
            except Exception as e:
                sc2.error(f"시군구 불러오기 실패: {e} — 쿠키가 만료되었으면 SGIS에 다시 로그인한 뒤 새 쿠키를 붙여넣으십시오.")
    sgg_opts = cache.get(sido_code, [])
    sgg_map = dict(sgg_opts)
    all_codes = [c for c, _ in sgg_opts]
    sel = sc2.multiselect("시군구 (여러 개 가능)", all_codes,
                          format_func=lambda c: f"{sgg_map.get(c, c)} · {c}", key="apply_sgg_pick")
    if sgg_opts:
        # 위젯 key를 직접 바꾸면 StreamlitAPIException → on_click 콜백에서 세팅(위젯 생성 전 실행돼 허용됨).
        def _pick_all(codes=all_codes):
            ss.apply_sgg_pick = list(codes)

        def _pick_none():
            ss.apply_sgg_pick = []

        ba1, ba2 = sc2.columns(2)
        ba1.button(f"＋ 이 시도 전체 선택 ({len(sgg_opts)}곳)", use_container_width=True,
                   help="이 시도의 모든 시군구를 한 번에 선택합니다.", on_click=_pick_all)
        ba2.button("－ 선택 비우기", use_container_width=True, on_click=_pick_none)
    picked_sgg = [(c, sgg_map.get(c, c)) for c in sel]
    if picked_sgg:
        st.caption("선택: " + ", ".join(f"{n}({c})" for c, n in picked_sgg))

    region = st.text_input("또는 시군구코드 직접 입력(쉼표)", value=ss.get("apply_region", ""),
                           key="in_apply_region", placeholder="예: 35011,35012",
                           help="목록을 안 쓰거나 다른 시도 코드를 섞을 때. 위 선택과 합쳐집니다.")
    st.caption("📍 **시군구코드 = 그 지역 '집계구'(14자리) 데이터**를 받아요 — 앱이 집계구→행정동으로 자동 묶습니다.")

    # ── ④ 수집 연도: 부문별로 개별 선택·삭제 ──────────────────────────────
    st.markdown("**④ 수집 연도 (부문별로 개별 선택·삭제)**")
    st.caption("증감률 지표(총인구·사업체·종사자)는 기본 전 연도. **성연령·건축연도는 기본 최신(2024)** 이고, "
               "다른 해로 비교하려면(예: 2023) 그 해를 추가하십시오. (엔진은 분석 기준연도 1개만 사용)")
    dom_keys = list(SR.YEAR_DOMAINS.keys())
    ycols = st.columns(2)
    for i, dk in enumerate(dom_keys):
        label, years, default_years, help_txt = SR.YEAR_DOMAINS[dk]
        ss.setdefault(f"apply_years_{dk}", list(default_years))
        ycols[i % 2].multiselect(label, years, key=f"apply_years_{dk}", help=help_txt)

    st.markdown("**⑤ 받을 항목**")
    st.caption("✅ 아래 **필수 5종 = 복합쇠퇴 전체 세트.** 이 항목을 모두 받으면 전체 결과표가 산출됩니다(기본 선택됨).")
    checked = []
    req_names = [n for n in SR.ITEM_CATALOG if SR.ITEM_META[n][0] == "필수"]
    opt_names = [n for n in SR.ITEM_CATALOG if SR.ITEM_META[n][0] == "선택"]
    for name in req_names:
        if st.checkbox(f"**{name}**", value=(name in SR.DEFAULT_CHECKED), key=f"apply_it_{name}"):
            checked.append(name)
        st.caption("　└ " + SR.ITEM_META[name][1])
    with st.expander("➕ 선택 항목 (새 지표 만들 때만 — 기본 분석엔 불필요)", expanded=False):
        for name in opt_names:
            if st.checkbox(name, value=(name in SR.DEFAULT_CHECKED), key=f"apply_it_{name}"):
                checked.append(name)
            st.caption("　└ " + SR.ITEM_META[name][1])

    test_first = st.checkbox("먼저 1건만 시험 신청(작동 확인용)", value=True, key="apply_test_first",
                             help="처음에는 켜서 1건만 신청한 뒤 신청내역을 확인하고, 끄고 전체 신청하는 것을 권장합니다.")

    def _persist_apply():
        ss.apply_cookie, ss.apply_region = cookie_raw, region
        ss.apply_userid, ss.apply_company = ap_userid, ap_company
        ss.apply_email, ss.apply_tel, ss.apply_goal = ap_email, ap_tel, ap_goal

    ab_sel, ab_all = st.columns([1, 1])
    if ab_sel.button("🏛 선택 시군구 신청", type="primary", use_container_width=True):
        cookie = SR.extract_cookie(cookie_raw or "")
        # 지역: 목록에서 고른 것 + 직접 입력을 합침(중복 제거·순서 유지)
        sgcodes = [c for c, _ in picked_sgg] + [s.strip() for s in region.split(",") if s.strip()]
        sgcodes = list(dict.fromkeys(sgcodes))
        # 목록을 불러왔으면 알려진 코드인지 검증 — 오타·미제공 코드는 전송 차단(타임아웃 예방)
        known = set()
        for lst in ss.get("sgg_cache", {}).values():
            known |= {c for c, _ in lst}
        loaded_sido = set(ss.get("sgg_cache", {}).keys())     # 시군구를 불러온 시도들
        # 그 시도를 불러온 적 있는데 목록에 없는 코드만 오타로 간주(안 불러온 시도는 검증 skip)
        unknown = [c for c in sgcodes if c[:2] in loaded_sido and c not in known]
        items = _apply_build_items(checked)
        if not cookie:
            st.error("쿠키를 붙여넣으십시오(JSESSIONID·accessToken이 확인되지 않으면 로그인·복사를 다시 하십시오).")
        elif not sgcodes:
            st.error("시군구를 1개 이상 선택하거나 코드를 입력하십시오.")
        elif unknown:
            st.error(f"목록에 없는 시군구코드: {', '.join(unknown)} — 오타이거나 SGIS가 제공하지 않는 지역입니다. "
                     "위 목록에서 선택하는 것이 안전합니다(잘못된 코드는 신청이 시간초과됩니다).")
        elif not items:
            st.error("받을 항목·연도를 1개 이상 선택하십시오.")
        elif not ap_email.strip():
            st.error("② 신청자 정보의 **이메일**을 입력하십시오 — 비어 있으면 SGIS가 신청을 거부합니다(응답 2).")
        else:
            _persist_apply()
            applicant = _apply_build_applicant(ap_userid, ap_company, ap_email, ap_tel, ap_goal)
            with st.spinner(f"{len(sgcodes)}개 시군구 신청 전송 중…"):
                results = _apply_submit_batch(cookie, sgcodes, items, applicant, only_first=test_first)
            ok = _apply_render_results(results)
            if test_first and ok:
                st.info("1건 시험 신청 성공 → 위 ‘먼저 1건만 시험 신청’ 체크를 끄고 다시 눌러 전체 신청하십시오.")

    # 전국 전체 신청·다운로드·빌드는 사이드바 '🚀 전국 디폴트 원스톱' 화면으로 일원화함(중복 제거).
    if ab_all.button("🌏 전국 전체는 원스톱에서 →", use_container_width=True,
                     help="전국 신청·자동 다운로드·전국 빌드는 '전국 디폴트 원스톱' 전용 화면에서 한 번에 처리합니다."):
        ss.nationwide = True
        st.rerun()
    st.caption("※ **전국 전체**(신청→다운로드→빌드)는 사이드바 **🚀 전국 디폴트 원스톱**에서 청크 처리로 진행하십시오.")


def _sgis_input_block(mapping_items):
    """🌐 SGIS API 직접받기 — 시계열(총인구·총사업체·산업 종사자수) 자동수집.
    ※ SGIS OpenAPI는 집계구(14자리)를 제공하지 않음 → 최소 단위가 읍면동(=행정동 8자리).
      즉 API가 이미 '집계구→행정동 합산'까지 끝낸 값을 준다(병합 단계 불필요).
      성연령·건축연도(노후건축물)는 API 정밀도 문제로 아래 보완 파일에서 채운다(하이브리드)."""
    st.caption("연도범위·지역을 정하면 SGIS OpenAPI로 **시계열**(총인구·총사업체·산업 종사자수)을 자동수집합니다. "
               "API는 **읍면동(=행정동) 단위**까지만 제공(집계구 없음) — 어차피 진단은 행정동 기준이라 그대로 사용됩니다.")
    with st.expander("SGIS 인증키 (consumer_key / secret)", expanded=not ss.get("sgis_key")):
        ck = st.text_input("서비스 ID (consumer_key)", value=ss.get("sgis_key", ""), type="password", key="in_sgis_key")
        cs = st.text_input("보안 Key (consumer_secret)", value=ss.get("sgis_secret", ""), type="password", key="in_sgis_secret")
        st.caption("발급: sgis.kostat.go.kr/developer → 인증키. 키는 이 세션에만 보관(파일 저장 안 함).")
    c1, c2, c3 = st.columns(3)
    yf = c1.number_input("시작 연도", 2000, 2024, int(ss.get("sgis_yf", 2000)), 1, key="in_sgis_yf")
    yt = c2.number_input("끝 연도", 2000, 2024, int(ss.get("sgis_yt", 2024)), 1, key="in_sgis_yt")
    region = c3.text_input("지역 시군구코드(쉼표)", value=ss.get("sgis_region", "35011,35012"), key="in_sgis_region",
                           help="전주=완산구 35011·덕진구 35012. 시군구 5자리 코드를 쉼표로 구분.")
    if st.button("🌐 SGIS에서 수집", type="primary"):
        sgcodes = [s.strip() for s in region.split(",") if s.strip()]
        if not ck or not cs:
            st.error("인증키(consumer_key/secret)를 입력하십시오.")
        elif not sgcodes:
            st.error("시군구코드를 1개 이상 입력하십시오.")
        else:
            ss.sgis_key, ss.sgis_secret = ck, cs
            ss.sgis_yf, ss.sgis_yt, ss.sgis_region = int(yf), int(yt), region
            years = list(range(int(yf), int(yt) + 1))
            total = max(len(years) * len(sgcodes), 1)
            prog = st.progress(0.0, text="수집 준비…")
            done = {"n": 0}

            def _cb(msg):
                if "수집 완료" in str(msg):
                    done["n"] += 1
                    prog.progress(min(done["n"] / total, 1.0), text=str(msg))

            try:
                token = SC.authenticate(ck, cs)
                ss.sgis_raw = SC.collect_raw(token, sgcodes, years, progress=_cb)
                prog.progress(1.0, text="완료")
                n = sum(len(v) for v in ss.sgis_raw.values())
                st.success(f"SGIS 수집 완료 — 총 {n:,}행 (to_in·to_fa·cp_bem)")
            except Exception as e:
                st.error(f"SGIS 수집 오류: {e}")

    api_raw = ss.get("sgis_raw")
    if api_raw:
        got = {b: len(v) for b, v in api_raw.items() if len(v)}
        st.info("API 수집됨: " + ", ".join(f"{b}({n}행)" for b, n in got.items())
                + "  ·  성연령·건축연도는 아래 보완 파일 필요")

    with st.expander("보완 파일 업로드 (성연령·건축연도 등 — API 미제공분)", expanded=False):
        st.caption("원시 SGIS CSV/TXT(연도,집계구,CODE,값). **집계구(14자리)면 행정동으로 자동 롤업**되어 API 데이터와 합쳐집니다.")
        supp = st.file_uploader("보완 CSV/TXT", type=["csv", "txt", "zip"], accept_multiple_files=True, key="sgis_supp")
        if supp:
            try:
                supp_raw = cached_load_uploads(_files_sig(supp), tuple(mapping_items or []))
                got2 = {b: len(v) for b, v in supp_raw.items() if len(v)}
                ss.sgis_supp_raw = supp_raw
                st.success("보완 파일 인식: " + ", ".join(f"{b}({n})" for b, n in got2.items()))
            except Exception as e:
                st.error(f"보완 파일 오류: {e}")

    if not api_raw and not ss.get("sgis_supp_raw"):
        return None
    return _merge_raw(api_raw, ss.get("sgis_supp_raw"))


# ══════════════════════════════════════════════════════════════════════════
# STEP 1 — 자료 신청 (SGIS 집계구 자료제공 신청)
# ══════════════════════════════════════════════════════════════════════════
def step1_apply():
    sec("① 자료 신청", "SGIS 집계구 자료를 **신청(제출)까지 자동화**합니다. 승인(약 10분)·다운로드 후 "
        "**② 데이터 입력**에서 받은 폴더를 불러오십시오.")
    _sgis_apply_block()
    st.markdown("")
    _, nav_r = st.columns([3, 1])
    if nav_r.button("다음: 데이터 입력 →", type="primary", use_container_width=True):
        goto(2)


def _fetch_all_sigungu(cookie, progress=None):
    """17개 시도의 전 시군구코드 수집(캐시 재사용). progress(done,total,accum)."""
    ss.setdefault("sgg_cache", {})
    cache = ss.sgg_cache
    allcodes = []
    for i, (sc, _n) in enumerate(SR.SIDO_LIST):
        if sc not in cache:
            try:
                cache[sc] = SR.fetch_sigungu_list(cookie, sc)
            except Exception:
                cache[sc] = []
        allcodes += [c for c, _ in cache.get(sc, [])]
        if progress:
            progress(i + 1, len(SR.SIDO_LIST), len(allcodes))
    return list(dict.fromkeys(allcodes))


def _sgis_download_block(mapping_items):
    """☁ SGIS 승인 자료 자동 다운로드 — 신청→승인 후, 쿠키로 승인 zip을 받아 바로 raw로.
    (SGIS 정부서버라 국내 IP에서만 동작. 데이터신청과 같은 JSESSIONID·accessToken 사용.)"""
    ss.setdefault("dl_items", None)
    ss.setdefault("dl_files", None)
    st.caption("신청 → **승인(약 10분, 이메일)** 후, SGIS 로그인 쿠키로 **승인된 자료 zip을 자동으로 받아 바로 불러옵니다.** "
               "여러 건을 한 번에 받아 전국 배치 빌드로 이어갈 수 있습니다. (국내 IP에서만 동작)")
    ck_raw = st.text_area("SGIS 쿠키 (sgis.mods.go.kr 요청 Copy as cURL 통째로)",
                          value=ss.get("apply_cookie", ""), height=80, key="dl_cookie_raw",
                          help="① 신청 화면에 넣은 쿠키와 같은 걸 써도 됩니다. F12 → Network → F5 → "
                               "맨 위 sgis.mods.go.kr 요청 우클릭 → Copy as cURL → 붙여넣기.")
    if ck_raw:
        ss.apply_cookie = ck_raw   # 매 렌더 저장 → 스텝 이동해도 쿠키 유지(①과 공유)
    if st.button("📋 승인된 자료 목록 불러오기"):
        ck = SR.extract_cookie(ck_raw or "")
        if not ck:
            st.error("쿠키를 붙여넣으십시오(JSESSIONID·accessToken이 확인되지 않으면 SGIS 로그인·복사를 다시 하십시오).")
        else:
            try:
                with st.spinner("SGIS 다운로드 목록 조회 중…"):
                    ss.dl_items = SR.fetch_download_list(ck)
                ss.apply_cookie = ck_raw
                if not ss.dl_items:
                    st.warning("다운로드 가능한(승인완료) 자료가 없습니다. 승인 이메일을 기다리거나 SGIS 신청내역을 확인하십시오.")
            except Exception as e:
                st.error(f"목록 불러오기 실패: {e} — 쿠키가 만료되었으면 SGIS에 다시 로그인해 새 쿠키를 받으십시오. "
                         "(해외 서버에서 실행하면 SGIS가 차단합니다 → 국내 PC·Cloudtype에서 실행)")
    items = ss.get("dl_items")
    if items:
        items = sorted(items, key=lambda it: str(it.get("date", "")), reverse=True)  # 최근 신청 먼저
        if "dl_done" not in ss:
            ss.dl_done = set()          # 이 세션에서 이미 받은 req_id
        done = ss.dl_done
        df = pd.DataFrame(items)[["req_id", "label", "date", "expire"]].copy()
        df.insert(0, "받음", ["✅" if it["req_id"] in done else "" for it in items])
        st.dataframe(df, use_container_width=True, height=240, hide_index=True)
        st.caption("※ 목록 = **승인완료(만료 전) 자료 전부**입니다. 이미 받은 것도 만료 전까지는 다시 표시됩니다 "
                   "(‘받음 ✅’ = 이번 세션에서 받은 표시). SGIS는 만료 전까지 재다운로드가 가능합니다.")
        # ── 신청일·개수로 골라 받기 ──
        dates = sorted({it.get("date", "") for it in items if it.get("date")}, reverse=True)
        f1, f2, f3 = st.columns([2, 1, 1])
        pick_dates = f1.multiselect("신청일로 거르기 (기본 전체)", dates, default=dates, key="dl_dates")
        recent_n = f2.number_input("최근 N건만 (0=전체)", 0, len(items), 0, 1, key="dl_recent_n",
                                   help="신청 늦은(최근) 순서부터 N건.")
        hide_done = f3.checkbox("이미 받은 건 빼기", value=False, key="dl_hide_done")
        sel = [it for it in items if it.get("date", "") in pick_dates]
        if hide_done:
            sel = [it for it in sel if it["req_id"] not in done]
        if recent_n:
            sel = sel[:int(recent_n)]                       # 이미 최근순 정렬됨
        st.caption(f"➡ 받을 자료: **{len(sel)}건** / 전체 {len(items)}건")
        if st.button(f"⬇ 선택 {len(sel)}건 다운로드 → 데이터로 불러오기", type="primary", disabled=not sel):
            ck = SR.extract_cookie(ck_raw or "")
            files, errs = [], []
            prog = st.progress(0, text="다운로드 중…")
            for n, it in enumerate(sel):
                try:
                    blob = SR.download_zip(ck, it["zippath"])
                    files.append((f"{it['req_id']}_{it['zippath'].split('/')[-1]}", blob))
                    done.add(it["req_id"])
                except Exception as e:
                    errs.append(f"{it['req_id']}: {e}")
                prog.progress(int((n + 1) / max(1, len(sel)) * 100), text=f"{n + 1}/{len(sel)}")
            ss.dl_files = files
            if errs:
                st.warning("일부 실패:\n\n" + "\n".join(f"- {e}" for e in errs))
            if files:
                st.success(f"{len(files)}개 zip 다운로드 완료 → 아래 인식 결과를 확인한 뒤 진행하십시오.")
    if ss.get("dl_files"):
        n_files = len(ss.dl_files)
        total_mb = sum(len(b) for _, b in ss.dl_files) / 1e6
        big = (total_mb > 40) or (n_files > 8)
        # ── 대용량: 시군구별 스트리밍 산출 (전체 로드 없이 → 서울 25구 같은 큰 지역도 안 터짐) ──
        with st.expander("🏙 대용량: 시군구별 스트리밍 산출 (전체 로드 없이 · 큰 지역 OK)", expanded=big):
            st.caption("받은 파일을 **시군구별로 하나씩** 처리해 정본 양식(계산방법+복합종합) 엑셀을 각각 생성한 뒤 "
                       "zip으로 묶습니다. 전체를 메모리에 올리지 않아 **큰 지역도 처리 가능합니다.** "
                       "가중치는 ③설정값(설정한 경우)·없으면 정본 기본값을 사용합니다.")
            so = st.text_input("저장 폴더(비워두십시오 · 로컬에서 직접 실행할 때만 사용)", value=ss.get("stream_out", ""),
                               key="in_stream_out", placeholder="비워두면 zip 다운로드만 (권장)",
                               help="웹에서는 비워두십시오. 결과는 아래 ‘zip 다운로드’로 받습니다. "
                                    "이 칸은 본인 PC에서 코드를 직접 실행할 때 각 파일을 폴더에 저장하는 용도입니다.")
            if st.button(f"⚙ 시군구별 스트리밍 산출 시작 ({n_files}개 파일 · {total_mb:.0f}MB)",
                         type="primary", use_container_width=True, key="stream_run"):
                import batch_build as BB
                import time as _time
                inds = TE.indicators_from_cfg(ss.get("cfg") or {})
                am = ss.get("active_map") or {}
                active_recipes = [rc for rc in (ss.get("recipes") or []) if am.get(rc["name"], True)]
                prog = st.progress(0, text="시작…")
                _tm = {"t0": _time.time(), "build0": None}

                def _fmt(sec):   # 초 → "1분 20초" / "45초"
                    sec = int(sec)
                    return f"{sec // 60}분 {sec % 60}초" if sec >= 60 else f"{sec}초"

                def _scb(done, total, sgg, phase="build"):
                    el = _time.time() - _tm["t0"]
                    if phase == "spool":   # 1단계: 원시 파일 정제(총 개수 미정)
                        prog.progress(5, text=f"1/2단계 · 원시 파일 정제 중 {done}개째… (경과 {_fmt(el)})")
                    else:                  # 2단계: 시군구 하나씩 산출(경과+예상 남은시간)
                        if _tm["build0"] is None:
                            _tm["build0"] = _time.time()
                        bel = _time.time() - _tm["build0"]
                        eta = (bel / done) * (total - done) if done else 0
                        pct = int(done / max(1, total) * 100)
                        prog.progress(pct, text=f"2/2단계 · 시군구 {done}/{total} ({sgg}) · "
                                                f"경과 {_fmt(el)} · 남은시간 약 {_fmt(eta)}")

                try:
                    with st.spinner("시군구별 스트리밍 산출 중… (하나씩 처리)"):
                        zb, summ = BB.stream_sigungu_templates(
                            list(ss.dl_files), indicators=inds, custom_df=ss.get("custom_df"),
                            recipes=active_recipes, admin_path=TE.DEFAULT_ADMIN_PATH,
                            sido_name_map=dict(getattr(SR, "SIDO_LIST", [])),
                            selected_years=ss.get("selected_years"),
                            out_dir=(so.strip() or None), progress=_scb)
                    ss.stream_out = so
                    ss.stream_zip, ss.stream_summary = zb, summ
                    ss.stream_secs = _time.time() - _tm["t0"]
                    prog.progress(100, text=f"완료 · 총 {_fmt(ss.stream_secs)}")
                except Exception as e:
                    st.error(f"스트리밍 산출 실패: {e}")
            if ss.get("stream_zip") is not None:
                sm = ss.stream_summary
                ok = int((sm["상태"] == "OK").sum()) if sm is not None else 0
                _secs = ss.get("stream_secs")
                _tt = f" · 소요 {int(_secs)//60}분 {int(_secs)%60}초" if _secs else ""
                st.success(f"완료 — 성공 {ok} / {len(sm)}곳{_tt}")
                st.dataframe(sm, use_container_width=True, height=220, hide_index=True)
                st.download_button("⬇ 시군구별 정본 zip 다운로드", ss.stream_zip,
                                   "쇠퇴진단_시군구별_정본.zip", "application/zip",
                                   use_container_width=True, key="dl_stream")
        # ── 전체 정제(작은 지역용) — 큰 데이터면 자동 로드를 건너뛰어 터짐 방지 ──
        do_full = st.checkbox("받은 데이터 전체를 이 앱으로 정제·분석 (작은 지역만 권장)",
                              value=not big, key="dl_full_load",
                              help="서울 전체처럼 크면 꺼두십시오 — 위 ‘스트리밍 산출’로 시군구별 결과만 받는 것이 안전합니다.")
        if do_full:
            st.caption(f"받아둔 zip {len(ss.dl_files)}개로 데이터 구성 중… (연도/기준연도는 아래에서)")
            try:
                return cached_load_uploads(tuple(ss.dl_files), tuple(mapping_items or []))
            except ValueError as e:
                st.error(str(e))
    return None


# ══════════════════════════════════════════════════════════════════════════
# STEP 2 — 데이터 입력
# ══════════════════════════════════════════════════════════════════════════
def _merge_raw(a, b):
    """기존 raw(a)에 새 raw(b)를 버킷별로 합침(중복 자동 제거). '추가' 모드용."""
    out = {}
    for k in set(a) | set(b):
        frames = [d for d in (a.get(k), b.get(k)) if d is not None and len(d)]
        if not frames:
            out[k] = a.get(k) if a.get(k) is not None else b.get(k)
            continue
        m = pd.concat(frames, ignore_index=True)
        subset = [c for c in ("연도", "집계구", "CODE") if c in m.columns]
        if subset:
            m = m.drop_duplicates(subset=subset, keep="last")
        out[k] = m.reset_index(drop=True)
    return out


def step1_data():
    sec("② 데이터 입력", "원시 SGIS 자료를 넣으면 내부 항목코드 기준으로 자동 분류합니다. "
        "필요하면 보조 파일(매핑·참조코드)을 먼저 올리십시오.")

    if ss.raw is not None:
        try:
            _nsgg = len(BB.list_sigungu(ss.raw))
        except Exception:
            _nsgg = 0
        _rows = sum(len(v) for v in ss.raw.values() if v is not None)
        st.success(f"✅ 데이터가 이미 로드되어 있습니다 — 시군구 **{_nsgg}곳** · 총 **{_rows:,}행**. "
                   "**다시 올리지 않아도 됩니다.** (화면의 업로드 칸이 비어 보여도 데이터는 유지됩니다.) "
                   "아래는 **교체·추가**할 때만 사용하십시오.")

    # 보조 파일
    mapping_items, name_map, code_label_map = None, None, {}
    with st.expander("보조 파일 (선택) — 집계구↔행정동 매핑 · 참조 코드표", expanded=False):
        ca, cb = st.columns(2)
        with ca:
            st.markdown("**집계구↔행정동 매핑**")
            map_template_df = pd.DataFrame({"집계구코드": ["35011100010001"], "행정동코드": ["35011100"], "행정동명": ["예시동"]})
            st.download_button("매핑 템플릿", xlsx_bytes(map_template_df, "매핑"),
                               "집계구_행정동_매핑템플릿.xlsx",
                               "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)
            map_file = st.file_uploader("매핑 (열: 집계구코드, 행정동코드, 행정동명)", type=["xlsx", "csv"],
                                        help="선택 사항. 전국 행정동 이름은 자동으로 붙습니다 — 이 파일은 이름을 덮어쓰거나 보완할 때만 사용합니다.", key="map_file")
            st.caption("✅ **전국 행정동 이름 자동** (행정구역코드 전국표 내장) — 어느 지역이든 파일 없이 이름이 붙습니다. "
                       "이름이 비거나 다르게 나올 때만 이 파일로 보완하십시오.")
        with cb:
            st.markdown("**참조 코드표**")
            st.caption("속성명 인식이 애매할 때만. 코드/명칭 열 자동 추정.")
            ref_file = st.file_uploader("ref_code / 항목코드표 / 산업분류표", type=["xlsx", "xls", "csv"], key="ref_file")

        if map_file is not None:
            try:
                mdf = read_mapping_file(map_file)
                mapping_items, name_map = parse_dong_mapping(mdf)
                if not mapping_items and not name_map:
                    st.error("매핑 파일에서 행정동코드/행정동명 열을 찾지 못했습니다. 열: [행정동코드, 행정동명] 또는 [집계구코드, 행정동코드, 행정동명]")
                else:
                    msg = []
                    if mapping_items:
                        msg.append(f"집계구→행정동 {len(mapping_items)}건")
                    if name_map:
                        msg.append(f"행정동 이름 {len(name_map)}개")
                    st.success("매핑 로드: " + " · ".join(msg))
            except Exception as e:
                st.error(f"매핑 파일 오류: {e}")
        if ref_file is not None:
            try:
                code_label_map = read_ref_code_file(ref_file)
                st.success(f"참조 코드 {len(code_label_map):,}건 로드")
            except Exception as e:
                st.error(f"참조 코드표 오류: {e}")

    # 입력 방식  (자료 신청은 ① 단계에서 별도)
    st.caption("SGIS 집계구 자료를 아직 받지 않았으면 **① 자료 신청**에서 먼저 신청하십시오.")
    src = st.radio("입력 방식", ["원시 SGIS 자료 (드래그·폴더/zip 경로)", "☁ SGIS 승인 자료 자동 다운로드", "원시 SGIS CSV 파일 업로드(인구·산업·물리 탭)", "🌐 SGIS API 직접받기(행정동·보조)", "이미 만든 법적 data 시트(xlsx)"],
                   horizontal=True, key="src_mode")

    raw_new = None
    if src.startswith("☁"):
        raw_new = _sgis_download_block(mapping_items)
    elif src.startswith("🌐 SGIS"):
        raw_new = _sgis_input_block(mapping_items)
    elif src.startswith("이미"):
        up = st.file_uploader("법적 data 개별 시트가 든 xlsx (인구/사업체/건축 등 6개 시트)", type=["xlsx"], key="xlsx_up")
        if up is not None:
            raw_new, miss = extract_raw(up.getvalue(), mapping_items)
            if miss:
                st.error("다음 필수 시트가 없습니다: " + ", ".join(miss))
                raw_new = None
        else:
            st.info("필요한 시트: " + ", ".join(f"`{s}`" for s in golden_io.REQUIRED_SHEETS))

    elif src.startswith("원시 SGIS CSV 파일"):
        st.caption("인구/산업/물리 카테고리별로 CSV/TXT 또는 **zip(압축 그대로)** 을 올리면 내부 CODE 기준으로 자동 분류합니다. **zip은 자동으로 해제됩니다.**")
        with st.expander("컬럼 매핑 보정 (파일 구조가 다를 때만)", expanded=False):
            mc1, mc2, mc3, mc4 = st.columns(4)
            col_year = mc1.number_input("연도 열", 1, 50, 1, 1)
            col_jgu = mc2.number_input("집계구 열", 1, 50, 2, 1)
            col_code = mc3.number_input("항목코드 열", 1, 50, 3, 1)
            col_value = mc4.number_input("값 열", 1, 50, 4, 1)
        tabs = st.tabs(["인구", "산업", "물리"])
        files = []
        with tabs[0]:
            files += st.file_uploader("인구 CSV/TXT", type=["csv", "txt", "zip"], accept_multiple_files=True, key="csv_pop")
        with tabs[1]:
            files += st.file_uploader("산업 CSV/TXT", type=["csv", "txt", "zip"], accept_multiple_files=True, key="csv_biz")
        with tabs[2]:
            files += st.file_uploader("물리 CSV/TXT", type=["csv", "txt", "zip"], accept_multiple_files=True, key="csv_phy")
        if files:
            mapping = dict(mapping_items) if mapping_items else None
            column_map = {"연도": int(col_year) - 1, "집계구": int(col_jgu) - 1,
                          "CODE": int(col_code) - 1, "값": int(col_value) - 1}
            try:
                with st.spinner(f"원시 파일 {len(files)}개 정제 중…"):
                    raw_new = cached_load_uploads(_files_sig(files), tuple(mapping_items or []), column_map=tuple(sorted(column_map.items())) if isinstance(column_map, dict) else column_map)
                    split_summary = L.summarize_uploaded_files(files, column_map=column_map) if hasattr(L, "summarize_uploaded_files") else pd.DataFrame()
                if len(split_summary):
                    with st.expander("분할 CSV 자동 통합 요약", expanded=False):
                        st.dataframe(split_summary, use_container_width=True, height=220)
            except ValueError as e:
                st.error(str(e))
        else:
            st.info("원시 SGIS CSV/TXT 파일을 업로드하십시오.")

    else:  # 드래그앤드롭 또는 폴더/zip 경로
        st.markdown("**① 파일을 여기로 끌어다 놓기** (zip 그대로 OK) — 또는 **② 아래에 폴더/zip 경로 입력**")
        drop = st.file_uploader("여기로 zip·CSV/TXT 드래그앤드롭 (여러 개 OK)",
                                type=["csv", "txt", "zip"], accept_multiple_files=True, key="path_drop")
        st.caption("💡 파일을 **경로 입력칸**에 드래그하면 브라우저 보안상 경로가 입력되지 않습니다 → **위 드롭존**을 사용하십시오. "
                   "폴더/zip 경로 방식은 앱과 **같은 PC**에서 실행할 때만 동작합니다.")
        paths = st.text_area("원시 CSV 폴더/zip 경로 (줄바꿈으로 여러 개)", height=80,
                             placeholder="D:/…/260413_SGIS_베이스자료\nD:/…/원본데이터.zip")
        folders = [p.strip() for p in paths.splitlines() if p.strip()]

        bad = [f for f in folders
               if not (os.path.isdir(f) or (f.lower().endswith(".zip") and os.path.isfile(f)))]
        if drop:                                # ① 드래그앤드롭 우선
            try:
                with st.spinner(f"업로드 {len(drop)}개(zip 자동해제) 정제 중…"):
                    raw_new = cached_load_uploads(_files_sig(drop), tuple(mapping_items or []))
                    split_summary = L.summarize_uploaded_files(drop) if hasattr(L, "summarize_uploaded_files") else pd.DataFrame()
                got = {b: len(v) for b, v in raw_new.items() if len(v)}
                st.success("드래그 파일 인식: " + ", ".join(f"{b}({n})" for b, n in got.items()) if got else "인식된 데이터가 없습니다(형식 확인).")
                if len(split_summary):
                    with st.expander("분할/압축 통합 요약", expanded=False):
                        st.dataframe(split_summary, use_container_width=True, height=220)
            except ValueError as e:
                st.error(str(e))
        elif folders and bad:                   # ② 경로 — 잘못된 경로
            st.error("경로를 찾을 수 없음: " + ", ".join(bad))
        elif folders:                           # ② 경로 — 정상
            try:
                with st.spinner("폴더/zip의 원시 CSV/TXT 정제 중…"):
                    files_found = L.find_csvs(folders)
                    raw_new = cached_load_folders(tuple(folders), tuple(mapping_items or []))
                    split_summary = cached_summarize_folders(tuple(folders))
                st.caption(f"원시 CSV/TXT {len(files_found)}개 + zip 내부 자동해제")
                if len(split_summary):
                    with st.expander("분할 CSV 자동 통합 요약", expanded=False):
                        by_sigungu = (split_summary.groupby("시군구코드")
                                      .agg(파일수=("파일", "nunique"), 행수=("행수", "sum"),
                                           집계구수=("집계구수", "sum"), 항목수=("항목수", "max"),
                                           연도목록=("연도목록", lambda x: ", ".join(sorted(set(", ".join(x).split(", "))))))
                                      .reset_index())
                        st.dataframe(by_sigungu, use_container_width=True)
                        st.dataframe(split_summary, use_container_width=True, height=220)
            except ValueError as e:
                st.error(str(e))
        else:
            st.info("① 파일을 드래그하거나 ② 폴더/zip 경로를 입력하십시오.")

    # 기존 세션 데이터 재사용 / 교체 · 추가
    if raw_new is not None and ss.raw is not None:
        _mode = st.radio("새로 올린 데이터 처리", ["기존에 추가(병합)", "기존을 교체"],
                         horizontal=True, key="merge_mode",
                         help="추가=기존 시군구에 새 데이터를 합침(중복 자동 제거). 교체=기존을 버리고 새 것만.")
        if _mode.startswith("기존에 추가"):
            raw = _merge_raw(ss.raw, raw_new)
            st.success("기존 데이터에 **추가(병합)** 했습니다. (중복 자동 제거)")
        else:
            raw = raw_new
            st.info("기존 데이터를 **교체** 했습니다.")
    else:
        raw = raw_new if raw_new is not None else ss.raw
    if raw is None:
        return

    # 연도 선택 + 상태 요약
    st.markdown("---")
    sec("데이터 인식 결과", "분류별 행 수와 기준연도를 확인한 뒤 다음 단계로 넘어갑니다.")
    years = sheet_builder.all_years(raw)
    default_years = [y for y in (ss.selected_years or years) if y in years] or years
    picked = st.multiselect("출력/산출에 사용할 연도", years, default=default_years,
                            help="지웠다 다시 넣어도 칸 순서와 무관하게 항상 연도순(오름차순)으로 처리됩니다.")
    selected_years = sorted(picked)          # 선택 순서와 무관하게 항상 연도 오름차순
    if not selected_years:
        st.error("최소 1개 이상의 연도를 선택하십시오.")
        return
    st.caption("사용 연도(정렬됨): " + ", ".join(map(str, selected_years)))
    raw_selected = sheet_builder.filter_raw_years(raw, selected_years)

    n_gu = raw["to_in"]["집계구"].nunique()
    n_dong = raw["to_in"]["행정동코드"].nunique()
    recognized = sum(1 for v in raw_selected.values() if len(v))
    c1, c2, c3, c4 = st.columns(4)
    with c1:
        tile("선택 연도", f"{min(selected_years)}–{max(selected_years)}", f"{len(selected_years)}개 연도")
    with c2:
        tile("집계구", f"{n_gu:,}", "원시자료 기준")
    with c3:
        tile("행정동", f"{n_dong:,}", "매핑 적용 후")
    with c4:
        variant = "danger" if recognized < 6 else "accent"
        tile("인식 분류", f"{recognized}/6", f"누락 {6 - recognized}개", variant=variant, danger_value=(recognized < 6))

    pop_cand = bucket_years(raw_selected, ["to_in", "in_age", "ho_yr", "ho_ar"]) or selected_years
    biz_cand = bucket_years(raw_selected, ["to_fa", "cp_bem"]) or selected_years
    r1, r2 = st.columns(2)
    pop_ref = r1.number_input("인구/주택 기준연도", int(min(selected_years)), int(max(selected_years)),
                              int(max(pop_cand)), 1, help="여기 입력한 연도를 기준연도로 사용합니다.")
    biz_ref = r2.number_input("산업 기준연도", int(min(selected_years)), int(max(selected_years)),
                              int(max(biz_cand)), 1, help="총사업체수/종사자수 증감률 기준연도입니다.")

    # ⚠️ 경고는 '선택한 기준연도' 기준으로 계산해야 하므로 coverage_warnings 호출 전에 반영한다.
    #    (이 한 줄이 없으면 기준연도를 바꿔도 경고가 모듈 기본값 2024 기준으로 떠서 오해를 유발)
    C.YEAR_POP_LATEST = int(pop_ref)
    C.YEAR_BIZ_LATEST = int(biz_ref)

    warns = L.coverage_warnings(raw_selected)
    if warns:
        for w in warns:
            st.warning("⚠️ " + w)
    else:
        st.success(f"기준연도(인구 {int(pop_ref)} · 산업 {int(biz_ref)}) 데이터가 모두 확인되었습니다.")

    with st.expander("입력 데이터 분류 상태", expanded=False):
        bs = pd.DataFrame([
            {"분류": k, "행수": len(v), "연도": ", ".join(map(str, sorted(v["연도"].unique().tolist()))) if len(v) else ""}
            for k, v in raw_selected.items()
        ])
        st.dataframe(bs, use_container_width=True, height=240)

    # ── 코드표 안전장치(③): 엔진이 합산할 코드가 실제 데이터에 있는지 점검 ──
    code_warns, code_df = CA.audit(raw_selected, pop_year=int(pop_ref), biz_year=int(biz_ref))
    if code_warns:
        st.warning("⚠️ 코드 점검: 엔진이 합산할 코드 일부가 데이터에 없습니다(값이 조용히 과소계산될 수 있음). "
                   "아래 '코드 점검 상세'에서 확인하십시오.")
    with st.expander(f"코드 점검 상세 — {'⚠ 누락 ' + str(len(code_warns)) + '건' if code_warns else '전부 정상'}",
                     expanded=bool(code_warns)):
        st.caption("기본 12지표는 SGIS 코드를 '위치(positional)'로 합산합니다. 코드체계가 바뀌면 조용히 틀리므로, "
                   "기준연도 데이터에 필요한 코드가 실제 있는지 점검합니다. (계산은 그대로 — 확인·경고용)")
        st.dataframe(code_df, use_container_width=True, height=300, hide_index=True)
        for w in code_warns:
            st.markdown(f"- {w}")

        st.markdown("**엔진이 계산 시 가정하는 코드셋** — 데이터엔 코드의 '뜻'이 없으니, 아래 표로 "
                    "*엔진이 어떤 코드를 더하는지*를 직접 확인하십시오. (SGIS 코드표가 바뀌면 config.py 수정 필요)")
        st.dataframe(CA.assumptions_table(raw_selected, pop_year=int(pop_ref), biz_year=int(biz_ref)),
                     use_container_width=True, height=430, hide_index=True)

    def _empty(k):
        return k not in raw_selected or raw_selected[k].empty
    missing_core = [k for k in REQUIRED_CORE if _empty(k)]
    if missing_core:
        st.info("아래 **필수 분류**가 없으면 최종 진단이 부정확할 수 있습니다(개별 시트는 가능): " + ", ".join(missing_core))
    missing_opt = [k for k in OPTIONAL_BUCKETS if _empty(k)]
    if missing_opt:
        st.caption("· 선택 분류 미포함(없어도 됨): " + ", ".join(missing_opt)
                   + " — 소형주택비율 등 ‘복제’ 지표를 쓸 때만 필요합니다.")

    # (구) '전국 시군구 배치 빌드' expander는 ⑤ 산출의 '선택 시군구 개별 산출'로 이동·개조함
    #  — ③설정 가중치를 쓰려면 설정 뒤(⑤)에 있어야 하므로 위치를 옮김.
    st.caption("💡 여러 시군구를 골라 **③설정 가중치 그대로** 개별 정본 양식으로 뽑으려면 "
               "**⑤ 진단 산출 화면의 ‘선택 시군구 개별 산출’** 을 사용하십시오. "
               "(전국 전체를 디폴트로 자동 신청·빌드하려면 🚀 전국 디폴트 원스톱)")

    st.markdown("")
    nav_l, nav_r = st.columns([3, 1])
    if nav_r.button("다음: 설정 →", type="primary", use_container_width=True):
        ss.raw = raw
        ss.selected_years = selected_years
        ss.pop_ref_year, ss.biz_ref_year = int(pop_ref), int(biz_ref)
        # 업로드 매핑이 있으면 내장표(전주)에 덮어써 합침 — 없으면 내장표 유지
        ss.name_map = {**_national_name_map(), **name_map} if name_map is not None else ss.name_map
        ss.code_label_map = code_label_map or ss.code_label_map
        goto(3)


# ══════════════════════════════════════════════════════════════════════════
# STEP 2 — 설정 (가중치 · 등급 · 추가지표)
# ══════════════════════════════════════════════════════════════════════════
def step2_settings():
    if ss.raw is None:
        st.info("먼저 ② 데이터 입력에서 데이터를 불러오십시오.")
        return
    sec("③ 설정", "모든 지표(기본·계산식·값)를 **지표 마스터** 한 표에서 켜고/끄고 가중치를 정합니다. "
        "기본지표는 '복제해서 수정', 새 지표는 아래 편집기에서 만들고, 외부에서 계산한 값은 파일로 가져올 수 있습니다.")

    # 마스터 표를 화면 위쪽에 두기 위한 자리(내용은 아래 편집기 처리 후 채운다)
    master_box = st.container()
    st.markdown("---")

    # ── (내짐 편집기 1) 계산식 지표 직접 만들기·수정 ──────────────────────────
    with st.expander("🧮 계산식 지표 직접 만들기 · 수정  (분자/분모/카테고리로 정의)",
                     expanded=bool(ss.recipes)):
        st.caption("한 줄 = 한 지표. 코드범위 예: `4-13, 65,66`  ·  분모 비우면 카테고리 전체.  "
                   "카테고리: in_age=성연령·to_in=총인구·to_fa=총사업체·cp_bem=종사자·ho_yr=건축연도·ho_ar=연건평")

        # 추가/샘플 버튼
        ab1, ab2, ab3 = st.columns([2, 2, 2])
        if ab1.button("➕ 계산식 한 줄 추가", use_container_width=True):
            ss.recipes = ss.recipes + [RE.blank_recipe([r["name"] for r in ss.recipes])]
            ss.weights_ver += 1
            st.rerun()
        if ab2.button("🎓 샘플 3종 다시 넣기", use_container_width=True):
            have = {r["name"] for r in ss.recipes}
            ss.recipes = ss.recipes + [r for r in RE.sample_recipes() if r["name"] not in have]
            ss.weights_ver += 1
            st.rerun()

        rdf = RE.recipes_to_df(ss.recipes)
        if len(rdf):
            rdf.insert(0, "삭제", False)
        edited_r = st.data_editor(
            rdf, num_rows="fixed", use_container_width=True, key=f"recipe_ed_{ss.weights_ver}",
            column_config={
                "삭제": st.column_config.CheckboxColumn("삭제", help="체크 후 아래 '체크한 줄 삭제' 버튼", width="small", default=False),
                "부문": st.column_config.SelectboxColumn(options=list(C.SECTORS)),
                "방향": st.column_config.SelectboxColumn(options=["+", "-"]),
                "유형": st.column_config.SelectboxColumn(options=list(RE.TYPE_LABELS.keys())),
                "카테고리": st.column_config.SelectboxColumn(options=list(RE.CATEGORY_LABELS.keys())),
                "스케일": st.column_config.NumberColumn(format="%.2f", step=1.0),
            })
        body = edited_r.drop(columns=["삭제"]) if "삭제" in edited_r.columns else edited_r
        checked = edited_r["삭제"].astype(bool) if "삭제" in edited_r.columns else None

        db1, _ = st.columns([2, 4])
        if db1.button("🗑 체크한 줄 삭제", use_container_width=True,
                      disabled=(checked is None or not checked.any())):
            keep = body[~checked]
            ss.recipes, _ = RE.df_to_recipes(keep)
            ss.weights_ver += 1
            st.rerun()

        new_recipes, rec_errors = RE.df_to_recipes(body)
        ss.recipes = new_recipes
        for e in rec_errors:
            st.warning("지표 정의: " + e)
        if new_recipes:
            st.markdown("**미리보기(수식)** — 이대로 계산됩니다")
            for rc in new_recipes:
                st.caption(f"· **{rc['name']}** ({rc['sector']} · 방향 {rc['direction']}) = {RE.formula_text(rc)}")

    # ── (내짐 편집기 2) 이미 계산된 값 가져오기 (파일) ────────────────────────
    with st.expander("📥 이미 계산된 값 가져오기 (엑셀/CSV) — 외부에서 산출한 지표 흡수", expanded=False):
        st.caption("한 행 = 한 지표 × 한 지역 값. 값↑이 쇠퇴면 +, 값↓이 쇠퇴면 −. 단위코드는 앱 지역코드와 같아야 합니다.")
        st.download_button("값 지표 템플릿 내려받기", CI.template_xlsx_bytes(), "추가지표_템플릿.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        custom_file = st.file_uploader("값 지표 파일 (xlsx/csv)", type=["xlsx", "csv"], key="custom_file")
        if custom_file is not None:
            try:
                cdf = CI.normalize(CI.read_uploaded(custom_file))
                ss.custom_df = cdf
                ss.custom_meta = CI.metadata(cdf)
                if len(ss.custom_meta):
                    st.success(f"값 지표 {len(ss.custom_meta)}개 로드 — 위 마스터 표에 '값' 출처로 합류")
            except Exception as e:
                ss.custom_df, ss.custom_meta = pd.DataFrame(), pd.DataFrame()
                st.error(f"값 지표 오류: {e}")
        if len(ss.custom_meta):
            if st.button("🗑 가져온 값 지표 비우기"):
                ss.custom_df, ss.custom_meta = pd.DataFrame(), pd.DataFrame()
                st.rerun()

    # ── 지표 3갈래(기본/계산식/값) 메타 + 수식을 하나로 통합 ───────────────────
    custom_meta = ss.custom_meta
    sector_of_all = dict(C.SECTOR_OF)
    sign_all = dict(C.SIGN)
    label_all = dict(C.INDLABEL)
    origin_all = {ind: "기본" for ind in C.IND_IDS}
    formula_all = dict(RE.BUILTIN_FORMULAS)
    all_inds = list(C.IND_IDS)
    for _, row in custom_meta.iterrows():
        ind = row["지표"]
        sector_of_all[ind] = row["부문"]
        sign_all[ind] = 10 if row["방향"] == "+" else -10
        label_all[ind] = ind
        origin_all[ind] = "값"
        formula_all[ind] = "외부 계산값(파일)"
        if ind not in all_inds:
            all_inds.append(ind)
    for rc in ss.recipes:
        ind = rc["name"]
        sector_of_all[ind] = rc["sector"]
        sign_all[ind] = 10 if rc["direction"] == "+" else -10
        label_all[ind] = ind
        origin_all[ind] = "계산식"
        formula_all[ind] = RE.formula_text(rc)
        if ind not in all_inds:
            all_inds.append(ind)

    # 최초 진입 시 사용여부/내부가중치 초기화 (DEFAULT_OFF 반영 + 부문 합 100 균등)
    if ss.active_map is None:
        ss.active_map = {ind: (ind not in DEFAULT_OFF) for ind in C.IND_IDS}
    for ind in all_inds:
        ss.active_map.setdefault(ind, ind not in DEFAULT_OFF)   # 샘플 계산식은 꺼진 채 등장
    # 사라진 지표(레시피/값 삭제분)는 사용맵에서 정리
    for gone in [k for k in list(ss.active_map) if k not in all_inds]:
        ss.active_map.pop(gone, None)
    if not ss.internal_wmap:
        ss.internal_wmap = normalize_weights(ss.active_map, {}, sector_of_all)

    # ── 마스터 표 렌더(위쪽 자리에) ────────────────────────────────────────────
    with master_box:
        if ss.clone_notice:
            st.info(ss.clone_notice)
            ss.clone_notice = None

        # 기본지표 복제 툴바
        st.markdown("**기본지표 복제해서 수정** — 검증된 원본은 보존하고, 편집 가능한 사본을 만듭니다(원본은 자동으로 꺼짐).")
        tc1, tc2 = st.columns([3, 1])
        clone_opts = list(RE.BUILTIN_SEEDS.keys())
        clone_sel = tc1.selectbox(
            "복제할 기본지표", clone_opts, key="clone_target",
            format_func=lambda n: f"{n}   {'· 정확 재현' if RE.BUILTIN_SEEDS[n]['exact'] else '· ⚠근사(8차 연도 코드분기 미반영)'}",
            label_visibility="collapsed")
        if tc2.button("📋 복제", use_container_width=True):
            recipe = RE.seed_from_builtin(clone_sel, [rc["name"] for rc in ss.recipes])
            ss.recipes.append(recipe)
            ss.active_map[clone_sel] = False          # 원본 끄기(보존·되돌리기 가능)
            ss.active_map[recipe["name"]] = True
            ss.weights_ver += 1
            approx = "" if RE.BUILTIN_SEEDS[clone_sel]["exact"] else \
                "  ⚠ 이 지표는 산업 8차(≤2005년) 코드분기가 있어 사본은 최신 차수 코드 기준 '근사'입니다."
            ss.clone_notice = (f"'{clone_sel}' → **{recipe['name']}** 사본을 만들었습니다. "
                               f"원본은 꺼두었고, 아래 '🧮 계산식 지표' 편집기에서 수정하십시오.{approx}")
            st.rerun()

        left, right = st.columns([3, 2], gap="large")

        # ── 가중치 ──
        with left:
            st.markdown('<div class="panel">', unsafe_allow_html=True)
            st.markdown("**부문 비율** — 세 부문 합이 100%가 되도록")
            if ss.sector_df is None:
                ss.sector_df = pd.DataFrame({"부문": C.SECTORS, "부문비율": [40.0, 20.0, 40.0]}).set_index("부문")
            sector_edit = st.data_editor(
                ss.sector_df, use_container_width=True,
                column_config={"부문비율": st.column_config.NumberColumn(format="%.2f", step=1.0)},
                key="sector_weights")
            ss.sector_df = sector_edit
            sector_ratio = {s: float(sector_edit.at[s, "부문비율"]) for s in C.SECTORS}
            if abs(sum(sector_ratio.values()) - 100.0) > 1e-6:
                st.warning(f"부문비율 합계 = {sum(sector_ratio.values()):.2f}% → 100%로 맞추십시오.")

            st.markdown("**지표 마스터** — ✅사용으로 켜고/끄고, 부문 내부가중치는 부문 안에서 합 100 기준")
            base_rows = []
            for ind in all_inds:
                base_rows.append({
                    "지표ID": ind, "지표": label_all.get(ind, ind), "출처": origin_all.get(ind, ""),
                    "부문": sector_of_all.get(ind, ""),
                    "방향": "+" if sign_all.get(ind, 10) == 10 else "-",
                    "정의": formula_all.get(ind, ""),
                    "사용": bool(ss.active_map.get(ind, True)),
                    "내부가중치": float(ss.internal_wmap.get(ind, 0.0)),
                    "삭제": False,
                })
            internal_df = pd.DataFrame(base_rows).set_index("지표ID")
            w_edit = st.data_editor(
                internal_df, use_container_width=True, height=460,
                disabled=["지표", "출처", "부문", "방향", "정의"],
                column_order=["사용", "지표", "출처", "부문", "방향", "내부가중치", "정의", "삭제"],
                column_config={
                    "사용": st.column_config.CheckboxColumn("사용", help="체크 해제하면 종합점수 계산에서 제외됩니다.", default=True),
                    "출처": st.column_config.TextColumn("출처", help="기본=검증된 12지표 · 계산식=레시피 · 값=파일", width="small"),
                    "정의": st.column_config.TextColumn("정의(수식)", width="large"),
                    "내부가중치": st.column_config.NumberColumn(format="%.3f", step=1.0),
                    "삭제": st.column_config.CheckboxColumn("삭제", help="계산식·값 지표만 삭제됩니다(기본지표는 '사용' 해제만). 체크 후 아래 '선택 지표 삭제' 버튼.", default=False),
                },
                key=f"internal_weights_{ss.weights_ver}")
            ss.active_map = {ind: bool(w_edit.at[ind, "사용"]) for ind in w_edit.index}
            ss.internal_wmap = {ind: float(w_edit.at[ind, "내부가중치"]) for ind in w_edit.index}
            active_inds = [ind for ind in all_inds if ss.active_map.get(ind)]

            # ── 개별 삭제 (계산식·값 지표만; 기본지표는 삭제 불가 → '사용' 해제만) ──
            del_sel = [ind for ind in w_edit.index if bool(w_edit.at[ind, "삭제"])]
            if del_sel:
                base_locked = [i for i in del_sel if origin_all.get(i) == "기본"]
                removable = [i for i in del_sel if origin_all.get(i) != "기본"]
                dc1, dc2 = st.columns([3, 2])
                note = f"삭제 대기: {', '.join(removable)}" if removable else "삭제할 계산식·값 지표를 체크하십시오."
                if base_locked:
                    note += f"  ·  기본지표는 삭제 불가(사용 해제만): {', '.join(base_locked)}"
                dc1.caption(note)
                if dc2.button("🗑 선택 지표 삭제", use_container_width=True, disabled=not removable):
                    rem = set(removable)
                    ss.recipes = [rc for rc in ss.recipes if rc["name"] not in rem]
                    if ss.custom_df is not None and len(ss.custom_df):
                        ss.custom_df = ss.custom_df[~ss.custom_df["지표명"].isin(rem)].reset_index(drop=True)
                        ss.custom_meta = CI.metadata(ss.custom_df)
                    for i in rem:
                        ss.active_map.pop(i, None)
                        ss.internal_wmap.pop(i, None)
                    ss.weights_ver += 1
                    st.rerun()

            # 부문별 사용 지표 합계 표시 + 자동 맞춤
            sums = []
            for s in C.SECTORS:
                ssum = sum(float(ss.internal_wmap.get(i, 0.0)) for i in active_inds if sector_of_all.get(i) == s)
                mark = "✓" if abs(ssum - 100.0) <= 0.01 else "⚠"
                sums.append(f"{mark} {s} {ssum:.1f}")
            cA, cB, cC = st.columns([3, 2, 2])
            cA.caption("부문 내부합: " + "  ·  ".join(sums))
            if cB.button("⚖ 비율유지 100", use_container_width=True,
                         help="현재 값들의 비율을 유지하며 각 부문 합을 100으로 재조정"):
                ss.internal_wmap = normalize_weights(ss.active_map, ss.internal_wmap, sector_of_all)
                ss.weights_ver += 1
                st.rerun()
            if cC.button("🟰 부문 균등분배", use_container_width=True,
                         help="각 부문의 사용 지표에 100을 똑같이 나눔(예: 5개면 각 20)"):
                ss.internal_wmap = equalize_weights(ss.active_map, sector_of_all)
                ss.weights_ver += 1
                st.rerun()

            final_weight = {
                ind: sector_ratio[sector_of_all[ind]] / 100.0 * float(ss.internal_wmap.get(ind, 0.0)) / 100.0
                for ind in active_inds
            }
            st.caption(f"사용 지표 {len(active_inds)}개 · 최종 가중치 합계 = {sum(final_weight.values()) * 100:.2f}%")
            st.markdown("</div>", unsafe_allow_html=True)

        # ── 등급 방식 ──
        with right:
            st.markdown('<div class="panel">', unsafe_allow_html=True)
            st.markdown("**등급 방식**")
            method_label = st.radio("10등급 분류", ["Natural (Jenks)", "Quantile (등위수)", "Pretty (균등간격)"],
                                    index=0, key="grade_method",
                                    help="기본은 Natural(Jenks). 골든본 고정컷과 다를 수 있음(정상).")
            method = {"Natural (Jenks)": "jenks", "Quantile (등위수)": "quantile", "Pretty (균등간격)": "pretty"}[method_label]
            n_classes = st.number_input("등급 수", 2, 20, 10, 1, key="n_classes")
            decimals = st.number_input("소수점 자릿수(엑셀 값)", 0, 6, 2, 1, key="decimals",
                                       help="엑셀에 쓰는 통계값(값·Z·T·부문·종합 등)의 반올림 자릿수. 기본 2자리.")

            st.markdown("---")
            cnt_base = sum(1 for i in all_inds if origin_all.get(i) == "기본")
            cnt_rec = sum(1 for i in all_inds if origin_all.get(i) == "계산식")
            cnt_val = sum(1 for i in all_inds if origin_all.get(i) == "값")
            st.caption(f"**지표 구성**\n\n· 기본 {cnt_base}개\n· 계산식 {cnt_rec}개\n· 값(파일) {cnt_val}개")
            st.caption("지표 추가·정의는 아래 두 편집기에서, 사용·가중치·삭제(계산식/값)는 위 마스터 표에서 조정하십시오.")
            st.markdown("</div>", unsafe_allow_html=True)

    # 계산 결과 스냅샷 (다음 단계에서 사용) — '사용' 지표만 반영
    ss.cfg = {
        "method": method, "n_classes": int(n_classes), "final_weight": final_weight,
        "indicator_ids": active_inds, "label_map": label_all,
        "sector_of": sector_of_all, "sign_map": sign_all,
        "decimals": int(decimals),
    }

    st.markdown("")
    b1, b2, _ = st.columns([1, 1, 2])
    if b1.button("← 데이터", use_container_width=True):
        goto(2)
    if b2.button("다음: 검토 →", type="primary", use_container_width=True):
        goto(4)


# ══════════════════════════════════════════════════════════════════════════
# STEP 3 — 시트 검토
# ══════════════════════════════════════════════════════════════════════════
def step3_review():
    if ss.raw is None:
        st.info("먼저 ② 데이터 입력에서 데이터를 불러오십시오.")
        return
    sec("④ 시트 검토", "필요한 개별 DATA/피벗 시트를 미리 보고, 원하면 개별 시트만 먼저 Excel로 만들 수 있습니다.")
    raw_selected = sheet_builder.filter_raw_years(ss.raw, ss.selected_years)

    with st.expander("개별 DATA 시트 미리보기", expanded=True):
        preview_key = st.selectbox(
            "미리보기 데이터",
            [("to_in", "법적인구/총인구"), ("to_fa", "법적산업/총사업체수"), ("ho_yr", "법적물리/건축연도"),
             ("in_age", "복합인문/성연령"), ("cp_bem", "복합산업/종사자수"), ("ho_ar", "복합물리/소형주택")],
            format_func=lambda x: x[1])[0]
        if len(raw_selected.get(preview_key, [])):
            st.dataframe(raw_selected[preview_key].head(200), use_container_width=True, height=280)
        else:
            st.info("이 분류의 데이터가 없습니다.")

    with st.expander("만들 시트 선택", expanded=False):
        s1, s2, s3, s4 = st.columns(4)
        with s1:
            st.markdown("**법적 DATA**")
            o_lp = st.checkbox("법적인구", True); o_lb = st.checkbox("법적산업", True); o_lph = st.checkbox("법적물리", True)
        with s2:
            st.markdown("**복합 DATA**")
            o_ch = st.checkbox("복합인문", True); o_cbt = st.checkbox("복합산업-사업체", True)
            o_cbw = st.checkbox("복합산업-종사자", True); o_csh = st.checkbox("복합물리-소형", True); o_coh = st.checkbox("복합물리-노후", True)
        with s3:
            st.markdown("**행정동 피벗**")
            o_ph = st.checkbox("피벗-인문", True); o_pbt = st.checkbox("피벗-사업체", True)
            o_pbw = st.checkbox("피벗-종사자", True); o_psh = st.checkbox("피벗-소형주택", True); o_poh = st.checkbox("피벗-노후건축물", True)
        with s4:
            st.markdown("**최종 진단**")
            o_fld = st.checkbox("법적진단-행정동", True); o_flj = st.checkbox("법적진단-집계구", True)
            o_fcd = st.checkbox("복합지수-행정동", True); o_fcj = st.checkbox("복합지수-집계구", True)
    sheet_options = {
        "legal_population": o_lp, "legal_business": o_lb, "legal_physical": o_lph,
        "complex_human": o_ch, "complex_business_total": o_cbt, "complex_business_worker": o_cbw,
        "complex_small_house": o_csh, "complex_old_house": o_coh,
        "pivot_human": o_ph, "pivot_business_total": o_pbt, "pivot_business_worker": o_pbw,
        "pivot_small_house": o_psh, "pivot_old_house": o_poh,
        "final_legal_dong": o_fld, "final_legal_jgu": o_flj,
        "final_complex_dong": o_fcd, "final_complex_jgu": o_fcj, "summary": True,
    }
    ss.sheet_options = sheet_options

    cM, cL = st.columns(2)
    download_mode = cM.radio("엑셀 다운로드 모드", ["수식 엑셀", "값 엑셀"], horizontal=True, key="download_mode",
                             help="수식 엑셀은 최종 복합표의 값/Z/T/가중합을 엑셀 수식으로 작성합니다.")
    level_label = cL.radio("복합·피벗 집계 단위", ["행정동", "집계구", "행정동+집계구"], horizontal=True,
                           key="pivot_level_ui",
                           help="복합 DATA 피벗을 어느 단위로 만들지 선택. 최종 진단표는 항상 행정동·집계구 둘 다 산출됩니다.")
    ss.pivot_level = {"행정동": "dong", "집계구": "jgu", "행정동+집계구": "both"}[level_label]

    if st.button("📄 선택한 개별 시트 제작", type="secondary"):
        cfg = ss.cfg or {}
        weight_map = cfg.get("final_weight", {ind: float(C.WEIGHT.get(ind, 0.0)) for ind in C.IND_IDS})
        individual = dict(sheet_options)
        for k in ("final_legal_dong", "final_legal_jgu", "final_complex_dong", "final_complex_jgu"):
            individual[k] = False
        if download_mode.startswith("수식"):
            st.info("개별 시트 수식 모드: 행정동 피벗이 원시 DATA를 SUMIFS로 참조합니다.")
        with st.spinner("개별 DATA/피벗 워크북 생성 중…"):
            wb = export.build_integrated_workbook(
                raw_selected, selected_years=ss.selected_years, sheet_options=individual,
                name_map=ss.name_map, indicator_ids=cfg.get("indicator_ids", list(C.IND_IDS)),
                label_map=cfg.get("label_map", dict(C.INDLABEL)), sector_of=cfg.get("sector_of", dict(C.SECTOR_OF)),
                weight=weight_map, sign_map=cfg.get("sign_map", dict(C.SIGN)),
                code_label_map=ss.code_label_map, formula_mode=download_mode.startswith("수식"),
                pivot_level=ss.pivot_level, decimals=int(cfg.get("decimals", 2)))
            buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        st.success("개별 시트 제작 완료.")
        st.download_button("⬇ 개별 시트 xlsx 다운로드", buf.getvalue(), "쇠퇴진단_개별시트.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    st.markdown("")
    b1, b2, _ = st.columns([1, 1, 2])
    if b1.button("← 설정", use_container_width=True):
        goto(3)
    if b2.button("다음: 산출 →", type="primary", use_container_width=True):
        goto(5)


# ══════════════════════════════════════════════════════════════════════════
# STEP 4 — 진단 산출
# ══════════════════════════════════════════════════════════════════════════
def step4_run():
    if ss.raw is None:
        st.info("먼저 ② 데이터 입력에서 데이터를 불러오십시오.")
        return
    sec("⑤ 진단 산출", "법적쇠퇴진단 + 복합쇠퇴지수를 계산하고 통합 Excel을 내려받습니다.")
    raw_selected = sheet_builder.filter_raw_years(ss.raw, ss.selected_years)
    cfg = ss.cfg or {}
    # 원시 6키 방어적 보장(빈 분류라도 키 존재 → 엔진 KeyError 방지)
    for b in REQUIRED_FULL:
        if b not in raw_selected:
            raw_selected[b] = pd.DataFrame(columns=["연도", "집계구", "CODE", "값", "행정동코드"])
    available = {b for b in REQUIRED_FULL if not raw_selected[b].empty}

    b1, _ = st.columns([1, 3])
    if b1.button("← 검토", use_container_width=True):
        goto(4)

    label_map = cfg.get("label_map", dict(C.INDLABEL))

    # ── 하드차단 완화(②): 6분류 강제 대신 '사용 지표가 필요한 분류'만 검사 ──
    #    없는 분류가 필요한 지표는 제외하고 진행(막지 않고 확인만).
    recipe_bucket = {rc["name"]: rc.get("category") for rc in ss.recipes}

    def _bucket_of(ind):
        if ind in C.INDICATOR_BUCKET:
            return C.INDICATOR_BUCKET[ind]      # 기본 12지표
        return recipe_bucket.get(ind)           # 계산식=카테고리 / 값(파일)=None(분류 불필요)

    excluded = {}   # bucket -> [지표,...]
    indicator_ids = []
    for ind in cfg.get("indicator_ids", list(C.IND_IDS)):
        b = _bucket_of(ind)
        if b is not None and b not in available:
            excluded.setdefault(b, []).append(ind)
        else:
            indicator_ids.append(ind)

    if excluded:
        lines = [f"- **{b}** 없음 → 제외: {', '.join(label_map.get(i, i) for i in inds)}"
                 for b, inds in excluded.items()]
        st.warning("데이터가 없어 아래 복합지표를 **빼고** 진행합니다(막지 않고 확인만):\n\n" + "\n".join(lines))

    legal_missing = [b for b in ("to_in", "to_fa", "ho_yr") if b not in available]
    if legal_missing:
        st.info("법적쇠퇴진단 관련 분류 없음 → 해당 부문은 '미충족(x)'으로 처리됩니다: " + ", ".join(legal_missing))

    if not indicator_ids:
        st.error("사용 중인 복합지표에 필요한 데이터가 하나도 없습니다. "
                 "③ 설정에서 데이터가 있는 지표를 켜거나, ② 데이터에서 분류를 보완하십시오.")
        miss_all = [b for b in REQUIRED_FULL if b not in available]
        if miss_all:
            st.dataframe(missing_bucket_message(miss_all), use_container_width=True)
        return

    # ── 코드표 안전장치(③): 산출 직전 재점검 — 누락/신규 코드가 있으면 배너로 경고 ──
    pre_warns, pre_df = CA.audit(raw_selected,
                                 pop_year=ss.get("pop_ref_year"), biz_year=ss.get("biz_ref_year"))
    if pre_warns:
        st.warning("⚠️ 코드 점검: 엔진이 합산할 코드에 문제가 있습니다(값이 조용히 과소/과대계산될 수 있음). "
                   "필요하면 ② 데이터로 돌아가 확인하십시오.")
        with st.expander(f"산출 직전 코드 점검 상세 — ⚠ {len(pre_warns)}건", expanded=False):
            st.dataframe(pre_df, use_container_width=True, height=260, hide_index=True)
            for w in pre_warns:
                st.markdown(f"- {w}")

    method = cfg.get("method", "jenks")
    n_classes = cfg.get("n_classes", 10)
    weight_map = cfg.get("final_weight", {ind: float(C.WEIGHT.get(ind, 0.0)) for ind in C.IND_IDS})
    sector_of = cfg.get("sector_of", dict(C.SECTOR_OF))
    sign_map = cfg.get("sign_map", dict(C.SIGN))
    sheet_options = ss.get("sheet_options", {"summary": True})
    download_mode = ss.get("download_mode", "수식 엑셀")

    # (제거) '최종 4시트(유림_17시)' 옵션 — 정본이 아니라 옛 골든 참조 형식이라 혼동만 줘서 삭제.
    #   출력은 ▶ 통합(26시트) 또는 정본 9시트('정본 양식 생성' / '선택 시군구 개별 산출')만.
    final_only = False
    sort_mode = st.selectbox("행정동/집계구 정렬(출력)",
                             ["기본(데이터순)", "행정동코드 오름차순", "행정동코드 내림차순",
                              "종합점수 높은순", "종합점수 낮은순"],
                             index=["기본(데이터순)", "행정동코드 오름차순", "행정동코드 내림차순",
                                    "종합점수 높은순", "종합점수 낮은순"].index(ss.get("sort_mode", "기본(데이터순)")),
                             key="sort_mode",
                             help="최종표(엑셀)·순위표의 행 정렬 순서. 기본은 데이터순(골든과 동일).")

    # ── 선택 시군구 개별 산출 (③설정 가중치 그대로 · 시군구별 정본 양식 · 따로 저장) ──
    with st.expander("🏙 선택 시군구 개별 산출 (내 설정 적용 · 시군구별 정본 양식)", expanded=False):
        import batch_build as BB
        st.caption("업로드한 데이터에서 시군구를 **골라**, ③설정의 **가중치·지표 구성 그대로** "
                   "시군구마다 **정본 양식(계산방법+복합종합) 엑셀을 따로** 만듭니다. "
                   "각 시군구는 그 안에서 독립 표준화되고, 커스텀·계산식 지표값도 반영됩니다.")
        sgg_all = BB.list_sigungu(raw_selected)
        if not sgg_all:
            st.info("데이터에서 시군구를 인식하지 못했습니다(집계구 14자리 또는 시군구 5자리 필요).")
        else:
            sido_name_map = dict(getattr(SR, "SIDO_LIST", []))
            pick = st.multiselect(
                "처리할 시군구 선택", sgg_all,
                format_func=lambda c: f"{c} {sido_name_map.get(c[:2], '')}", key="pick_sigungu")
            pick_out = st.text_input(
                "저장 폴더 경로(선택 · 로컬 실행 시 각 파일을 여기 저장)",
                value=ss.get("pick_out", ""), key="in_pick_out",
                placeholder=r"예: D:\쇠퇴진단_선택출력", help="비우면 zip 다운로드만.")
            if st.button(f"⚙ 선택 시군구 개별 산출 ({len(pick)}곳)", type="secondary",
                         use_container_width=True, disabled=not pick):
                try:
                    inds = TE.indicators_from_cfg(ss.cfg)
                    active_recipes = [rc for rc in ss.recipes
                                      if ss.active_map.get(rc["name"], True) and rc["name"] in indicator_ids]
                    prog = st.progress(0, text=f"0/{len(pick)}")

                    def _cb(done, total, sgg):
                        prog.progress(int(done / max(1, total) * 100), text=f"{done}/{total} · {sgg}")

                    with st.spinner("시군구별 정본 양식 생성 중…"):
                        zbytes, summary = BB.build_batch_zip(
                            raw_selected, sigungu=pick, sido_name_map=sido_name_map,
                            year_pop=int(ss.get("pop_ref_year", C.YEAR_POP_LATEST)),
                            year_biz=int(ss.get("biz_ref_year", C.YEAR_BIZ_LATEST)),
                            template_mode=True, indicators=inds,
                            custom_df=ss.custom_df, recipes=active_recipes,
                            admin_path=TE.DEFAULT_ADMIN_PATH,
                            out_dir=(pick_out.strip() or None), progress=_cb)
                    ss.pick_out = pick_out
                    ss.pick_zip, ss.pick_summary = zbytes, summary
                    prog.progress(100, text="완료")
                    if pick_out.strip():
                        st.info(f"📁 저장 완료: `{pick_out.strip()}` 폴더에 시군구별 정본 양식 저장됨.")
                except Exception as e:
                    st.error(f"선택 시군구 산출 실패: {e}")
            if ss.get("pick_zip") is not None:
                ok = int((ss.pick_summary["상태"] == "OK").sum()) if ss.get("pick_summary") is not None else 0
                st.success(f"완료 — 성공 {ok} / {len(ss.pick_summary)}곳")
                st.dataframe(ss.pick_summary, use_container_width=True, height=200, hide_index=True)
                st.download_button("⬇ 선택 시군구 zip 다운로드", ss.pick_zip, "쇠퇴진단_선택시군구.zip",
                                   "application/zip", use_container_width=True, key="dl_pick")

    run_clicked = st.button("▶ 최종 법적 + 복합 진단 산출", type="primary", use_container_width=True)

    if run_clicked:
        # 엔진 전역에 가중치·기준연도 반영
        for ind in C.IND_IDS:
            C.WEIGHT[ind] = float(weight_map.get(ind, 0.0))
        C.YEAR_POP_LATEST = int(ss.pop_ref_year)
        C.YEAR_BIZ_LATEST = int(ss.biz_ref_year)

        prog = st.progress(0, text="0% · 행정동 복합지수 계산 중")
        dong_base = E.run(raw_selected, level="dong", grade_method=method, n_classes=int(n_classes))
        prog.progress(25, text="25% · 집계구 복합지수 계산 중")
        jgu_base = E.run(raw_selected, level="jgu", grade_method=method, n_classes=int(n_classes))
        prog.progress(45, text="45% · 추가지표 및 가중치 반영 중")
        custom_df = ss.custom_df
        custom_dong = CI.build_scores(custom_df, dong_base[0].index, "dong")
        custom_jgu = CI.build_scores(custom_df, jgu_base[0].index, "jgu")
        # 계산식 지표(레시피) — '사용' 중이고 데이터가 있어 제외되지 않은 것만.
        active_recipes = [rc for rc in ss.recipes
                          if ss.active_map.get(rc["name"], True) and rc["name"] in indicator_ids]
        recipe_dong = RE.build_recipe_scores(active_recipes, raw_selected, "dong", dong_base[0].index)
        recipe_jgu = RE.build_recipe_scores(active_recipes, raw_selected, "jgu", jgu_base[0].index)
        dong_scores = CI.combine_scores(CI.combine_scores(dong_base[0], custom_dong), recipe_dong)
        jgu_scores = CI.combine_scores(CI.combine_scores(jgu_base[0], custom_jgu), recipe_jgu)
        dong_comp = CI.composite(dong_scores, indicator_ids, sector_of, weight_map)
        jgu_comp = CI.composite(jgu_scores, indicator_ids, sector_of, weight_map)
        dong_grades = E.assign_grades(dong_comp["종합"], int(n_classes), method)
        jgu_grades = E.assign_grades(jgu_comp["종합"], int(n_classes), method)
        dong = (dong_scores, dong_comp, dong_grades, dong_base[3])
        jgu = (jgu_scores, jgu_comp, jgu_grades, jgu_base[3])
        prog.progress(65, text="65% · 법적쇠퇴진단 계산 중")
        legal_dong = LG.run_legal(raw_selected, level="dong")
        legal_jgu = LG.run_legal(raw_selected, level="jgu")
        # ── 출력 정렬(선택) — 행정동/집계구 순서 재배열(값 무관·표시순만) ──
        if sort_mode != "기본(데이터순)":
            od = _sorted_order(dong_comp, sort_mode)
            dong_scores, dong_comp, dong_grades = dong_scores.reindex(od), dong_comp.reindex(od), dong_grades.reindex(od)
            legal_dong = legal_dong.reindex(od)
            oj = _sorted_order(jgu_comp, sort_mode)
            jgu_scores, jgu_comp, jgu_grades = jgu_scores.reindex(oj), jgu_comp.reindex(oj), jgu_grades.reindex(oj)
            legal_jgu = legal_jgu.reindex(oj)
            dong = (dong_scores, dong_comp, dong_grades, dong[3])
            jgu = (jgu_scores, jgu_comp, jgu_grades, jgu[3])

        prog.progress(80, text="80% · 통합 엑셀 생성 중")
        wb = export.build_integrated_workbook(
            raw_selected, selected_years=ss.selected_years, sheet_options=sheet_options,
            name_map=ss.name_map, dong_res=dong[:3], jgu_res=jgu[:3],
            n_classes=int(n_classes), method=method, legal_dong=legal_dong, legal_jgu=legal_jgu,
            indicator_ids=indicator_ids, label_map=label_map, sector_of=sector_of,
            weight=weight_map, sign_map=sign_map, code_label_map=ss.code_label_map,
            formula_mode=download_mode.startswith("수식"), pivot_level=ss.pivot_level,
            final_only=False, decimals=int(cfg.get("decimals", 2)))
        if download_mode.startswith("수식"):
            for sn in ("전주시 복합쇠퇴지수(행정동)", "전주시 복합쇠퇴지수(집계구)"):
                if sn in wb.sheetnames and not str(wb[sn]["C3"].value).startswith("="):
                    st.error(f"{sn} 수식 생성 확인 실패: C3 셀이 수식이 아닙니다.")
                    return
            if "복합지표값(행정동)" in wb.sheetnames and "전주시 법적쇠퇴진단(행정동)" in wb.sheetnames:
                if not str(wb["복합지표값(행정동)"]["B2"].value).startswith("="):
                    st.error("복합지표값(행정동) 인구변화율이 법적진단 참조 수식으로 생성되지 않았습니다.")
                    return
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        prog.progress(100, text="100% · 완료")

        n_decl = int((legal_dong["쇠퇴지역"] == "o").sum())
        # 정본 양식용: 집계구별 지표값(기본+커스텀+계산식 전체)을 jgu_scores에서 뽑아 저장
        template_values = TE.values_from_scores(jgu_scores, indicator_ids)
        ss.results = {
            "dong_comp": dong_comp, "dong_grades": dong_grades, "dong_stats": dong[3],
            "legal_dong": legal_dong, "n_dong": len(dong[0]), "n_jgu": len(jgu[0]),
            "n_decl": n_decl, "xlsx": buf.getvalue(), "formula": download_mode.startswith("수식"),
            "final_only": False,
            "template_values": template_values,
        }

    # 결과 표시
    res = ss.results
    if not res:
        st.caption("위 버튼을 눌러 진단을 산출하십시오. 산출 결과 요약과 순위 차트, 통합 Excel 다운로드가 여기에 표시됩니다.")
        return

    st.success("진단 산출 완료")
    comp, grades = res["dong_comp"], res["dong_grades"]
    avg_grade = pd.to_numeric(pd.Series(grades), errors="coerce").mean()
    worst_idx = comp["종합"].idxmax()
    worst_name = (ss.name_map or {}).get(worst_idx, str(worst_idx))

    t1, t2, t3, t4 = st.columns(4)
    with t1:
        tile("법적 쇠퇴지역", f"{res['n_decl']}", "부합 2개 이상 · 행정동", variant="danger", danger_value=True)
    with t2:
        tile("최고 쇠퇴 행정동", worst_name, f"종합점수 {comp['종합'].max():.1f}")
    with t3:
        tile("평균 등급", f"{avg_grade:.1f}" if pd.notna(avg_grade) else "—", f"{len(comp)}개 행정동")
    with t4:
        tile("산출 규모", f"{res['n_dong']} · {res['n_jgu']:,}", "행정동 · 집계구")

    dl_label = "⬇ 통합 xlsx 다운로드 (개별 DATA + 법적 + 복합)"
    dl_name = ("쇠퇴진단_통합결과_수식.xlsx" if res["formula"] else "쇠퇴진단_통합결과_값.xlsx")
    st.download_button(
        dl_label, res["xlsx"], dl_name,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary", use_container_width=True)

    # ── 정본 양식(차장님 형태 · ③설정 가중치 그대로 반영) ──
    with st.container(border=True):
        st.markdown("**📐 정본 양식 (계산방법 + 복합쇠퇴진단 종합 · 집계구 단위)**")
        st.caption("③ 설정의 **방향부호·최종가중치**가 첫 시트(계산방법 E24~)에 그대로 써지고, "
                   "복합쇠퇴진단 종합 시트가 그 셀을 **엑셀 함수로 참조**합니다. "
                   "설정을 바꿔 다시 산출하면 값이 함수로 자동 재계산됩니다.")
        if st.button("정본 양식 생성 (가중치 반영)", use_container_width=True, key="build_template"):
            try:
                inds = TE.indicators_from_cfg(ss.cfg)
                with st.spinner("정본 양식 생성 중… (원본 9시트 · 원시데이터 많으면 시간 걸림)"):
                    # 산출 때 저장한 집계구 지표값(기본+커스텀+계산식) 재활용 → 재계산 없음
                    # raw가 있으면 9시트 전부, 없으면 2시트(계산방법+복합종합)로 폴백
                    if ss.get("raw"):
                        twb = TE.build_full_workbook(ss.raw, values=res.get("template_values"),
                                                     indicators=inds)
                    else:
                        twb = TE.build_composite_workbook(indicators=inds,
                                                          values=res.get("template_values"))
                    tbuf = io.BytesIO(); TE.save_wb(twb, tbuf); tbuf.seek(0)
                ss.results["template_xlsx"] = tbuf.getvalue()
                st.success(f"정본 양식 생성 완료 — 지표 {len(inds)}개 · "
                           f"최종가중치 합계 {sum(i[4] for i in inds) * 100:.2f}%")
            except FileNotFoundError:
                st.error(f"행정구역코드 표를 찾을 수 없습니다: {TE.DEFAULT_ADMIN_PATH}\n"
                         "행정구역코드_전국.xlsx 위치를 확인하십시오.")
            except Exception as e:
                st.error(f"정본 양식 생성 실패: {e}")
        if res.get("template_xlsx"):
            st.download_button(
                "⬇ 정본 양식 xlsx 다운로드 (계산방법 + 복합종합)",
                res["template_xlsx"], "복합쇠퇴진단_정본양식.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True, key="dl_template")

    st.markdown("")
    chart_col, table_col = st.columns([2, 3], gap="large")
    with chart_col:
        st.markdown("**행정동 복합쇠퇴 순위 (상위 15)**")
        rank = pd.DataFrame({"행정동": [ (ss.name_map or {}).get(i, str(i)) for i in comp.index ],
                             "종합점수": comp["종합"].round(2).values,
                             "등급": pd.to_numeric(pd.Series(grades, index=comp.index), errors="coerce").values})
        rank = rank.sort_values("종합점수", ascending=False).head(15)
        import altair as alt          # 지연 import(결과 화면에서만 필요)
        chart = (alt.Chart(rank).mark_bar(cornerRadiusEnd=3)
                 .encode(
                     x=alt.X("종합점수:Q", title="종합점수"),
                     y=alt.Y("행정동:N", sort="-x", title=None),
                     color=alt.Color("종합점수:Q", scale=alt.Scale(scheme="teals"), legend=None),
                     tooltip=["행정동", "종합점수", "등급"])
                 .properties(height=max(260, 26 * len(rank))))
        st.altair_chart(chart, use_container_width=True)

    with table_col:
        rtab = st.tabs(["복합쇠퇴지수", "법적쇠퇴진단", "표준화 통계"])
        with rtab[0]:
            prev = pd.DataFrame({
                "종합점수": comp["종합"].round(2), "종합등급": grades,
                "인문사회": comp["인문사회"].round(2), "산업경제": comp["산업경제"].round(2),
                "물리환경": comp["물리환경"].round(2),
            }).sort_values("종합점수", ascending=False)
            st.dataframe(prev, use_container_width=True, height=420)
        with rtab[1]:
            lg = res["legal_dong"].copy()
            for c in ("인구증감률", "사업체증감률", "노후건축물비율"):
                if c in lg:
                    lg[c] = lg[c].round(2)
            st.dataframe(lg.sort_values("부합개수", ascending=False), use_container_width=True, height=420)
        with rtab[2]:
            st.dataframe(pd.DataFrame(res["dong_stats"]).T[["mean", "std"]].round(4),
                         use_container_width=True, height=420)


def _finish_nw_zip(outdir, rows, localdir=None):
    """청크 빌드로 outdir에 쌓인 시군구 xlsx들 + 요약을 하나의 zip(bytes)으로 묶어 반환.
    반환: (zip_bytes, 요약 DataFrame). 메모리에 xlsx를 계속 안 들고 디스크에서 읽어 묶음."""
    import os as _os
    import io as _io
    import zipfile as _zip
    summary = pd.DataFrame(rows) if rows else pd.DataFrame()
    buf = _io.BytesIO()
    with _zip.ZipFile(buf, "w", _zip.ZIP_DEFLATED, compresslevel=1) as zf:
        try:
            for fn in sorted(_os.listdir(outdir)):
                fp = _os.path.join(outdir, fn)
                if _os.path.isfile(fp):
                    zf.write(fp, fn)
        except Exception:
            pass
        try:
            csv_bytes = summary.to_csv(index=False).encode("utf-8-sig")
            zf.writestr("_요약.csv", csv_bytes)
            if localdir:
                try:
                    with open(_os.path.join(localdir, "_요약.csv"), "wb") as fh:
                        fh.write(csv_bytes)
                except Exception:
                    pass
        except Exception:
            pass
    buf.seek(0)
    return buf.getvalue(), summary


def nationwide_onestop():
    """🚀 전국 디폴트 원스톱 — 일반 단계별 분석과 분리된 전용 화면.
    전국 시군구를 '디폴트' 형태로 처음부터 끝까지 자동 구축:
      1) 전국 신청  →  (승인 대기: SGIS 이메일)  →  2) 자동 다운로드 → 시군구별 디폴트 엑셀 → zip/폴더 저장."""
    import batch_build as BB
    sec("🚀 전국 디폴트 원스톱",
        "전국 시군구를 **디폴트 지표·가중치** 그대로 집계구별로 받아, 시군구별 엑셀(함수 포함 전체 시트·"
        "집계구+행정동·최종표)을 한 번에 구축합니다. 일반 단계별 분석과 분리된 전용 화면입니다.")
    if st.button("← 일반 단계별 분석으로 돌아가기"):
        ss.nationwide = False
        st.rerun()

    st.info("흐름: **① 전국 신청** → (SGIS 승인 이메일 대기, 약 10분~) → **② 자동 다운로드 + 229개 빌드**. "
            "승인 사이의 대기만 사람이 기다리면 되고, 나머지는 버튼 두 번으로 끝납니다.")

    # ── 0. 쿠키(공통) ──
    ck_raw = st.text_area("SGIS 쿠키 (sgis.mods.go.kr 요청 Copy as cURL 통째로)",
                          value=ss.get("apply_cookie", ""), height=80, key="nw_cookie",
                          help="F12 → Network → F5 → 맨 위 sgis.mods.go.kr 요청 우클릭 → Copy as cURL → "
                               "붙여넣기. 국내 IP에서만 동작(배포 Cloudtype OK).")

    # ── 1. 전국 신청 ──
    with st.container(border=True):
        st.markdown("### 1. 전국 시군구 신청")
        st.caption("전국 모든 시군구의 집계구 자료를 **디폴트 필수항목**으로 한 번에 신청합니다. "
                   "승인 알림 이메일이 반드시 필요합니다.")
        nw_email = st.text_input("승인 알림 이메일", value=ss.get("apply_email", ""), key="nw_email")
        if nw_email.strip():          # 일반 화면과 이메일 공유(화면 전환에도 유지)
            ss.apply_email = nw_email
        if st.button("🏛 전국 신청 실행", key="nw_apply"):
            ck = SR.extract_cookie(ck_raw or "")
            checked = [n for n in SR.ITEM_CATALOG if SR.ITEM_META[n][0] == "필수"]
            items = _apply_build_items(checked)
            if not ck:
                st.error("쿠키를 붙여넣으십시오.")
            elif not nw_email.strip():
                st.error("승인 알림 이메일을 입력하십시오 — 비어 있으면 SGIS가 신청을 거부합니다.")
            else:
                ss.apply_cookie, ss.apply_email = ck_raw, nw_email
                lp = st.progress(0, text="전국 시군구 목록 수집 중…")
                allcodes = _fetch_all_sigungu(
                    ck, progress=lambda d, t, a: lp.progress(int(d / t * 100), text=f"{d}/{t} 시도 · 누적 {a}곳"))
                if not allcodes:
                    st.error("시군구 목록을 불러오지 못했습니다 — 쿠키 만료 또는 해외 IP 차단.")
                else:
                    applicant = _apply_build_applicant("", "", nw_email, "", "복합쇠퇴진단")
                    st.info(f"전국 **{len(allcodes)}개 시군구** 신청 시작…")
                    sp = st.progress(0, text=f"0/{len(allcodes)}")
                    results = _apply_submit_batch(
                        ck, allcodes, items, applicant, only_first=False,
                        progress=lambda d, t, sg: sp.progress(int(d / max(1, t) * 100), text=f"{d}/{t} · {sg}"))
                    _apply_render_results(results)
                    st.caption("↳ 승인 이메일이 오면 아래 **2번**을 실행하십시오.")

    # ── 2. 자동 다운로드 → 디폴트 빌드 → 저장 (청크 처리: 시군구 몇 개씩 나눠 빌드 → 타임아웃 방지) ──
    with st.container(border=True):
        st.markdown("### 2. 승인 자료 자동 다운로드 → 디폴트 시트 빌드 → 저장")
        st.caption("승인 이메일을 받은 뒤 실행하면, 승인된 자료를 **자동 다운로드**해 디스크에 스풀한 뒤, "
                   "시군구를 **몇 개씩 나눠(청크)** 빌드합니다. 각 단계가 짧아 **연결 끊김(타임아웃) 없이** "
                   "서울·전국 같은 대량도 처리되며, 중간에 멈춰도 받은 것까지 zip으로 받을 수 있습니다. "
                   "출력 = **정본 9시트 양식**(계산방법·법적종합·복합종합·원시 6시트), 함수식은 **계산값이 채워진 채** 저장됩니다.")
        oc1, oc2 = st.columns([3, 1])
        nw_out = oc1.text_input("저장 폴더 경로(선택 · 로컬 실행 시 착착 저장)", value=ss.get("batch_out", ""),
                                key="nw_out", placeholder=r"예: D:\쇠퇴진단_전국출력")
        nw_chunk = oc2.number_input("한 번에 빌드(청크)", 1, 20, int(ss.get("nw_chunk", 3)), 1, key="nw_chunk",
                                    help="한 요청에서 처리할 시군구 수. 작을수록 요청이 짧아 타임아웃에 안전(2~4 권장).")

        # 1) 다운로드 + 스풀(준비) — 파일을 하나씩 받아 시군구별 pkl로 디스크에 쌓음(메모리 안전)
        if st.button("① 다운로드 + 준비 (스풀)", key="nw_prep", disabled=bool(ss.get("nw_building"))):
            import os as _os, tempfile as _tf
            ck = SR.extract_cookie(ck_raw or "")
            if not ck:
                st.error("쿠키를 붙여넣으십시오.")
            else:
                try:
                    with st.spinner("승인 자료 목록 조회 중…"):
                        items = SR.fetch_download_list(ck)
                except Exception as e:
                    items = None
                    st.error(f"목록 조회 실패: {e} — 쿠키 만료/해외 IP 여부를 확인하십시오.")
                if items is not None and not items:
                    st.warning("다운로드 가능한(승인완료) 자료가 없습니다. 승인 이메일을 기다리십시오.")
                elif items:
                    ss.apply_cookie = ck_raw
                    # 이전 진행상태 정리
                    _old = ss.get("nw_tmp")
                    if _old:
                        import shutil as _sh
                        _sh.rmtree(_old, ignore_errors=True)
                    tmp = _tf.mkdtemp(prefix="nw_")
                    parts, ymeta, errs = {}, {}, []
                    dp = st.progress(0, text=f"0/{len(items)} 다운로드·스풀")
                    for n, it in enumerate(items):
                        try:
                            blob = SR.download_zip(ck, it["zippath"])
                            fname = f"{it['req_id']}_{it['zippath'].split('/')[-1]}"
                            BB.spool_files([(fname, blob)], tmp, sgg_parts=parts, year_meta=ymeta)
                            del blob
                        except Exception as e:
                            errs.append(f"{it['req_id']}: {e}")
                        dp.progress(int((n + 1) / len(items) * 100), text=f"{n + 1}/{len(items)} 다운로드·스풀")
                    order = sorted(parts.keys())
                    if errs:
                        st.warning("일부 다운로드 실패:\n\n" + "\n".join(f"- {e}" for e in errs))
                    if not order:
                        st.error("스풀된 시군구가 없습니다(자료 형식 확인).")
                    else:
                        ss.nw_tmp, ss.nw_parts, ss.nw_order = tmp, parts, order
                        ss.nw_done, ss.nw_rows = [], []
                        ss.nw_outdir = _os.path.join(tmp, "_out"); _os.makedirs(ss.nw_outdir, exist_ok=True)
                        ss.nw_localdir = (nw_out.strip() or None)
                        ss.nw_params = {"template_mode": True,   # 정본 9시트(함수식+계산값 캐시)
                                        "year_pop": int(ymeta.get("pop", 2024)),
                                        "year_biz": int(ymeta.get("biz", 2023))}
                        ss.nw_zip = None
                        ss.nw_building = True
                        st.info(f"시군구 **{len(order)}곳** 스풀 완료 · 기준연도 인구{ss.nw_params['year_pop']}/"
                                f"산업{ss.nw_params['year_biz']} → 청크 빌드 시작합니다.")
                        st.rerun()

        # 2) 청크 빌드 — K개씩 빌드하고 st.rerun으로 다음 청크 자동 진행(요청이 짧아 타임아웃 안전)
        if ss.get("nw_building"):
            order = ss.nw_order
            total = len(order)
            ndone = len(ss.nw_done)
            st.progress(ndone / total if total else 0.0,
                        text=f"청크 빌드 {ndone}/{total} 시군구 완료…")
            stop = st.button("⏸ 중지 (여기까지 받은 것만 zip)", key="nw_stop")
            if stop:
                ss.nw_building = False
            else:
                import os as _os
                p = ss.nw_params
                C.YEAR_POP_LATEST = int(p["year_pop"]); C.YEAR_BIZ_LATEST = int(p["year_biz"])
                sido_map = dict(getattr(SR, "SIDO_LIST", []))
                inds = TE.indicators_from_cfg(ss.get("cfg") or {})   # 정본 지표(디폴트=기본 12지표)
                done_set = set(ss.nw_done)
                remaining = [s for s in order if s not in done_set]
                for sgg in remaining[:int(nw_chunk)]:
                    fname, blob, row = BB.build_sigungu_from_pkls(
                        sgg, ss.nw_parts[sgg], template_mode=True,
                        indicators=inds, admin_path=TE.DEFAULT_ADMIN_PATH,
                        sido_name_map=sido_map)
                    if blob:
                        try:
                            with open(_os.path.join(ss.nw_outdir, fname), "wb") as fh:
                                fh.write(blob)
                        except Exception:
                            pass
                        if ss.get("nw_localdir"):
                            try:
                                _os.makedirs(ss.nw_localdir, exist_ok=True)
                                with open(_os.path.join(ss.nw_localdir, fname), "wb") as fh:
                                    fh.write(blob)
                            except Exception:
                                pass
                    ss.nw_rows.append(row)
                    ss.nw_done.append(sgg)
            # 다음 청크 자동 진행 or 마무리(zip)
            if ss.get("nw_building") and len(ss.nw_done) < len(order):
                st.rerun()
            else:
                ss.nw_building = False
                ss.nw_zip, ss.nw_summary = _finish_nw_zip(
                    ss.nw_outdir, ss.nw_rows, ss.get("nw_localdir"))
                _tmp = ss.get("nw_tmp")
                if _tmp:
                    import shutil as _sh
                    _sh.rmtree(_tmp, ignore_errors=True)
                ss.nw_tmp = None; ss.nw_parts = None
                st.rerun()

        if ss.get("nw_zip") is not None and not ss.get("nw_building"):
            sm = ss.get("nw_summary")
            ok = int((sm["상태"] == "OK").sum()) if sm is not None else 0
            st.success(f"전국 빌드 완료 — 성공 {ok}/{len(sm) if sm is not None else 0}곳")
            if sm is not None:
                st.dataframe(sm, use_container_width=True, height=260, hide_index=True)
            if ss.get("nw_localdir"):
                st.info(f"📁 저장 완료: `{ss.nw_localdir}` 폴더")
            st.download_button("⬇ 전국 배치 zip 다운로드", ss.nw_zip, "쇠퇴진단_전국배치.zip",
                               "application/zip", type="primary", use_container_width=True)


# ──────────────────────────────────────────────────────────────────────────
# 라우팅
# ──────────────────────────────────────────────────────────────────────────
st.markdown("")
if ss.get("nationwide"):
    nationwide_onestop()
else:
    {1: step1_apply, 2: step1_data, 3: step2_settings, 4: step3_review, 5: step4_run}[ss.step]()
