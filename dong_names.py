# -*- coding: utf-8 -*-
"""
dong_names.py — 행정동 코드→이름 표 (전국 자동 로드 + 전주 내장 폴백)
=====================================================================
SGIS 원시 파일(4열)엔 행정동 '이름'이 없어, 결과표에 코드(35011600)로만 나온다.

전국판: 같은 저장소의 ``행정구역코드/행정구역코드_전국.xlsx``(SGIS, 3,559동)를
자동으로 읽어 default_name_map()이 **전국 행정동 이름**을 반환한다.
그 파일이 없거나 openpyxl이 없으면 아래 **전주 34동 내장표로 폴백**(기존 동작 유지).
이 덕분에 배치/앱의 어느 진입점도 코드 수정 없이 전국 행정동 이름을 얻는다.

출처: SGIS addr/stage API(2026-07-08 수집). 코드 체계는 통계청/SGIS(전북=35).
표 갱신은 행정구역코드/sgis_admin_codes.py 재실행.
"""
from __future__ import annotations

import os

# 전주시 완산구(35011*)·덕진구(35012*) 행정동 34개 — 전국표가 없을 때의 폴백.
JEONJU = {
    "35011600": "동서학동", "35011610": "서서학동", "35011620": "중화산1동", "35011630": "중화산2동",
    "35011640": "평화1동", "35011650": "평화2동", "35011660": "서신동", "35011670": "삼천1동",
    "35011680": "삼천2동", "35011690": "삼천3동", "35011700": "효자1동", "35011710": "효자2동",
    "35011720": "효자3동", "35011740": "중앙동", "35011750": "풍남동", "35011760": "노송동",
    "35011770": "완산동", "35011780": "효자4동", "35011790": "효자5동", "35012540": "인후1동",
    "35012550": "인후2동", "35012560": "인후3동", "35012570": "덕진동", "35012600": "팔복동",
    "35012610": "우아1동", "35012620": "우아2동", "35012630": "호성동", "35012650": "송천1동",
    "35012660": "송천2동", "35012670": "조촌동", "35012690": "진북동", "35012700": "혁신동",
    "35012710": "여의동", "35012721": "금암동",
}

# 내장 기본표(폴백). 지자체 확장 시 여기에 dict를 합치면 됨.
BUILTIN = dict(JEONJU)

# 전국 코드표 위치 — 이 모듈과 같은 폴더의 행정구역코드/ 안.
ADMIN_XLSX = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "행정구역코드", "행정구역코드_전국.xlsx",
)

# 로드 결과 캐시 — 배치에서 default_name_map()을 시군구마다 불러도 xlsx는 1회만 읽는다.
_CACHE = {"dong": None, "sgg": None, "sido": None, "loaded": False}


def _load_national(path=None):
    """행정구역코드_전국.xlsx → (dong8→명, sgg5→명, sido2→명).
    파일/openpyxl 부재 등 어떤 실패든 (None, None, None) 반환(폴백 신호)."""
    path = path or ADMIN_XLSX
    if not os.path.isfile(path):
        return None, None, None
    try:
        import openpyxl
    except Exception:
        return None, None, None
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception:
        return None, None, None
    try:
        ws = wb["전체"] if "전체" in wb.sheetnames else wb[wb.sheetnames[0]]
        dong, sgg, sido = {}, {}, {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            # 헤더: 연번 | 시도코드 | 시도명 | 시군구코드 | 시군구명 | 행정동코드 | 행정동명
            if not row or len(row) < 7:
                continue
            sd = str(row[1] or "").strip()
            sdn = str(row[2] or "").strip()
            sg = str(row[3] or "").strip()
            sgn = str(row[4] or "").strip()
            dc = str(row[5] or "").strip()
            dn = str(row[6] or "").strip()
            if sd:
                sido[sd] = sdn
            if sg:
                sgg[sg] = sgn
            if dc:
                dong[dc] = dn
        return dong, sgg, sido
    finally:
        wb.close()


def _ensure():
    """전국표를 한 번만 읽어 캐시. 이후 호출은 캐시 반환."""
    if not _CACHE["loaded"]:
        d, s, sd = _load_national()
        _CACHE.update(dong=d, sgg=s, sido=sd, loaded=True)
    return _CACHE


def national_loaded() -> bool:
    """전국표를 실제로 읽었는지(True) 아니면 전주 폴백인지(False)."""
    return bool(_ensure()["dong"])


def default_name_map():
    """행정동 코드→이름 표(사본).
    전국표가 있으면 전주 내장에 덮어써 **전국 3,559동**, 없으면 전주 34동만."""
    c = _ensure()
    if c["dong"]:
        return {**BUILTIN, **c["dong"]}
    return dict(BUILTIN)


def default_sigungu_map():
    """시군구코드(5자리)→이름 표(사본). 전국표 없으면 빈 dict."""
    return dict(_ensure()["sgg"] or {})


def default_sido_map():
    """시도코드(2자리)→이름 표(사본). 전국표 없으면 빈 dict."""
    return dict(_ensure()["sido"] or {})
