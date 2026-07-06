# -*- coding: utf-8 -*-
"""
app.py — 복합쇠퇴진단 자동화 앱 (Streamlit)
============================================
골든본/정제 xlsx 업로드 → 가중치·등급방식 설정 → 복합쇠퇴지수 xlsx 다운로드.
집계구·행정동 두 레벨을 모두 산출한다.

실행:  .venv/bin/streamlit run app.py
"""
import io
import os
import pandas as pd
import openpyxl
import streamlit as st

import config as C
import custom_indicators as CI
import decline_engine as E
import export
import golden_io
import loader as L
import legal_engine as LG
import sheet_builder

st.set_page_config(page_title="쇠퇴진단 자동화 시스템", page_icon="▦", layout="wide")
st.markdown(
    """
    <style>
    .stApp { background: #F6F8FB; }
    .block-container { padding-top: 1.5rem; padding-bottom: 3rem; max-width: 1440px; }
    [data-testid="stSidebar"] { background: #FFFFFF; border-right: 1px solid #E3E8EF; }
    .app-hero {
        border: 1px solid #DCE3EC;
        border-radius: 10px;
        background: linear-gradient(135deg, #FFFFFF 0%, #EEF4F8 100%);
        padding: 24px 28px;
        margin-bottom: 18px;
    }
    .brand-kicker {
        color: #2E7D6B;
        font-size: 0.78rem;
        font-weight: 700;
        letter-spacing: .08em;
        text-transform: uppercase;
        margin-bottom: 6px;
    }
    .brand-title {
        color: #17202A;
        font-size: 2.0rem;
        line-height: 1.15;
        font-weight: 760;
        margin-bottom: 8px;
    }
    .brand-subtitle {
        color: #52616F;
        font-size: 0.98rem;
        margin-bottom: 16px;
    }
    .step-row { display: flex; gap: 8px; flex-wrap: wrap; }
    .step-pill {
        border: 1px solid #D7DEE8;
        background: #FFFFFF;
        color: #263442;
        border-radius: 999px;
        padding: 7px 12px;
        font-size: 0.84rem;
        font-weight: 650;
    }
    .section-title {
        color: #17202A;
        font-size: 1.08rem;
        font-weight: 760;
        margin: 20px 0 2px 0;
    }
    .section-subtitle {
        color: #65727E;
        font-size: 0.88rem;
        margin-bottom: 8px;
    }
    .status-card {
        border: 1px solid #DCE3EC;
        border-radius: 8px;
        background: #FFFFFF;
        padding: 14px 16px;
        min-height: 76px;
    }
    .status-label {
        color: #65727E;
        font-size: 0.78rem;
        font-weight: 650;
        margin-bottom: 4px;
    }
    .status-value {
        color: #17202A;
        font-size: 1.34rem;
        font-weight: 760;
    }
    .status-note {
        color: #7B8794;
        font-size: 0.76rem;
        margin-top: 4px;
    }
    div.stButton > button {
        border-radius: 7px;
        border: 1px solid #CAD3DF;
        font-weight: 700;
    }
    div.stDownloadButton > button {
        border-radius: 7px;
        font-weight: 700;
    }
    </style>
    <div class="app-hero">
      <div class="brand-kicker">SGIS Decline Analyzer</div>
      <div class="brand-title">쇠퇴진단 자동화 시스템</div>
      <div class="brand-subtitle">SGIS 원시자료 기반 법적쇠퇴진단 · 복합쇠퇴지수 · 검토용 Excel 산출</div>
      <div class="step-row">
        <div class="step-pill">1 자료 입력</div>
        <div class="step-pill">2 분류 인식</div>
        <div class="step-pill">3 개별 시트 검토</div>
        <div class="step-pill">4 진단 산출</div>
        <div class="step-pill">5 Excel 다운로드</div>
      </div>
    </div>
    """,
    unsafe_allow_html=True,
)


def section_header(title, subtitle=""):
    st.markdown(f'<div class="section-title">{title}</div>', unsafe_allow_html=True)
    if subtitle:
        st.markdown(f'<div class="section-subtitle">{subtitle}</div>', unsafe_allow_html=True)


def status_card(label, value, note=""):
    st.markdown(
        f"""
        <div class="status-card">
          <div class="status-label">{label}</div>
          <div class="status-value">{value}</div>
          <div class="status-note">{note}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


# ── 업로드 파싱 (파일 바이트 기준 캐시) ──
@st.cache_data(show_spinner="원시 데이터 추출 중…")
def extract_raw(file_bytes: bytes, mapping_items):
    mapping = dict(mapping_items) if mapping_items else None
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        miss = golden_io.missing_sheets(wb)
        if miss:
            return None, miss
        return golden_io.load_raw_from_workbook(wb, mapping), []
    finally:
        wb.close()


def xlsx_bytes(df, sheet_name):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    buf.seek(0)
    return buf.getvalue()


def read_mapping_file(uploaded):
    name = uploaded.name.lower()
    if name.endswith(".xlsx"):
        return pd.read_excel(uploaded, dtype=str).fillna("")
    for enc in ("utf-8-sig", "cp949", "euc-kr", "utf-8"):
        try:
            uploaded.seek(0)
            return pd.read_csv(uploaded, dtype=str, encoding=enc).fillna("")
        except Exception:
            continue
    raise ValueError("매핑 파일을 읽지 못했습니다. xlsx 형식을 권장합니다.")


def read_ref_code_file(uploaded):
    name = uploaded.name.lower()
    if name.endswith(".xlsx") or name.endswith(".xls"):
        df = pd.read_excel(uploaded, dtype=str).fillna("")
    else:
        df = None
        for enc in ("utf-8-sig", "cp949", "euc-kr", "utf-8"):
            try:
                uploaded.seek(0)
                df = pd.read_csv(uploaded, dtype=str, encoding=enc).fillna("")
                break
            except Exception:
                continue
        if df is None:
            raise ValueError("참조 코드표를 읽지 못했습니다.")
    cols = list(df.columns)
    code_col = next((c for c in cols if "CODE" in str(c).upper() or "코드" in str(c)), cols[0])
    name_col = next((c for c in cols if c != code_col and ("명" in str(c) or "항목" in str(c) or "name" in str(c).lower())), cols[1] if len(cols) > 1 else cols[0])
    return dict(zip(df[code_col].astype(str).str.strip(), df[name_col].astype(str).str.strip()))


@st.cache_data(show_spinner=False)
def cached_load_folders(folders_tuple, mapping_items):
    mapping = dict(mapping_items) if mapping_items else None
    return L.load_raw_from_folders(list(folders_tuple), mapping=mapping)


@st.cache_data(show_spinner=False)
def cached_summarize_folders(folders_tuple):
    return L.summarize_folders(list(folders_tuple)) if hasattr(L, "summarize_folders") else pd.DataFrame()


def bucket_years(raw, keys):
    years = set()
    for key in keys:
        if key in raw and len(raw[key]):
            years.update(pd.to_numeric(raw[key]["연도"], errors="coerce").dropna().astype(int).tolist())
    return sorted(years)


BUCKET_GUIDE = pd.DataFrame([
    {"분류": "to_in", "용도": "법적인구/인구변화율", "필요 CODE": "to_in_001", "파일 예": "총인구"},
    {"분류": "in_age", "용도": "복합 인문사회", "필요 CODE": "in_age_001~021, 065~068", "파일 예": "성연령별인구"},
    {"분류": "to_fa", "용도": "법적산업/총사업체수", "필요 CODE": "to_fa_010", "파일 예": "총괄사업체수"},
    {"분류": "cp_bem", "용도": "복합 산업경제", "필요 CODE": "cp_bem_*", "파일 예": "산업분류별 종사자수"},
    {"분류": "ho_yr", "용도": "법적물리/노후건축물", "필요 CODE": "ho_yr_*", "파일 예": "건축년도별주택"},
    {"분류": "ho_ar", "용도": "복합 물리환경/소형주택", "필요 CODE": "ho_ar_*", "파일 예": "연건평별주택"},
])


def missing_bucket_message(missing):
    return BUCKET_GUIDE[BUCKET_GUIDE["분류"].isin(missing)].reset_index(drop=True)


# ── 사이드바: 설정 ──
with st.sidebar:
    st.header("⚙️ 설정")

    st.subheader("등급 방식")
    method_label = st.radio(
        "10등급 분류 방식", ["Natural (Jenks)", "Quantile (등위수)", "Pretty (균등간격)"],
        index=0, help="기본은 Natural(Jenks). 골든본 고정컷과 다를 수 있음(정상).")
    method = {"Natural (Jenks)": "jenks", "Quantile (등위수)": "quantile",
              "Pretty (균등간격)": "pretty"}[method_label]
    n_classes = st.number_input("등급 수", 2, 20, 10, 1)

    st.subheader("추가지표")
    st.caption("한 행 = 한 지표 × 한 지역 값. 방향은 값이 높을수록 쇠퇴면 +, 낮을수록 쇠퇴면 -.")
    st.download_button(
        "추가지표 템플릿 다운로드",
        data=CI.template_xlsx_bytes(),
        file_name="추가지표_템플릿.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    custom_file = st.file_uploader("추가지표 xlsx/csv", type=["xlsx", "csv"])
    try:
        custom_df = CI.normalize(CI.read_uploaded(custom_file)) if custom_file is not None else pd.DataFrame()
        custom_meta = CI.metadata(custom_df)
        if len(custom_meta):
            st.success(f"추가지표 {len(custom_meta)}개 로드")
    except Exception as e:
        custom_df = pd.DataFrame()
        custom_meta = pd.DataFrame()
        st.error(f"추가지표 오류: {e}")

    st.subheader("가중치")
    st.caption("부문 기본값\n인구 40 · 산업 20 · 물리 40")
    sector_df = pd.DataFrame({"부문": C.SECTORS, "부문비율": [40.0, 20.0, 40.0]}).set_index("부문")
    sector_edit = st.data_editor(
        sector_df,
        use_container_width=True,
        column_config={"부문비율": st.column_config.NumberColumn(format="%.2f", step=1.0)},
        key="sector_weights",
    )
    base_rows = []
    sec_cnt = {s: sum(1 for i in C.IND_IDS if C.SECTOR_OF[i] == s) for s in C.SECTORS}
    for ind in C.IND_IDS:
        base_rows.append({
            "지표ID": ind,
            "지표": C.INDLABEL[ind],
            "부문": C.SECTOR_OF[ind],
            "방향": "+" if C.SIGN[ind] == 10 else "-",
            "내부가중치": round(100.0 / sec_cnt[C.SECTOR_OF[ind]], 6),
            "출처": "기본",
        })
    for _, row in custom_meta.iterrows():
        base_rows.append({
            "지표ID": row["지표"],
            "지표": row["지표"],
            "부문": row["부문"],
            "방향": row["방향"],
            "내부가중치": float(row["내부가중치"]),
            "출처": "추가",
        })
    internal_df = pd.DataFrame(base_rows).set_index("지표ID")
    w_edit = st.data_editor(
        internal_df,
        use_container_width=True,
        disabled=["지표", "부문", "방향", "출처"],
        column_config={"내부가중치": st.column_config.NumberColumn(format="%.3f", step=1.0)},
        key="internal_weights",
    )
    sector_ratio = {sec: float(sector_edit.at[sec, "부문비율"]) for sec in C.SECTORS}
    final_weight = {
        ind: sector_ratio[w_edit.at[ind, "부문"]] / 100.0 * float(w_edit.at[ind, "내부가중치"]) / 100.0
        for ind in w_edit.index
    }
    final_sum = sum(final_weight.values())
    if abs(sum(sector_ratio.values()) - 100.0) > 1e-6:
        st.warning(f"부문비율 합계 = {sum(sector_ratio.values()):.2f}% 입니다. 100%가 되도록 조정하세요.")
    for sec in C.SECTORS:
        s = float(w_edit.loc[w_edit["부문"] == sec, "내부가중치"].sum())
        if abs(s - 100.0) > 0.01:
            st.warning(f"{sec} 내부가중치 합계 = {s:.2f}% 입니다.")
    st.caption("내부가중치 합계는 각 부문 안에서 100% 기준입니다. 산업경제 부문비율 20%와는 별도로 계산됩니다.")
    st.caption(f"최종 가중치 합계 = {final_sum * 100:.2f}%")

    st.subheader("집계구↔행정동 매핑 (선택)")
    map_template_df = pd.DataFrame({
        "집계구코드": ["35011100010001"],
        "행정동코드": ["35011100"],
        "행정동명": ["예시동"],
    })
    st.download_button(
        "매핑 xlsx 템플릿 다운로드",
        xlsx_bytes(map_template_df, "매핑"),
        "집계구_행정동_매핑템플릿.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    map_file = st.file_uploader(
        "매핑 파일 (xlsx 권장, 열: 집계구코드, 행정동코드, 행정동명)", type=["xlsx", "csv"],
        help="없으면 집계구코드 앞 8자리를 행정동코드로 사용.")

    st.subheader("참조 코드표 (선택)")
    ref_file = st.file_uploader(
        "ref_code / 항목코드표 / 산업분류표", type=["xlsx", "xls", "csv"],
        help="속성명 인식이 애매할 때 넣습니다. 코드 열과 명칭 열을 자동 추정합니다.")

# 매핑 파싱
mapping_items, name_map = None, None
if map_file is not None:
    try:
        mdf = read_mapping_file(map_file)
        cols = list(mdf.columns)
        c_gu, c_dong = cols[0], cols[1]
        mapping_items = tuple(zip(mdf[c_gu].str.strip(), mdf[c_dong].str.strip()))
        if len(cols) >= 3:
            name_map = dict(zip(mdf[c_dong].str.strip(), mdf[cols[2]].str.strip()))
        st.sidebar.success(f"매핑 {len(mapping_items)}건 로드")
    except Exception as e:
        st.sidebar.error(f"매핑 파일 오류: {e}")

code_label_map = {}
if ref_file is not None:
    try:
        code_label_map = read_ref_code_file(ref_file)
        st.sidebar.success(f"참조 코드 {len(code_label_map):,}건 로드")
    except Exception as e:
        st.sidebar.error(f"참조 코드표 오류: {e}")

# ── 메인: 입력 방식 선택 & 실행 ──
section_header("자료 입력", "원시 SGIS CSV/TXT 폴더 또는 파일을 넣으면 내부 항목코드 기준으로 자동 분류합니다.")
src = st.radio("입력 방식",
               ["원시 SGIS CSV 폴더 경로", "원시 SGIS CSV 파일 업로드", "이미 만든 법적 data 시트(xlsx) 업로드"],
               horizontal=True)

raw = None
if src.startswith("이미"):
    up = st.file_uploader(
        "법적 data 개별 시트가 든 xlsx 업로드 (인구/사업체/건축 등 6개 시트)", type=["xlsx"])
    if up is None:
        st.info("법적 data(원시 CSV에 항목 속성만 붙여 나열한) 시트가 든 xlsx를 업로드하세요. "
                "필요한 시트: " + ", ".join(f"`{s}`" for s in golden_io.REQUIRED_SHEETS))
        st.stop()
    raw, miss = extract_raw(up.getvalue(), mapping_items)
    if miss:
        st.error("다음 필수 시트가 없습니다: " + ", ".join(miss))
        st.stop()
elif src.startswith("원시 SGIS CSV 파일"):
    st.caption("인구/산업/물리 카테고리별로 CSV/TXT를 올리면 내부 CODE 기준으로 자동 분류합니다.")
    with st.expander("컬럼 매핑 보정", expanded=False):
        st.caption("SGIS 기본 CSV/TXT는 1~4열이 연도·집계구·항목코드·값입니다. 파일 구조가 다를 때만 바꾸세요.")
        mc1, mc2, mc3, mc4 = st.columns(4)
        col_year = mc1.number_input("연도 열", 1, 50, 1, 1)
        col_jgu = mc2.number_input("집계구 열", 1, 50, 2, 1)
        col_code = mc3.number_input("항목코드 열", 1, 50, 3, 1)
        col_value = mc4.number_input("값 열", 1, 50, 4, 1)
    tabs = st.tabs(["인구", "산업", "물리"])
    files = []
    with tabs[0]:
        files += st.file_uploader("인구 CSV/TXT", type=["csv", "txt"], accept_multiple_files=True, key="csv_pop")
    with tabs[1]:
        files += st.file_uploader("산업 CSV/TXT", type=["csv", "txt"], accept_multiple_files=True, key="csv_biz")
    with tabs[2]:
        files += st.file_uploader("물리 CSV/TXT", type=["csv", "txt"], accept_multiple_files=True, key="csv_phy")
    if not files:
        st.info("원시 SGIS CSV/TXT 파일을 업로드하세요.")
        st.stop()
    mapping = dict(mapping_items) if mapping_items else None
    column_map = {"연도": int(col_year) - 1, "집계구": int(col_jgu) - 1,
                  "CODE": int(col_code) - 1, "값": int(col_value) - 1}
    progress = st.progress(0, text="0% · 업로드 파일 확인 중")
    try:
        progress.progress(30, text=f"30% · 원시 파일 {len(files)}개 읽는 중")
        raw = L.load_raw_from_uploaded_files(files, mapping=mapping, column_map=column_map)
        progress.progress(75, text="75% · 시군구/연도 커버리지 요약 중")
        split_summary = L.summarize_uploaded_files(files, column_map=column_map) if hasattr(L, "summarize_uploaded_files") else pd.DataFrame()
        progress.progress(100, text="100% · 정제 완료")
    except ValueError as e:
        st.error(str(e))
        st.stop()
    if len(split_summary):
        with st.expander("분할 CSV 자동 통합 요약", expanded=True):
            st.dataframe(split_summary, use_container_width=True, height=220)
else:
    st.caption("원시 CSV/TXT 폴더 경로를 줄바꿈으로 여러 개 입력(로컬/네트워크 경로). cp949·헤더없음 4열 형식.")
    paths = st.text_area("원시 CSV 폴더 경로", height=90,
                         placeholder="/Volumes/NO NAME/260412_SGIS_베이스자료\n/Volumes/NO NAME/260413_SGIS_베이스자료")
    folders = [p.strip() for p in paths.splitlines() if p.strip()]
    if not folders:
        st.info("원시 CSV/TXT가 담긴 폴더 경로를 입력하세요.")
        st.stop()
    bad = [f for f in folders if not os.path.isdir(f)]
    if bad:
        st.error("폴더를 찾을 수 없음: " + ", ".join(bad))
        st.stop()
    mapping = dict(mapping_items) if mapping_items else None
    progress = st.progress(0, text="0% · 폴더와 원시 파일 확인 중")
    try:
        files_found = L.find_csvs(folders)
        progress.progress(20, text=f"20% · 원시 CSV/TXT {len(files_found)}개 발견")
        raw = cached_load_folders(tuple(folders), tuple(mapping_items or []))
        progress.progress(75, text="75% · 시군구/연도 커버리지 요약 중")
        split_summary = cached_summarize_folders(tuple(folders))
        progress.progress(100, text="100% · 정제 완료")
    except ValueError as e:
        st.error(str(e))
        st.stop()
    if len(split_summary):
        with st.expander("분할 CSV 자동 통합 요약", expanded=True):
            by_sigungu = (split_summary.groupby("시군구코드")
                          .agg(파일수=("파일", "nunique"), 행수=("행수", "sum"),
                               집계구수=("집계구수", "sum"), 항목수=("항목수", "max"),
                               연도목록=("연도목록", lambda x: ", ".join(sorted(set(", ".join(x).split(", "))))))
                          .reset_index())
            st.dataframe(by_sigungu, use_container_width=True)
            st.dataframe(split_summary, use_container_width=True, height=220)

# 가중치 반영 (엔진 전역 override + 추가지표 메타)
for ind in C.IND_IDS:
    C.WEIGHT[ind] = float(final_weight.get(ind, 0.0))

indicator_ids = list(C.IND_IDS) + [i for i in w_edit.index if i not in C.IND_IDS]
label_map = dict(C.INDLABEL)
sector_of = dict(C.SECTOR_OF)
sign_map = dict(C.SIGN)
weight_map = dict(final_weight)
for ind in indicator_ids:
    if ind not in label_map:
        label_map[ind] = ind
    if ind in w_edit.index:
        sector_of[ind] = w_edit.at[ind, "부문"]
        sign_map[ind] = 10 if w_edit.at[ind, "방향"] == "+" else -10

# 입력 요약
section_header("데이터 인식 결과", "분류별 행 수와 기준연도를 확인한 뒤 개별 시트 또는 최종 진단을 산출합니다.")
years = sheet_builder.all_years(raw)
selected_years = st.multiselect("출력/산출에 사용할 연도", years, default=years)
if not selected_years:
    st.error("최소 1개 이상의 연도를 선택하세요.")
    st.stop()
raw_selected = sheet_builder.filter_raw_years(raw, selected_years)
n_gu = raw["to_in"]["집계구"].nunique()
n_dong = raw["to_in"]["행정동코드"].nunique()
recognized_now = sum(1 for k, v in raw_selected.items() if len(v))
missing_now = 6 - recognized_now
card1, card2, card3, card4 = st.columns(4)
with card1:
    status_card("선택 연도", f"{min(selected_years)}–{max(selected_years)}", f"{len(selected_years)}개 연도")
with card2:
    status_card("집계구", f"{n_gu:,}", "원시자료 기준")
with card3:
    status_card("행정동", f"{n_dong:,}", "매핑 적용 후")
with card4:
    status_card("인식 분류", f"{recognized_now}/6", f"누락 {missing_now}개")

pop_ref_candidates = bucket_years(raw_selected, ["to_in", "in_age", "ho_yr", "ho_ar"]) or selected_years
biz_ref_candidates = bucket_years(raw_selected, ["to_fa", "cp_bem"]) or selected_years
ref1, ref2 = st.columns(2)
pop_ref_year = ref1.number_input(
    "인구/주택 기준연도",
    min_value=int(min(selected_years)),
    max_value=int(max(selected_years)),
    value=int(max(pop_ref_candidates)),
    step=1,
    help="앱이 임의로 2024를 기준으로 쓰지 않습니다. 여기 입력한 연도를 기준연도로 사용합니다.",
)
biz_ref_year = ref2.number_input(
    "산업 기준연도",
    min_value=int(min(selected_years)),
    max_value=int(max(selected_years)),
    value=int(max(biz_ref_candidates)),
    step=1,
    help="총사업체수/종사자수 증감률 기준연도입니다.",
)
C.YEAR_POP_LATEST = int(pop_ref_year)
C.YEAR_BIZ_LATEST = int(biz_ref_year)
for w in L.coverage_warnings(raw_selected):
    st.warning("⚠️ " + w)

bucket_status = pd.DataFrame([
    {"분류": k, "행수": len(v), "연도": ", ".join(map(str, sorted(v["연도"].unique().tolist()))) if len(v) else ""}
    for k, v in raw_selected.items()
])
with st.expander("입력 데이터 분류 상태", expanded=False):
    st.dataframe(bucket_status, use_container_width=True, height=220)

section_header("개별 시트 검토", "필요한 원시 DATA/피벗 시트를 선택해 먼저 검토용 Excel로 만들 수 있습니다.")
with st.expander("개별 DATA 시트 미리보기", expanded=False):
    preview_key = st.selectbox(
        "미리보기 데이터",
        [("to_in", "법적인구/총인구"), ("to_fa", "법적산업/총사업체수"), ("ho_yr", "법적물리/건축연도"),
         ("in_age", "복합인문/성연령"), ("cp_bem", "복합산업/종사자수"), ("ho_ar", "복합물리/소형주택")],
        format_func=lambda x: x[1],
    )[0]
    st.dataframe(raw_selected[preview_key].head(200), use_container_width=True, height=260)

with st.expander("만들 시트 선택", expanded=False):
    st.caption("필요한 개별 시트만 선택해서 제작할 수 있습니다.")
    s1, s2, s3, s4 = st.columns(4)
    with s1:
        st.markdown("**법적 DATA**")
        opt_legal_population = st.checkbox("법적인구", value=True)
        opt_legal_business = st.checkbox("법적산업", value=True)
        opt_legal_physical = st.checkbox("법적물리", value=True)
    with s2:
        st.markdown("**복합 DATA**")
        opt_complex_human = st.checkbox("복합인문", value=True)
        opt_complex_business_total = st.checkbox("복합산업-사업체", value=True)
        opt_complex_business_worker = st.checkbox("복합산업-종사자", value=True)
        opt_complex_small_house = st.checkbox("복합물리-소형", value=True)
        opt_complex_old_house = st.checkbox("복합물리-노후", value=True)
    with s3:
        st.markdown("**행정동 피벗**")
        opt_pivot_human = st.checkbox("피벗-인문", value=True)
        opt_pivot_business_total = st.checkbox("피벗-사업체", value=True)
        opt_pivot_business_worker = st.checkbox("피벗-종사자", value=True)
        opt_pivot_small_house = st.checkbox("피벗-소형주택", value=True)
        opt_pivot_old_house = st.checkbox("피벗-노후건축물", value=True)
    with s4:
        st.markdown("**최종 진단**")
        opt_final_legal_dong = st.checkbox("법적진단-행정동", value=True)
        opt_final_legal_jgu = st.checkbox("법적진단-집계구", value=True)
        opt_final_complex_dong = st.checkbox("복합지수-행정동", value=True)
        opt_final_complex_jgu = st.checkbox("복합지수-집계구", value=True)
    sheet_options = {
        "legal_population": opt_legal_population,
        "legal_business": opt_legal_business,
        "legal_physical": opt_legal_physical,
        "complex_human": opt_complex_human,
        "complex_business_total": opt_complex_business_total,
        "complex_business_worker": opt_complex_business_worker,
        "complex_small_house": opt_complex_small_house,
        "complex_old_house": opt_complex_old_house,
        "pivot_human": opt_pivot_human,
        "pivot_business_total": opt_pivot_business_total,
        "pivot_business_worker": opt_pivot_business_worker,
        "pivot_small_house": opt_pivot_small_house,
        "pivot_old_house": opt_pivot_old_house,
        "final_legal_dong": opt_final_legal_dong,
        "final_legal_jgu": opt_final_legal_jgu,
        "final_complex_dong": opt_final_complex_dong,
        "final_complex_jgu": opt_final_complex_jgu,
        "summary": True,
    }

download_mode = st.radio(
    "엑셀 다운로드 모드",
    ["수식 엑셀", "값 엑셀"],
    horizontal=True,
    help="수식 엑셀은 최종 복합표의 값/Z/T/가중합을 엑셀 수식으로 작성합니다.",
)

required_for_full = ["to_in", "in_age", "to_fa", "cp_bem", "ho_yr", "ho_ar"]
missing_buckets = [k for k in required_for_full if k not in raw_selected or raw_selected[k].empty]
if missing_buckets:
    with st.expander("최종 진단 산출 전 보충이 필요한 원시 분류", expanded=True):
        st.warning("아래 분류가 없어서 최종 법적+복합 진단은 아직 계산하지 않습니다. 개별 시트 제작은 가능합니다.")
        st.dataframe(missing_bucket_message(missing_buckets), use_container_width=True)
        st.caption("속성 인식이 애매하면 SGIS 원시 파일의 항목코드가 위 CODE와 맞는지 확인하세요. 산업은 8/9/10차 대분류 종사자수(cp_bem_*)와 총괄사업체수(to_fa_010)가 필요합니다.")

section_header("산출", "개별 시트 제작과 최종 진단 산출을 분리해 실행합니다.")
btn1, btn2 = st.columns(2)

if btn1.button("📄 선택한 개별 시트 제작", type="secondary"):
    individual_options = dict(sheet_options)
    for k in ("final_legal_dong", "final_legal_jgu", "final_complex_dong", "final_complex_jgu"):
        individual_options[k] = False
    if download_mode.startswith("수식"):
        st.info("개별 시트 수식 모드에서는 행정동 피벗 시트가 원시 DATA 시트를 SUMIFS로 참조합니다. 원시 DATA 값 자체는 원천값으로 유지됩니다.")
    progress = st.progress(0, text="0% · 개별 DATA/피벗 워크북 생성 중")
    wb = export.build_integrated_workbook(
        raw_selected,
        selected_years=selected_years,
        sheet_options=individual_options,
        name_map=name_map,
        indicator_ids=indicator_ids,
        label_map=label_map,
        sector_of=sector_of,
        weight=weight_map,
        sign_map=sign_map,
        code_label_map=code_label_map,
        formula_mode=download_mode.startswith("수식"),
    )
    progress.progress(80, text="80% · 엑셀 저장 중")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    progress.progress(100, text="100% · 완료")
    st.success("선택한 개별 시트 제작이 완료되었습니다.")
    st.download_button(
        "⬇ 개별 시트 xlsx 다운로드",
        data=buf.getvalue(),
        file_name="쇠퇴진단_개별시트.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

if btn2.button("▶ 최종 법적+복합 진단 산출", type="primary"):
    if missing_buckets:
        st.error("최종 진단 산출에 필요한 원시 분류가 부족합니다: " + ", ".join(missing_buckets))
        st.dataframe(missing_bucket_message(missing_buckets), use_container_width=True)
        st.stop()

    progress = st.progress(0, text="0% · 행정동 복합지수 계산 중")
    dong_base = E.run(raw_selected, level="dong", grade_method=method, n_classes=int(n_classes))
    progress.progress(25, text="25% · 집계구 복합지수 계산 중")
    jgu_base = E.run(raw_selected, level="jgu", grade_method=method, n_classes=int(n_classes))
    progress.progress(45, text="45% · 추가지표 및 가중치 반영 중")
    custom_dong = CI.build_scores(custom_df, dong_base[0].index, "dong")
    custom_jgu = CI.build_scores(custom_df, jgu_base[0].index, "jgu")
    dong_scores = CI.combine_scores(dong_base[0], custom_dong)
    jgu_scores = CI.combine_scores(jgu_base[0], custom_jgu)
    dong_comp = CI.composite(dong_scores, indicator_ids, sector_of, weight_map)
    jgu_comp = CI.composite(jgu_scores, indicator_ids, sector_of, weight_map)
    dong_grades = E.assign_grades(dong_comp["종합"], int(n_classes), method)
    jgu_grades = E.assign_grades(jgu_comp["종합"], int(n_classes), method)
    dong = (dong_scores, dong_comp, dong_grades, dong_base[3])
    jgu = (jgu_scores, jgu_comp, jgu_grades, jgu_base[3])
    progress.progress(65, text="65% · 법적쇠퇴진단 계산 중")
    legal_dong = LG.run_legal(raw_selected, level="dong")
    legal_jgu = LG.run_legal(raw_selected, level="jgu")
    progress.progress(80, text="80% · 통합 엑셀 생성 중")
    wb = export.build_integrated_workbook(
        raw_selected,
        selected_years=selected_years,
        sheet_options=sheet_options,
        name_map=name_map,
        dong_res=dong[:3],
        jgu_res=jgu[:3],
        n_classes=int(n_classes),
        method=method,
        legal_dong=legal_dong,
        legal_jgu=legal_jgu,
        indicator_ids=indicator_ids,
        label_map=label_map,
        sector_of=sector_of,
        weight=weight_map,
        sign_map=sign_map,
        code_label_map=code_label_map,
        formula_mode=download_mode.startswith("수식"),
    )
    if download_mode.startswith("수식"):
        for sheet_name in ("전주시 복합쇠퇴지수(행정동)", "전주시 복합쇠퇴지수(집계구)"):
            if sheet_name in wb.sheetnames:
                sample = wb[sheet_name]["C3"].value
                if not str(sample).startswith("="):
                    st.error(f"{sheet_name} 수식 생성 확인 실패: C3 셀이 수식이 아닙니다.")
                    st.stop()
        if "복합지표값(행정동)" in wb.sheetnames and "전주시 법적쇠퇴진단(행정동)" in wb.sheetnames:
            sample = wb["복합지표값(행정동)"]["B2"].value
            if not str(sample).startswith("="):
                st.error("복합지표값(행정동)의 인구변화율이 법적진단 시트를 참조하는 수식으로 생성되지 않았습니다.")
                st.stop()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    progress.progress(100, text="100% · 완료")

    n_decl = int((legal_dong["쇠퇴지역"] == "o").sum())
    st.success(f"완료 — 행정동 {len(dong[0])}개 · 집계구 {len(jgu[0])}개 · 법적 쇠퇴지역(행정동) {n_decl}개")

    tab1, tab2 = st.tabs(["복합쇠퇴지수(행정동)", "법적쇠퇴진단(행정동)"])
    with tab1:
        scores, comp, grades = dong[:3]
        prev = pd.DataFrame({
            "종합점수": comp["종합"].round(2), "종합등급": grades,
            "인문사회": comp["인문사회"].round(2), "산업경제": comp["산업경제"].round(2),
            "물리환경": comp["물리환경"].round(2),
        }).sort_values("종합점수", ascending=False)
        st.dataframe(prev, use_container_width=True, height=420)
    with tab2:
        lg = legal_dong.copy()
        lg["인구증감률"] = lg["인구증감률"].round(2)
        lg["사업체증감률"] = lg["사업체증감률"].round(2)
        lg["노후건축물비율"] = lg["노후건축물비율"].round(2)
        st.dataframe(lg.sort_values("부합개수", ascending=False), use_container_width=True, height=420)

    st.download_button(
        "⬇ 통합 xlsx 다운로드 (개별 DATA + 법적 + 복합)",
        data=buf.getvalue(),
        file_name="쇠퇴진단_통합결과_수식.xlsx" if download_mode.startswith("수식") else "쇠퇴진단_통합결과_값.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    with st.expander("표준화 통계 (지표별 평균 / 모표준편차)"):
        stats = dong[3]
        st.dataframe(pd.DataFrame(stats).T[["mean", "std"]].round(4), use_container_width=True)
