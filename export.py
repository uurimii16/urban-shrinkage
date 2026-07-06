# -*- coding: utf-8 -*-
"""
export.py — 엔진 결과 → 멀티시트 xlsx (명세서 §2 구조)
=====================================================
build_workbook(raw, dong_res, jgu_res, ...) 로 openpyxl Workbook 을 만든다.

시트 구성:
  · 복합쇠퇴지수(행정동)   : 12지표×{값,Z,T,등급} + 부문3 + 종합 + 종합등급
  · 복합쇠퇴지수(집계구)   : 위와 동일(집계구 단위)
  · 원시_{분류}            : 연도|집계구|연도집계구|CODE|값|항목명|행정동코드 (6개)
  · 행정동집계_{분류}       : 행정동 × 연도 합계 + 총합 (6개)

지표별 등급은 각 지표 T점수에 engine.assign_grades 를 적용(엔진 기존 함수 재사용).
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import decline_engine as E
import config as C

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_SUB_FILL = PatternFill("solid", fgColor="D9E1F2")


# ── 코드 → 한글 항목명 (원시 long 시트 '항목명' 열용, best-effort) ──
_AGE_LABEL = {
    1: '0~4세', 2: '5~9세', 3: '10~14세', 4: '15~19세', 5: '20~24세',
    6: '25~29세', 7: '30~34세', 8: '35~39세', 9: '40~44세', 10: '45~49세',
    11: '50~54세', 12: '55~59세', 13: '60~64세', 14: '65~69세', 15: '70~74세',
    16: '75~79세', 17: '80~84세', 18: '85~89세', 19: '90~94세', 20: '95~99세',
    21: '100세 이상', 22: '나이미상',
}
_HO_YR_LABEL = {1: '1979 이전', 2: '1980~89', 3: '1990~99', 4: '2000~04',
                5: '2005~09', 6: '2010', 7: '2011', 8: '2012', 9: '2013', 10: '2014'}
_HO_AR_LABEL = {1: '20㎡ 이하', 2: '20~40㎡', 3: '40~60㎡', 4: '60~85㎡', 5: '85~100㎡',
                6: '100~130㎡', 7: '130~165㎡', 8: '165~230㎡', 9: '230㎡ 초과'}


def code_name(code: str) -> str:
    """항목코드 → 한글명(best-effort). 미상 코드는 코드 원문 반환."""
    if not isinstance(code, str):
        return str(code)
    try:
        no = int(code[-3:])
    except Exception:
        return code
    if code.startswith('to_in'):
        return '총인구'
    if code.startswith('to_fa'):
        return '총사업체수'
    if code.startswith('in_age'):
        prefix = '남' if 31 <= no <= 52 else ('여' if 61 <= no <= 82 else '전체')
        base = no if no <= 22 else (no - 30 if no <= 52 else no - 60)
        return f'{prefix} {_AGE_LABEL.get(base, code)}'
    if code.startswith('ho_yr'):
        return f'건축 {_HO_YR_LABEL.get(no, code)}'
    if code.startswith('ho_ar'):
        return f'연건평 {_HO_AR_LABEL.get(no, code)}'
    if code.startswith('cp_bem'):
        return f'종사자수(산업 {no:03d})'
    return code


# ── 지표별 등급 계산 ──
def indicator_grades(scores, n_classes=10, method='jenks', indicator_ids=None):
    """지표별 T점수 → 등급 DataFrame(unit index × 지표)."""
    indicator_ids = indicator_ids or C.IND_IDS
    out = {}
    for ind in indicator_ids:
        if (ind, 'T') in scores.columns:
            out[ind] = E.assign_grades(scores[(ind, 'T')], n_classes, method)
    return pd.DataFrame(out)


# ── 복합쇠퇴지수 시트 ──
def _write_composite(ws, scores, comp, grades, ind_grades, name_map, key_label,
                     indicator_ids=None, label_map=None):
    ci = ind_grades  # alias
    indicator_ids = indicator_ids or C.IND_IDS
    label_map = label_map or C.INDLABEL
    # 헤더 2줄: 지표명(병합) / 값·Z·T·등급
    hdr1 = [key_label, '명칭']
    hdr2 = ['', '']
    for ind in indicator_ids:
        hdr1 += [label_map.get(ind, ind), '', '', '']
        hdr2 += ['값', 'Z', 'T', '등급']
    hdr1 += ['인문사회', '산업경제', '물리환경', '종합', '종합등급']
    hdr2 += ['', '', '', '', '']
    ws.append(hdr1)
    ws.append(hdr2)
    for r in (1, 2):
        for cell in ws[r]:
            cell.fill = _HEADER_FILL if r == 1 else _SUB_FILL
            if r == 1:
                cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
    # 데이터
    for unit in scores.index:
        row = [unit, ((name_map or {}).get(unit) or (name_map or {}).get(str(unit)[:8], ''))]
        for ind in indicator_ids:
            row += [
                _num(scores.at[unit, (ind, '값')]),
                _num(scores.at[unit, (ind, 'Z')]),
                _num(scores.at[unit, (ind, 'T')]),
                _int(ci.at[unit, ind]) if ind in ci.columns else None,
            ]
        row += [
            _num(comp.at[unit, '인문사회']), _num(comp.at[unit, '산업경제']),
            _num(comp.at[unit, '물리환경']), _num(comp.at[unit, '종합']),
            _int(grades.get(unit)),
        ]
        ws.append(row)
    ws.freeze_panes = 'C3'


def _num(v):
    try:
        f = float(v)
        return round(f, 6)
    except Exception:
        return None


def _int(v):
    try:
        return int(v)
    except Exception:
        return None


# ── 원시 long 시트 ──
def _write_raw(ws, df, key='행정동코드'):
    ws.append(['연도', '집계구', '연도집계구', 'CODE', '값', '항목명', '행정동코드'])
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    dong = df['행정동코드'] if '행정동코드' in df.columns else df['집계구'].str[:8]
    for yr, gu, code, val, dcode in zip(df['연도'], df['집계구'], df['CODE'], df['값'], dong):
        ws.append([int(yr), str(gu), f'{int(yr)}{gu}', str(code), _num(val),
                   code_name(str(code)), str(dcode)])


# ── 행정동 집계 피벗 ──
def _write_dong_pivot(ws, df):
    d = df.copy()
    d['값'] = pd.to_numeric(d['값'], errors='coerce')
    piv = (d.groupby(['행정동코드', '연도'])['값'].sum(min_count=1)
             .unstack('연도').sort_index())
    piv['총합계'] = piv.sum(axis=1, min_count=1)
    years = list(piv.columns)
    ws.append(['행정동코드'] + [str(y) for y in years])
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    for dcode, r in piv.iterrows():
        ws.append([str(dcode)] + [_num(v) for v in r.values])


_RAW_SHEET_LABEL = {
    'to_in': '인구총괄', 'in_age': '성연령인구', 'to_fa': '총사업체수',
    'cp_bem': '종사자수', 'ho_yr': '건축연도주택', 'ho_ar': '연건평주택',
}


_LEGAL_COLS = ['인구증감률', '인구감소해당', '인구연속감소',
               '사업체증감률', '사업체감소해당', '사업체연속감소',
               '노후건축물비율', '물리노후해당', '부합개수', '쇠퇴지역']


def _write_legal(ws, legal_df, name_map, key_label):
    """법적쇠퇴진단 시트: 단위 + (행정동명) + 판정열."""
    hdr = [key_label, '명칭'] + _LEGAL_COLS
    ws.append(hdr)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal='center')
    for unit in legal_df.index:
        row = [unit, ((name_map or {}).get(unit) or (name_map or {}).get(str(unit)[:8], ''))]
        for c in _LEGAL_COLS:
            v = legal_df.at[unit, c]
            if c in ('인구증감률', '사업체증감률', '노후건축물비율'):
                row.append(_num(v))
            elif c == '부합개수':
                row.append(_int(v))
            else:
                row.append(str(v))
        ws.append(row)
    ws.freeze_panes = 'C2'


def build_workbook(raw, dong_res, jgu_res, name_map=None,
                   n_classes=10, method='jenks', include_raw=True,
                   legal_dong=None, legal_jgu=None,
                   indicator_ids=None, label_map=None):
    """엔진 결과 → openpyxl Workbook.
    dong_res/jgu_res = (scores, comp, grades) 튜플.
    legal_dong/legal_jgu = legal_engine.run_legal 결과 DataFrame(선택)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for level_label, lg, key_label in (
        ('법적쇠퇴진단(행정동)', legal_dong, '행정동코드'),
        ('법적쇠퇴진단(집계구)', legal_jgu, '집계구'),
    ):
        if lg is not None:
            ws = wb.create_sheet(level_label)
            nm = name_map if key_label == '행정동코드' else None
            _write_legal(ws, lg, nm, key_label)

    for level_label, res, key_label in (
        ('복합쇠퇴지수(행정동)', dong_res, '행정동코드'),
        ('복합쇠퇴지수(집계구)', jgu_res, '집계구'),
    ):
        scores, comp, grades = res
        ci = indicator_grades(scores, n_classes, method, indicator_ids)
        ws = wb.create_sheet(level_label)
        nm = name_map if key_label == '행정동코드' else None
        _write_composite(ws, scores, comp, grades, ci, nm, key_label, indicator_ids, label_map)

    if include_raw:
        for key, df in raw.items():
            ws = wb.create_sheet(f'원시_{_RAW_SHEET_LABEL.get(key, key)}')
            _write_raw(ws, df)
        for key, df in raw.items():
            ws = wb.create_sheet(f'행정동집계_{_RAW_SHEET_LABEL.get(key, key)}')
            _write_dong_pivot(ws, df)

    return wb


def build_integrated_workbook(raw, *, selected_years=None, sheet_options=None,
                              name_map=None, dong_res=None, jgu_res=None,
                              legal_dong=None, legal_jgu=None,
                              n_classes=10, method='jenks',
                              indicator_ids=None, label_map=None,
                              sector_of=None, weight=None,
                              sign_map=None, code_label_map=None,
                              formula_mode=False, pivot_level="dong", final_only=False):
    """원시 DATA 개별 시트 + 피벗 + 최종 진단을 한 xlsx로 병합한다."""
    import sheet_builder
    return sheet_builder.build_integrated_workbook(
        raw,
        selected_years=selected_years,
        sheet_options=sheet_options,
        name_map=name_map,
        dong_res=dong_res,
        jgu_res=jgu_res,
        legal_dong=legal_dong,
        legal_jgu=legal_jgu,
        n_classes=n_classes,
        method=method,
        indicator_ids=indicator_ids,
        label_map=label_map,
        sector_of=sector_of,
        weight=weight,
        sign_map=sign_map,
        code_label_map=code_label_map,
        formula_mode=formula_mode,
        pivot_level=pivot_level,
        final_only=final_only,
    )
