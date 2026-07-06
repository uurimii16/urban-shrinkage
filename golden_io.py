# -*- coding: utf-8 -*-
"""
golden_io.py — 골든본/정제 xlsx 에서 6개 원시 long 추출 (부작용 없음)
====================================================================
verify_jeonju.py 의 추출 로직을 Streamlit 등에서 재사용하기 위해
stdout 조작 같은 import 부작용 없이 분리한 모듈.
"""
import numpy as np
import pandas as pd
import openpyxl

# (시트, 연도열, 집계구열, CODE열, 값열, 접두어) — 0-based 열 인덱스
RAW_SPECS = {
    'to_in':  ('법적DATA_1.인구총괄(총인구)(수정)',       0, 1, 3, 4, 'to_in'),
    'to_fa':  ('복합DATA_2.산업경제-총괄사업체수(수정)',   0, 1, 2, 3, 'to_fa'),
    'cp_bem': ('복합DATA_2.산업경제-종사자수(수정)',       0, 1, 2, 5, 'cp_bem'),
    'in_age': ('복합 DATA_1.인문사회(수정)',              0, 1, 2, 3, 'in_age'),
    'ho_ar':  ('복합DATA_3.물리환경-소형주택비율(수정)',   0, 1, 2, 5, 'ho_ar'),
    'ho_yr':  ('복합DATA_3.물리환경-주택건축물비율(수정)', 0, 1, 2, 4, 'ho_yr'),
}

REQUIRED_SHEETS = [spec[0] for spec in RAW_SPECS.values()]


def to_num(x):
    """숫자 아니면 NaN(N/A·공백 등은 결측)."""
    if x is None:
        return np.nan
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if s in ('', 'N/A', 'NA', 'na', 'n/a', '#N/A'):
        return np.nan
    try:
        return float(s)
    except Exception:
        return np.nan


def _clean_gu(gu):
    s = str(gu).strip()
    return s[:-2] if s.endswith('.0') else s


def load_raw_from_workbook(wb, mapping=None):
    """열린 openpyxl Workbook → dict[str, DataFrame]
    (컬럼: 연도|집계구|CODE|값|행정동코드).
    mapping: {집계구코드(14): 행정동코드(8)} — 없으면 집계구[:8]."""
    raw = {}
    for key, (sh, ciy, cig, cic, civ, pref) in RAW_SPECS.items():
        ws = wb[sh]
        maxc = max(ciy, cig, cic, civ) + 1
        rows = []
        for r in ws.iter_rows(min_row=2, max_col=maxc, values_only=True):
            code = r[cic]
            if code is None or not str(code).startswith(pref):
                continue
            yr, gu, val = r[ciy], r[cig], r[civ]
            if yr is None or gu is None:
                continue
            try:
                yr = int(yr)
            except Exception:
                continue
            rows.append((yr, _clean_gu(gu), str(code).strip(), to_num(val)))
        df = pd.DataFrame(rows, columns=['연도', '집계구', 'CODE', '값'])
        if mapping:
            df['행정동코드'] = df['집계구'].map(lambda g: mapping.get(g, g[:8]))
        else:
            df['행정동코드'] = df['집계구'].str[:8]
        raw[key] = df
    return raw


def load_raw_from_path(path, mapping=None):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return load_raw_from_workbook(wb, mapping)
    finally:
        wb.close()


def missing_sheets(wb):
    """업로드 파일에 필요한 (수정) 시트가 다 있는지 검사."""
    have = set(wb.sheetnames)
    return [s for s in REQUIRED_SHEETS if s not in have]
