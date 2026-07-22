# -*- coding: utf-8 -*-
"""template_export.py — 차장님 정본 양식(복합쇠퇴진단.xlsx) 그대로 산출.
9시트(집계구 단위): 계산방법 / 1 법적쇠퇴진단 종합 / 2 복합쇠퇴진단 종합 / 원시 6시트.
엔진 계산은 decline_engine·legal_engine 그대로 재사용, 출력 양식만 이 모듈이 담당.
지금은 계산방법 + 복합쇠퇴진단 종합 을 우선 구현(검증용). 법적·원시시트는 후속.
"""
import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side
import pandas as pd
import numpy as np

import decline_engine as E

# 12지표: (id, 헤더(R1), 부문, 방향부호, 기본최종가중치) — 계산방법 E24~E35에 대응
TEMPLATE_INDICATORS = [
    ('인구변화율',      '인구변화율\n(30년간 최다인구수 연도대비)',       '인문사회', -10, 0.220),
    ('노년부양비',      '2024년 노년부양비',                           '인문사회',  10, 0.058),
    ('경제활동인구비율','2024년 경제활동인구비율',                      '인문사회', -10, 0.113),
    ('소멸위험지수',    '2024년 소멸위험지수',                          '인문사회', -10, 0.009),
    ('총사업체수증감률','총사업체수 증감률\n(10년간 최다 사업체수 대비)', '산업경제', -10, 0.021),
    ('총종사자수증감률','총 종사자수 증감률\n(10년간 최다 종사자수 대비)','산업경제', -10, 0.028),
    ('제조업증감률',    '제조업 종사자수 증감률',                       '산업경제', -10, 0.057),
    ('고차산업증감률',  '고차산업종사자수 증감률',                      '산업경제', -10, 0.066),
    ('도소매증감률',    '도소매종사자수 증감률',                        '산업경제', -10, 0.024),
    ('음식숙박증감률',  '음식숙박업 종사자수 증감률',                    '산업경제', -10, 0.004),
    ('노후건축물비율',  '노후건축물비율',                              '물리환경',  10, 0.250),
    ('소형주택비율',    '19평 이하 소형주택비율',                       '물리환경',  10, 0.150),
]
SECTORS = ['인문사회', '산업경제', '물리환경']
CALC_SHEET = '계산방법(수식·알고리즘)'
# 지표 id → 정본 양식 헤더/부문/방향/가중치. 앱 cfg에 없는 값은 여기서 폴백.
_TEMPLATE_HDR = {iid: hdr for iid, hdr, *_ in TEMPLATE_INDICATORS}
_TEMPLATE_SEC = {iid: sec for iid, _h, sec, _s, _w in TEMPLATE_INDICATORS}
_TEMPLATE_SIGN = {iid: sign for iid, _h, _sec, sign, _w in TEMPLATE_INDICATORS}
_TEMPLATE_W = {iid: w for iid, _h, _sec, _s, w in TEMPLATE_INDICATORS}
# 전국 행정동명표(load_admin_names 기본 경로) — 이 파일 기준 상대경로.
DEFAULT_ADMIN_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                  '행정구역코드', '행정구역코드_전국.xlsx')
_THIN = Side(style='thin', color='B0B0B0')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HDR_FONT = Font(bold=True, size=9)
_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)


def indicators_from_cfg(cfg):
    """app_v2의 ss.cfg → 정본 양식용 indicators 리스트.
    반환: [(id, 헤더, 부문, 방향부호, 최종가중치), ...]  (순서 = cfg['indicator_ids']).
    · 방향부호·최종가중치·부문 = ③설정에서 사용자가 정한 값(sign_map/final_weight/sector_of).
    · 헤더 = 정본 양식 헤더 우선, 없으면(커스텀 지표) cfg의 label_map.
    cfg가 비어 있으면 정본 기본 12지표를 그대로 돌려준다."""
    ids = list(cfg.get('indicator_ids') or [i[0] for i in TEMPLATE_INDICATORS])
    sign_map = cfg.get('sign_map', {})
    weight = cfg.get('final_weight', {})
    sector_of = cfg.get('sector_of', {})
    label_map = cfg.get('label_map', {})
    out = []
    for iid in ids:
        hdr = _TEMPLATE_HDR.get(iid) or label_map.get(iid, iid)
        sec = sector_of.get(iid) or _TEMPLATE_SEC.get(iid) or SECTORS[0]
        # cfg에 값이 있으면(0 포함) 그대로, 아예 없으면 정본 기본값으로 폴백
        sign = int(sign_map[iid]) if iid in sign_map else _TEMPLATE_SIGN.get(iid, -10)
        w = float(weight[iid]) if iid in weight else _TEMPLATE_W.get(iid, 0.0)
        out.append((iid, hdr, sec, sign, w))
    return out


def load_admin_names(path):
    """행정구역코드_전국.xlsx → (행정동코드8→행정동명, 시군구코드5→시군구명)."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb['전체'] if '전체' in wb.sheetnames else wb[wb.sheetnames[0]]
    dong, sgg = {}, {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        # 연번, 시도코드, 시도명, 시군구코드, 시군구명, 행정동코드, 행정동명
        sg, sgn, dc, dn = (str(row[3] or '').strip(), str(row[4] or '').strip(),
                           str(row[5] or '').strip(), str(row[6] or '').strip())
        if dc:
            dong[dc] = dn
        if sg:
            sgg[sg] = sgn
    wb.close()
    return dong, sgg


def compute_values(raw_sub, ids=None, key='집계구'):
    """집계구(또는 행정동)별 지표 값 DataFrame. ids 순서·구성으로 정렬.
    ids=None이면 정본 기본 12지표. 엔진이 못 만드는 지표(커스텀 등)는 NaN 열로 남고,
    종합시트의 Z/T 수식이 빈칸을 0으로 처리하므로 구조는 그대로 유지된다."""
    if ids is None:
        ids = [i[0] for i in TEMPLATE_INDICATORS]
    df = E.derive_indicators(raw_sub, key=key)            # 11지표(config)
    if '소형주택비율' in ids and '소형주택비율' not in df.columns and 'ho_ar' in raw_sub:
        df = df.copy()                                     # 소형주택 되살림(요청 시)
        df['소형주택비율'] = E.derive_small_housing_ratio(raw_sub['ho_ar'], key)
    return df.reindex(columns=list(ids))


def values_from_scores(scores, ids):
    """앱이 이미 산출한 scores(멀티인덱스 (지표,'값'/'Z'/'T')) → 정본 지표값 DataFrame.
    기본+커스텀+계산식 지표 모두 (지표,'값') 열을 갖고 있으므로 그대로 뽑아 쓴다.
    scores에 없는 지표는 NaN(빈칸)으로 남고 종합시트 함수가 0으로 처리."""
    cols = {}
    for iid in ids:
        cols[iid] = scores[(iid, '값')] if (iid, '값') in scores.columns else np.nan
    return pd.DataFrame(cols, index=scores.index)


# ── 계산방법 시트 (원본 텍스트·산식 그대로) ─────────────────────────
_CALC_DOC = {
    1: '① 표준화 (Z점수 · T점수)',
    2: 'Z = (지표값 − 평균) ÷ 모표준편차(STDEV.P)',
    3: '   · 평균·표준편차는 그 단위 집합(행정동 또는 집계구) 전체에서 결측(N/A) 제외하고 계산',
    4: '   · 엑셀 등가:  평균=AVERAGE(범위) ,  모표준편차=STDEV.P(범위)',
    5: '   · 표준편차가 0이면(모두 같은 값) Z = 0 으로 처리',
    6: 'T = Z × 방향부호 + 50',
    7: '   · 방향부호(±10): 값이 클수록 쇠퇴가 심하면 +10, 값이 클수록 양호하면 −10',
    8: '   · 즉 일반 T점수(=Z×10+50)에 쇠퇴 방향을 부호로 합친 형태 (평균 50 기준)',
    10: '② 가중치 · 부문점수 · 종합',
    11: '최종가중치 = (부문비율 ÷ 100) × (부문 내부비율 ÷ 100)',
    12: '   · 부문비율 : 인문사회/산업경제/물리환경 3부문 사이의 배분(합 100%)',
    13: '   · 내부비율 : 한 부문 안에서 지표들 사이의 배분(부문별 합 100%)',
    14: '부문점수 = Σ ( 지표T × 그 지표 최종가중치 )   (그 부문에 속한 지표만 합산)',
    15: '종합점수 = 인문사회 + 산업경제 + 물리환경  (세 부문점수의 합)',
    16: '   · 엑셀 등가:  가중T = T×가중치 ,  부문 = SUM(부문 가중T들) ,  종합 = SUM(3부문)',
    18: '③ 등급',
    19: '분류 방식 : Natural Breaks(Jenks) · 10등급',
    20: '종합점수가 클수록 쇠퇴가 심함 → 1등급(가장 쇠퇴) … 큰 등급숫자일수록 양호',
    22: '④ 지표별 산식 · 방향 · 최종가중치',
}
# 지표 id → (계산방법 ④표 지표명, 부문, 산식설명). 복합종합은 E열(가중치)만 참조 → A~C는 표기용.
_CALC_ROWS = {
    '인구변화율':      ('인구변화율(법적)', '인문사회', '(기준연도 총인구 − 전체연도 중 최다 총인구) ÷ 최다 총인구 × 100'),
    '노년부양비':      ('노년부양비', '인문사회', '65세 이상 인구 ÷ 15~64세 인구 × 100'),
    '경제활동인구비율': ('경제활동인구비율', '인문사회', '15~64세 인구 ÷ 총인구 × 100'),
    '소멸위험지수':    ('소멸위험지수', '인문사회', '20~39세 여성 인구 ÷ 65세 이상 인구'),
    '총사업체수증감률': ('총사업체수 증감률', '산업경제', '(기준연도 총사업체수 − 전체연도 최다) ÷ 최다 × 100'),
    '총종사자수증감률': ('총종사자수 증감률', '산업경제', '(기준연도 총종사자수 − 전체연도 최다) ÷ 최다 × 100'),
    '제조업증감률':    ('제조업 종사자수 증감률', '산업경제', '(기준연도 제조업 종사자수 − 최다) ÷ 최다 × 100  (차수별 산업코드 합)'),
    '고차산업증감률':  ('고차산업 종사자수 증감률', '산업경제', '(기준연도 고차산업 종사자수 − 최다) ÷ 최다 × 100  (차수별 산업코드 합)'),
    '도소매증감률':    ('도소매 종사자수 증감률', '산업경제', '(기준연도 도소매 종사자수 − 최다) ÷ 최다 × 100'),
    '음식숙박증감률':  ('음식숙박 종사자수 증감률', '산업경제', '(기준연도 음식숙박 종사자수 − 최다) ÷ 최다 × 100'),
    '노후건축물비율':  ('노후건축물비율', '물리환경', '2004년 이전 건축 주택수 ÷ 전체 주택수 × 100'),
    '소형주택비율':    ('19평이하 소형주택비율', '물리환경', '연면적 60㎡이하 주택수 ÷ 전체 주택수 × 100'),
}


def write_calc_method(ws, indicators=TEMPLATE_INDICATORS):
    """알고리즘 설명 + ④지표별 산식표. 방향부호(D24~)·최종가중치(E24~)는 여기서 주입."""
    for r, line in _CALC_DOC.items():
        ws.cell(r, 1).value = line
    hdr = ['지표', '부문', '산식', '방향부호', '최종가중치']
    for c, h in enumerate(hdr, start=1):
        cell = ws.cell(23, c); cell.value = h; cell.font = _HDR_FONT; cell.border = _BORDER
    for i, (iid, _hdr, sec, sign, w) in enumerate(indicators):
        r = 24 + i
        name, csec, formula = _CALC_ROWS.get(iid, (_hdr or iid, sec, ''))
        ws.cell(r, 1).value = name
        ws.cell(r, 2).value = sec or csec          # ③설정 부문 우선
        ws.cell(r, 3).value = formula              # 산식 설명(원본 그대로)
        ws.cell(r, 4).value = sign                 # D열 방향부호
        ws.cell(r, 5).value = w                    # E열 최종가중치 ← 복합종합이 참조
        for c in range(1, 6):
            ws.cell(r, c).border = _BORDER
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['C'].width = 46


# ── 복합쇠퇴진단 종합 시트 ─────────────────────────────────────
def write_composite(ws, codes, values, dong_names, indicators=TEMPLATE_INDICATORS,
                    calc_sheet=CALC_SHEET, grades=None):
    """codes: 집계구코드 리스트(정렬). values: df(집계구 index × 12지표). dong_names: 8자리→명.
    grades: codes 순서의 Natural Break(Jenks) 등급 리스트(없으면 AW 비움)."""
    n = len(codes)
    N = 2 + n                              # 데이터 마지막 행(3..N)
    q = lambda s: f"'{s}'"                 # 시트명 인용

    # ── R1/R2 헤더 ──
    ws.cell(1, 1).value = '행정동'; ws.cell(1, 2).value = '집계구'; ws.cell(1, 3).value = '행정동명'
    for i, (iid, hdr, sec, sign, w) in enumerate(indicators):
        vcol = 4 + i * 3
        ws.cell(1, vcol).value = hdr
        ws.merge_cells(start_row=1, start_column=vcol, end_row=1, end_column=vcol + 2)
        ws.cell(2, vcol).value = '지표값'; ws.cell(2, vcol + 1).value = 'Z점수'; ws.cell(2, vcol + 2).value = 'T점수'
    base = 4 + len(indicators) * 3          # 40 (AN)
    sec_cols = {}
    for s_i, sec in enumerate(SECTORS):
        wcol = base + s_i * 2
        ws.cell(1, wcol).value = sec
        ws.merge_cells(start_row=1, start_column=wcol, end_row=1, end_column=wcol + 1)
        ws.cell(2, wcol).value = '가중치적용'; ws.cell(2, wcol + 1).value = '에러수정'
        sec_cols[sec] = wcol
    total_col = base + len(SECTORS) * 2      # 46 (AT)
    ws.cell(1, total_col).value = '종합'; ws.cell(2, total_col).value = '가중치적용'
    ws.cell(1, total_col + 1).value = '등급'
    ws.cell(2, total_col + 1).value = '등분위'; ws.cell(2, total_col + 2).value = '등간격'
    ws.cell(2, total_col + 3).value = 'Natural Break'

    # 부문별 지표 인덱스
    sec_idx = {s: [i for i, ind in enumerate(indicators) if ind[2] == s] for s in SECTORS}

    # ── 데이터 행(R3~) ──
    for ridx, code in enumerate(codes):
        r = 3 + ridx
        dong8 = str(code)[:8]
        ws.cell(r, 1).value = dong8
        ws.cell(r, 2).value = str(code)
        ws.cell(r, 3).value = dong_names.get(dong8, '')
        # 지표값 + Z/T 수식
        for i, (iid, hdr, sec, sign, w) in enumerate(indicators):
            vcol = 4 + i * 3
            V = get_column_letter(vcol); Z = get_column_letter(vcol + 1); Tc = get_column_letter(vcol + 2)
            val = values.iloc[ridx][iid] if iid in values.columns else None
            ws.cell(r, vcol).value = (None if val is None or (isinstance(val, float) and np.isnan(val)) else float(val))
            rng = f"{V}$3:{V}${N}"
            if i == 0:
                ws.cell(r, vcol + 1).value = f"=({V}{r}-AVERAGE({rng}))/_xlfn.STDEV.P({rng})"
                ws.cell(r, vcol + 2).value = f"=({Z}{r}*{sign})+50"
            else:
                ws.cell(r, vcol + 1).value = f'=IF({V}{r}="",0,(({V}{r}-AVERAGE({rng}))/_xlfn.STDEV.P({rng})))'
                ws.cell(r, vcol + 2).value = f'=IF({V}{r}="",0,({Z}{r}*{sign})+50)'
        # 부문 가중치적용 / 에러수정
        for sec in SECTORS:
            wcol = sec_cols[sec]
            terms, terms_err = [], []
            for i in sec_idx[sec]:
                Tc = get_column_letter(4 + i * 3 + 2)
                ecell = f"{q(calc_sheet)}!$E${24 + i}"
                terms.append(f"({Tc}{r}*{ecell})")
                terms_err.append(f"IFERROR({Tc}{r}*{ecell}, 0)")
            ws.cell(r, wcol).value = "=SUM(" + ",".join(terms) + ")"
            ws.cell(r, wcol + 1).value = "=SUM(" + ",".join(terms_err) + ")"
        # 종합 = 부문 에러수정들의 합
        err_cols = [get_column_letter(sec_cols[s] + 1) + str(r) for s in SECTORS]
        ws.cell(r, total_col).value = "=SUM(" + ",".join(err_cols) + ")"
        # 등급(종합 AT열 기준)
        AT = get_column_letter(total_col)
        rngT = f"$AT$3:${AT}${N}".replace('AT', AT)  # AT range
        rngT = f"${AT}$3:${AT}${N}"
        ws.cell(r, total_col + 1).value = (
            f"=MIN(10,ROUNDUP(RANK({AT}{r},{rngT},0)/(COUNT({rngT})/10),0))")      # 등분위
        ws.cell(r, total_col + 2).value = (
            f"=MAX(1,MIN(10,ROUNDUP((MAX({rngT})-{AT}{r})/((MAX({rngT})-MIN({rngT}))/10),0)))")  # 등간격
        # Natural Break(Jenks): 엑셀 Jenks 함수가 없어 파이썬 경계 계산값을 정적 등급으로 기록.
        ws.cell(r, total_col + 3).value = (int(grades[ridx]) if grades is not None
                                           and ridx < len(grades) else None)

    # 표시서식 0.00 (지표값·Z·T·부문·종합)
    for r in range(3, N + 1):
        for c in range(4, total_col + 1):
            cell = ws.cell(r, c)
            if isinstance(cell.value, (int, float)) or (isinstance(cell.value, str) and cell.value.startswith('=')):
                cell.number_format = '0.00'
    # 헤더 스타일
    for r in (1, 2):
        for c in range(1, total_col + 4):
            cell = ws.cell(r, c); cell.font = _HDR_FONT; cell.alignment = _CENTER
    ws.freeze_panes = 'D3'
    return N


# ══════════════════════════════════════════════════════════════════
#  원시 6시트 + 법적종합 시트 (원본 9시트 재현)
# ══════════════════════════════════════════════════════════════════
_cl = get_column_letter
# CODE→항목명(설명열 · VLOOKUP 미참조 → 형태용). 버킷별 패턴.
_HO_YR_BANDS = ['1979년 이전', '1980년~1989년', '1990년~1999년', '2000년~2004년', '2005년~2009년']
_HO_AR_BANDS = {1: '20㎡이하(호)', 2: '20㎡~40㎡이하(호)', 3: '40㎡~60㎡이하(호)', 4: '60㎡~85㎡이하(호)',
                5: '85㎡~100㎡이하(호)', 6: '100㎡~130㎡이하(호)', 7: '130㎡~165㎡이하(호)',
                8: '165㎡~230㎡이하(호)', 9: '230㎡초과(호)'}
# 원시 시트 정확 명칭 + 버킷 대응
RAW_SHEETS = [
    ('1.1. 총인구',        'to_in'),
    ('1.2. 총사업체',      'to_fa'),
    ('1.3. 건축년도별주택', 'ho_yr'),
    ('2.1. 성연령별인구',   'in_age'),
    ('2.2. (산업별)종사자', 'cp_bem'),
    ('2.3. 연건평별주택',   'ho_ar'),
]
LEGAL_SHEET = '1 법적쇠퇴진단 종합'


def _code_num(code):
    try:
        return int(str(code)[-3:])
    except (TypeError, ValueError):
        return 0


def _ho_yr_label(n):
    """건축연도 밴드 라벨. 1~5=고정, 6~=2010+(n-6)년."""
    if 1 <= n <= 5:
        return _HO_YR_BANDS[n - 1]
    return f'{2010 + (n - 6)}년'


def _item_name(bucket, code):
    n = _code_num(code)
    if bucket == 'to_in':
        return '총인구'
    if bucket == 'to_fa':
        return '총사업체수'
    if bucket == 'cp_bem':
        return f'종사자수(산업 {str(code)[-3:]})'
    if bucket == 'ho_yr':
        return _ho_yr_label(n)
    if bucket == 'ho_ar':
        return _HO_AR_BANDS.get(n, f'연면적코드 {n}')
    if bucket == 'in_age':
        if 31 <= n <= 52:
            pre, k = '남 ', n - 30
        elif 61 <= n <= 82:
            pre, k = '여 ', n - 60
        else:
            pre, k = '', n
        if k == 22:
            return pre + '연령미상'
        if k == 21:                       # 최고 밴드: 원본은 '100세이상'(상한 없음)
            return pre + '100세이상'
        if k == 1:                        # 최저 밴드: 원본은 '4세이하'(하한 없음)
            return pre + '4세이하'
        lo = (k - 1) * 5
        return f'{pre}{lo}세이상~{lo + 4}세이하'
    return ''


def _val_cell(v):
    """원시 값 → 셀값. 결측(NaN)은 'N/A' 텍스트(원본 형태)."""
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return 'N/A'
    if isinstance(v, str):
        return v
    return float(v)


def write_raw_sheet(ws, df, bucket):
    """원시 롱데이터 → 7열(연도·집계구·연도집계구·CODE·값·항목명·행정동코드). 데이터 행수 반환.
    ho_yr는 호출측에서 피벗(I~AG)을 추가로 붙인다."""
    hdr = ['연도', '집계구', '연도집계구', 'CODE', '값', '항목명', '행정동코드']
    ws.append(hdr)
    for c in range(1, 8):
        ws.cell(1, c).font = _HDR_FONT
    if df is None or len(df) == 0:
        return 0
    d = df.copy()
    d['집계구'] = d['집계구'].astype(str)
    d['CODE'] = d['CODE'].astype(str)
    d['연도'] = pd.to_numeric(d['연도'], errors='coerce').astype('Int64')
    d = d.sort_values(['연도', '집계구', 'CODE']).reset_index(drop=True)
    if '행정동코드' not in d.columns:
        d['행정동코드'] = d['집계구'].str[:8]
    # 항목명: 코드 단위로 한 번만 계산 → 매핑
    name_by_code = {c: _item_name(bucket, c) for c in d['CODE'].unique()}
    yrs = d['연도'].tolist(); gus = d['집계구'].tolist(); codes = d['CODE'].tolist()
    vals = d['값'].tolist(); dongs = d['행정동코드'].astype(str).tolist()
    for i in range(len(d)):
        y = int(yrs[i]); g = gus[i]
        ws.append([y, g, f'{y}{g}', codes[i], _val_cell(vals[i]), name_by_code[codes[i]], dongs[i]])
    return len(d)


def _ho_yr_pivot(df):
    """ho_yr 롱 → 집계구별 건축연도밴드 피벗. 반환: (집계구리스트, 밴드행렬(dict 집계구→{n:합}), 최대밴드)."""
    if df is None or len(df) == 0:
        return [], {}, 0
    d = df.copy()
    d['집계구'] = d['집계구'].astype(str)
    d['_n'] = d['CODE'].map(_code_num)
    d['값'] = pd.to_numeric(d['값'], errors='coerce').fillna(0.0)
    piv = {}
    maxn = int(d['_n'].max()) if len(d) else 0
    for gu, sub in d.groupby('집계구'):
        piv[gu] = sub.groupby('_n')['값'].sum().to_dict()
    return sorted(piv.keys()), piv, max(maxn, 20)


def write_ho_yr_pivot(ws, df, start_col=9):
    """1.3 시트 오른쪽에 피벗(I~AG) 부착. 반환: (집계구 피벗행수, 캐시맵)."""
    codes, piv, maxn = _ho_yr_pivot(df)
    nb = 20                                     # 밴드 20개(1979이전~2024년)
    Icol = start_col                            # I=9
    band0 = Icol + 1                            # J=10
    total_col = band0 + nb                      # AD=30
    aecol = total_col + 1                       # AE=31 (집계구)
    afcol = aecol + 1                           # AF=32 (노후=SUM(J:M))
    agcol = afcol + 1                           # AG=33 (전체=SUM(J:AC))
    ws.cell(1, Icol).value = '합계 : 값2'; ws.cell(1, band0).value = '열 레이블'
    ws.cell(2, Icol).value = '행 레이블'
    for k in range(nb):
        ws.cell(2, band0 + k).value = _ho_yr_label(k + 1)
    ws.cell(2, total_col).value = '총합계'
    ws.cell(2, aecol).value = '집계구'; ws.cell(2, afcol).value = '노후건축물'; ws.cell(2, agcol).value = '전체건축물'
    for c in range(Icol, agcol + 1):
        ws.cell(2, c).font = _HDR_FONT
    cache = {}
    Jl = _cl(band0); Ml = _cl(band0 + 3); ACl = _cl(band0 + nb - 1)
    for i, gu in enumerate(codes):
        r = 3 + i
        bands = piv[gu]
        ws.cell(r, Icol).value = gu
        tot = 0.0
        for k in range(nb):
            v = bands.get(k + 1, 0)
            if v:
                ws.cell(r, band0 + k).value = float(v)
                tot += float(v)
        ws.cell(r, total_col).value = float(tot)
        ws.cell(r, aecol).value = gu
        old = sum(float(bands.get(k, 0)) for k in (1, 2, 3, 4))     # 노후=1979이전~2004
        ws.cell(r, afcol).value = f'=SUM({Jl}{r}:{Ml}{r})'
        ws.cell(r, agcol).value = f'=SUM({Jl}{r}:{ACl}{r})'
        cache[f'{_cl(afcol)}{r}'] = repr(float(old))
        cache[f'{_cl(agcol)}{r}'] = repr(float(tot))
    return len(codes), cache


def write_legal(ws, codes, dong_names, raw_sub, npop, nbiz, npivot):
    """1 법적쇠퇴진단 종합 — 원시시트 VLOOKUP + 증감률 + 3년연속감소 + 노후비율 + 부합/쇠퇴.
    수식 그대로. 반환: 캐시맵(계산값 주입용)."""
    pop = raw_sub.get('to_in'); biz = raw_sub.get('to_fa')
    pop_years = sorted({int(y) for y in pd.to_numeric(pop['연도'], errors='coerce').dropna()}) if pop is not None and len(pop) else []
    biz_years = sorted({int(y) for y in pd.to_numeric(biz['연도'], errors='coerce').dropna()}) if biz is not None and len(biz) else []
    if pop_years:                       # 인구: 최근 30년 표본
        pop_years = [y for y in pop_years if y >= pop_years[-1] - 29]
    biz_years = biz_years[-10:]         # 사업체: 최근 10년(원본 U~AD = 2014~2023)
    npy, nby = len(pop_years), len(biz_years)
    # 열 배치
    pop_s = 5; pop_e = 4 + npy
    Rc = pop_e + 1; Sc = pop_e + 2; Tc = pop_e + 3
    biz_s = pop_e + 4; biz_e = biz_s + nby - 1
    AEc = biz_e + 1; AFc = biz_e + 2; AGc = biz_e + 3
    AHc = biz_e + 4; AIc = biz_e + 5; AJc = biz_e + 6; AKc = biz_e + 7
    ALc = biz_e + 8; AMc = biz_e + 9
    L = _cl
    # ── 헤더 R1~R3 ──
    ws.cell(1, 1).value = '연번'; ws.cell(1, 2).value = '행정동'; ws.cell(1, 3).value = '집계구'; ws.cell(1, 4).value = '행정동명'
    ws.cell(1, pop_s).value = '최근 30년간 인구수가 가장 많았던 연도 대비 증감률'
    ws.cell(1, Tc).value = '최근 5년간\n3년이상\n연속 감소'
    ws.cell(1, biz_s).value = '최근 10년간 사업체가 가장 많았던 연도 대비 증감률'
    ws.cell(1, AGc).value = '최근 5년간\n3년이상\n연속 감소'
    ws.cell(1, AHc).value = '사용승인 20년 이상 노후건축물 비율'
    ws.cell(1, ALc).value = '부합개수'; ws.cell(1, AMc).value = '쇠퇴지역'
    ws.cell(2, pop_s).value = '연도별 총인구'; ws.cell(2, Rc).value = '증감률\n(%)'; ws.cell(2, Sc).value = '해당여부'
    ws.cell(2, biz_s).value = '연도별 총사업체'; ws.cell(2, AEc).value = '증감률\n(%)'; ws.cell(2, AFc).value = '해당여부'
    ws.cell(2, AHc).value = '노후건축물'; ws.cell(2, AIc).value = '전체건축물'; ws.cell(2, AJc).value = '비율(%)'; ws.cell(2, AKc).value = '해당여부'
    for k, y in enumerate(pop_years):
        ws.cell(3, pop_s + k).value = int(y)
    for k, y in enumerate(biz_years):
        ws.cell(3, biz_s + k).value = int(y)
    for c in range(1, AMc + 1):
        for rr in (1, 2, 3):
            ws.cell(rr, c).font = _HDR_FONT; ws.cell(rr, c).alignment = _CENTER
    # ── 룩업 값(캐시 계산용) ──
    def lut(df):
        if df is None or len(df) == 0:
            return {}
        t = df.copy(); t['집계구'] = t['집계구'].astype(str)
        t['연도'] = pd.to_numeric(t['연도'], errors='coerce')
        t['값'] = pd.to_numeric(t['값'], errors='coerce')
        return {(int(y), g): v for y, g, v in zip(t['연도'], t['집계구'], t['값']) if pd.notna(y)}
    pop_lut, biz_lut = lut(pop), lut(biz)
    _, piv, _ = _ho_yr_pivot(raw_sub.get('ho_yr'))
    cache = {}
    E_ = L(pop_s); PE_ = L(pop_e); U_ = L(biz_s); BE_ = L(biz_e)
    pop_sh = "'1.1. 총인구'"; biz_sh = "'1.2. 총사업체'"; ho_sh = "'1.3. 건축년도별주택'"
    pop_rng = f"{pop_sh}!$C$2:$E${1 + npop}"
    biz_rng = f"{biz_sh}!$C$2:$E${1 + nbiz}"
    ho_rng = f"{ho_sh}!$AE$3:$AG${2 + npivot}"

    def seq_decline(vals5):
        """마지막 5개 값 [M,N,O,P,Q]에서 3연속 감소(원본 T식과 동일)."""
        if len(vals5) < 5 or any(v is None or (isinstance(v, float) and np.isnan(v)) for v in vals5):
            m, n_, o, p, q = ([float('nan')] * (5 - len(vals5)) + list(vals5))[-5:] if vals5 else [float('nan')] * 5
        else:
            m, n_, o, p, q = vals5
        def lt(a, b):
            return (not np.isnan(a)) and (not np.isnan(b)) and a < b
        w1 = lt(n_, m) and lt(o, n_) and lt(p, o)
        w2 = lt(o, n_) and lt(p, o) and lt(q, p)
        return 'O' if (w1 or w2) else 'X'

    for i, code in enumerate(codes):
        r = 4 + i
        ws.cell(r, 1).value = i + 1
        ws.cell(r, 2).value = str(code)[:8]
        ws.cell(r, 3).value = str(code)
        ws.cell(r, 4).value = dong_names.get(str(code)[:8], '')
        # 인구 연도열
        pv = []
        for k, y in enumerate(pop_years):
            cc = pop_s + k; cl = L(cc)
            ws.cell(r, cc).value = f"=VLOOKUP({cl}$3&$C{r},{pop_rng},3,FALSE)"
            v = pop_lut.get((y, str(code)))
            pv.append(v if v is not None else float('nan'))
            if v is not None and not (isinstance(v, float) and np.isnan(v)):
                cache[f'{cl}{r}'] = repr(float(v))
        pv_num = [x for x in pv if not (x is None or np.isnan(x))]
        mxp = max(pv_num) if pv_num else 0
        last_p = pv[-1] if pv else float('nan')
        R = 0.0 if (not pv_num or mxp == 0 or np.isnan(last_p)) else (last_p - mxp) / mxp * 100
        ws.cell(r, Rc).value = f"=IFERROR(({PE_}{r}-MAX({E_}{r}:{PE_}{r}))/MAX({E_}{r}:{PE_}{r})*100, 0)"
        ws.cell(r, Sc).value = f'=IF({L(Rc)}{r}<=-20,"O","X")'
        ws.cell(r, Tc).value = (f'=IF(OR(AND({L(pop_e-3)}{r}<{L(pop_e-4)}{r},{L(pop_e-2)}{r}<{L(pop_e-3)}{r},{L(pop_e-1)}{r}<{L(pop_e-2)}{r}),'
                                f'AND({L(pop_e-2)}{r}<{L(pop_e-3)}{r},{L(pop_e-1)}{r}<{L(pop_e-2)}{r},{L(pop_e)}{r}<{L(pop_e-1)}{r})),"O","X")') if npy >= 5 else '=\"X\"'
        cache[f'{L(Rc)}{r}'] = repr(float(R))
        Sv = 'O' if R <= -20 else 'X'
        cache[f'{L(Sc)}{r}'] = Sv
        Tv = seq_decline(pv[-5:]) if npy >= 5 else 'X'
        cache[f'{L(Tc)}{r}'] = Tv
        # 사업체 연도열
        bv = []
        for k, y in enumerate(biz_years):
            cc = biz_s + k; cl = L(cc)
            ws.cell(r, cc).value = f"=VLOOKUP({cl}$3&$C{r},{biz_rng},3,FALSE)"
            v = biz_lut.get((y, str(code)))
            bv.append(v if v is not None else float('nan'))
            if v is not None and not (isinstance(v, float) and np.isnan(v)):
                cache[f'{cl}{r}'] = repr(float(v))
        bv_num = [x for x in bv if not (x is None or np.isnan(x))]
        mxb = max(bv_num) if bv_num else 0
        last_b = bv[-1] if bv else float('nan')
        AE = 0.0 if (not bv_num or mxb == 0 or np.isnan(last_b)) else (last_b - mxb) / mxb * 100
        ws.cell(r, AEc).value = f"=IFERROR(({BE_}{r}-MAX({U_}{r}:{BE_}{r}))/MAX({U_}{r}:{BE_}{r})*100, 0)"
        ws.cell(r, AFc).value = f'=IF({L(AEc)}{r}<-5,"O","X")'
        ws.cell(r, AGc).value = (f'=IF(OR(AND({L(biz_e-3)}{r}<{L(biz_e-4)}{r},{L(biz_e-2)}{r}<{L(biz_e-3)}{r},{L(biz_e-1)}{r}<{L(biz_e-2)}{r}),'
                                 f'AND({L(biz_e-2)}{r}<{L(biz_e-3)}{r},{L(biz_e-1)}{r}<{L(biz_e-2)}{r},{L(biz_e)}{r}<{L(biz_e-1)}{r})),"O","X")') if nby >= 5 else '=\"X\"'
        cache[f'{L(AEc)}{r}'] = repr(float(AE))
        AFv = 'O' if AE < -5 else 'X'
        cache[f'{L(AFc)}{r}'] = AFv
        AGv = seq_decline(bv[-5:]) if nby >= 5 else 'X'
        cache[f'{L(AGc)}{r}'] = AGv
        # 건축(피벗 VLOOKUP)
        ws.cell(r, AHc).value = f"=VLOOKUP(C{r},{ho_rng},2,FALSE)"
        ws.cell(r, AIc).value = f"=VLOOKUP(C{r},{ho_rng},3,FALSE)"
        ws.cell(r, AJc).value = f"={L(AHc)}{r}/{L(AIc)}{r}*100"
        ws.cell(r, AKc).value = f'=IF({L(AJc)}{r}>=50,"O","X")'
        bands = piv.get(str(code))
        if bands:
            old = sum(float(bands.get(k, 0)) for k in (1, 2, 3, 4))
            tot = sum(float(v) for v in bands.values())
            cache[f'{L(AHc)}{r}'] = repr(old); cache[f'{L(AIc)}{r}'] = repr(tot)
            AJ = (old / tot * 100) if tot else 0
            cache[f'{L(AJc)}{r}'] = repr(float(AJ))
            AKv = 'O' if AJ >= 50 else 'X'
        else:
            AKv = 'X'
        cache[f'{L(AKc)}{r}'] = AKv
        # 부합개수 / 쇠퇴지역
        ws.cell(r, ALc).value = (f'=IF(OR({L(Sc)}{r}="o", {L(Tc)}{r}="o"), 1, 0) + '
                                 f'IF(OR({L(AFc)}{r}="o", {L(AGc)}{r}="o"), 1, 0) + IF({L(AKc)}{r}="o", 1, 0)')
        ws.cell(r, AMc).value = f'=IF({L(ALc)}{r}>=2,"O","X")'
        AL = (1 if Sv == 'O' or Tv == 'O' else 0) + (1 if AFv == 'O' or AGv == 'O' else 0) + (1 if AKv == 'O' else 0)
        cache[f'{L(ALc)}{r}'] = str(AL)
        cache[f'{L(AMc)}{r}'] = 'O' if AL >= 2 else 'X'
    ws.freeze_panes = 'E4'
    return cache


def build_full_workbook(raw_sub, values=None, admin_path=None,
                        indicators=TEMPLATE_INDICATORS, selected_years=None):
    """원본 9시트 전부(계산방법·법적종합·복합종합·원시6) 워크북. 집계구 단위.
    raw_sub: 엔진 raw dict(연도·집계구·CODE·값·행정동코드). values: 복합 지표값(없으면 raw로 계산)."""
    dong_names, _ = load_admin_names(admin_path or DEFAULT_ADMIN_PATH)
    ids = [i[0] for i in indicators]
    vals = values.reindex(columns=ids) if values is not None else compute_values(raw_sub, ids, key='집계구')
    vals = vals[vals.index.astype(str).str.len() == 14]
    codes = sorted(vals.index.astype(str))
    vals = vals.loc[codes]
    code_set = set(codes)
    # raw를 이 집계구 집합으로 한정(원시시트 무관 행 제거)
    rsub = {}
    for b, df in (raw_sub or {}).items():
        if df is None or len(df) == 0:
            rsub[b] = df; continue
        d = df.copy(); d['집계구'] = d['집계구'].astype(str)
        rsub[b] = d[d['집계구'].isin(code_set)]

    wb = openpyxl.Workbook()
    cache = {}
    # ① 계산방법
    write_calc_method(wb.active, indicators)
    wb.active.title = CALC_SHEET
    # ② 1 법적쇠퇴진단 종합 (원시시트 행수 먼저 계산 필요 → 시트 생성/기록은 뒤, 행수 산정)
    ws_legal = wb.create_sheet(LEGAL_SHEET)
    # ③ 2 복합쇠퇴진단 종합
    ws_comp = wb.create_sheet('2 복합쇠퇴진단 종합')
    grades = _composite_grades(codes, vals, indicators)
    write_composite(ws_comp, codes, vals, dong_names, indicators, grades=grades)
    cache['2 복합쇠퇴진단 종합'] = _composite_cached_values(codes, vals, indicators)
    # ④~⑨ 원시 6시트 (행수 기록 → 법적 VLOOKUP 범위에 사용)
    rowcount = {}
    for sheet_name, bucket in RAW_SHEETS:
        ws = wb.create_sheet(sheet_name)
        n = write_raw_sheet(ws, rsub.get(bucket), bucket)
        rowcount[bucket] = n
        if bucket == 'ho_yr':
            npiv, ho_cache = write_ho_yr_pivot(ws, rsub.get('ho_yr'))
            rowcount['ho_pivot'] = npiv
            if ho_cache:
                cache[sheet_name] = ho_cache
    # 법적종합 채우기(원시 행수 확정 후)
    lc = write_legal(ws_legal, codes, dong_names, rsub,
                     rowcount.get('to_in', 0), rowcount.get('to_fa', 0), rowcount.get('ho_pivot', 0))
    cache[LEGAL_SHEET] = lc
    wb._tmpl_cache = cache
    return wb


def _put_cache(m, ref, val):
    """유한수만 캐시맵에 넣는다(문자열 repr)."""
    try:
        f = float(val)
    except (TypeError, ValueError):
        return
    if not np.isfinite(f):
        return
    m[ref] = repr(f)


def _composite_at(codes, vals, indicators):
    """복합종합 수식과 동일 규칙으로 Z·T·부문점수·종합(AT)을 파이썬에서 계산.
    Z=(V-평균)/STDEV.P, T=Z×방향부호+50(빈칸=0), 부문=ΣT×가중치, 종합=Σ부문.
    반환: (Zarr{i→ndarray}, Tarr{i→ndarray}, sec_sum{부문→ndarray}, AT ndarray)."""
    n = len(codes)
    Zarr, Tarr = {}, {}
    for i, (iid, hdr, sec, sign, w) in enumerate(indicators):
        if iid in vals.columns:
            v = pd.to_numeric(vals[iid], errors='coerce').to_numpy(dtype=float)
        else:
            v = np.full(n, np.nan)
        mask = ~np.isnan(v)
        if mask.sum() > 0:
            mean = float(v[mask].mean()); std = float(v[mask].std(ddof=0))   # 모표준편차
        else:
            mean = std = 0.0
        if i == 0:                                   # 첫 지표: IF 가드 없음(빈칸=0 취급)
            vv = np.where(np.isnan(v), 0.0, v)
            z = (vv - mean) / std if std > 0 else np.zeros(n)
            t = z * sign + 50.0
        else:                                        # 나머지: 빈칸이면 Z=T=0
            z = np.where(np.isnan(v), 0.0, ((v - mean) / std) if std > 0 else 0.0)
            t = np.where(np.isnan(v), 0.0, z * sign + 50.0)
        Zarr[i], Tarr[i] = z, t
    sec_idx = {s: [i for i, ind in enumerate(indicators) if ind[2] == s] for s in SECTORS}
    sec_sum, AT = {}, np.zeros(n)
    for sec in SECTORS:
        ssum = np.zeros(n)
        for i in sec_idx[sec]:
            ssum = ssum + Tarr[i] * float(indicators[i][4])
        sec_sum[sec] = ssum
        AT = AT + ssum
    return Zarr, Tarr, sec_sum, AT


def _jenks_grades(at_values, k=10):
    """종합(AT) 값 → Natural Breaks(Jenks) k등급. 종합이 클수록 1등급(쇠퇴 심함).
    엑셀에는 Jenks 함수가 없어 파이썬(DP)으로 최적 자연분류 경계를 구해 등급을 매긴다.
    반환: at_values와 같은 순서의 정수 등급 리스트(1~k)."""
    at = np.asarray(list(at_values), dtype=float)
    n = at.size
    if n == 0:
        return []
    order = np.argsort(at, kind='mergesort')            # 오름차순(안정 정렬)
    s = at[order]
    keff = int(min(k, np.unique(s).size))
    if keff <= 1:                                        # 값이 사실상 동일 → 전부 1등급
        return [1] * n
    p1 = np.concatenate(([0.0], np.cumsum(s)))
    p2 = np.concatenate(([0.0], np.cumsum(s * s)))
    dp = np.full((keff + 1, n + 1), np.inf)
    split = np.zeros((keff + 1, n + 1), dtype=np.int64)
    idx = np.arange(1, n + 1)
    dp[1, 1:] = p2[1:] - p1[1:] ** 2 / idx               # segvar(1..l): 1분류 비용
    for j in range(2, keff + 1):                         # 분류 수
        for l in range(j, n + 1):                        # 앞 l개 점
            ss = np.arange(j - 1, l)                      # 직전 경계 후보(앞 분류 원소 수)
            a = ss + 1                                    # 마지막 분류 시작(1-index)
            cnt = l - a + 1
            seg = (p2[l] - p2[a - 1]) - (p1[l] - p1[a - 1]) ** 2 / cnt
            cost = dp[j - 1, ss] + seg
            m = int(np.argmin(cost))
            dp[j, l] = cost[m]; split[j, l] = int(ss[m])
    classes = np.empty(n, dtype=np.int64)
    l = n
    for j in range(keff, 0, -1):                         # 역추적: 경계로 class 배정
        st = int(split[j, l]) if j > 1 else 0
        classes[st:l] = j                                # 정렬순 st..l-1 → class j(값 클수록 큰 class)
        l = st
    # class(1=최저값 … keff=최고값) → 등급(최고값=1등급, 최저값=k등급으로 스케일)
    grade_sorted = 1 + np.rint((keff - classes) * (k - 1) / (keff - 1)).astype(int)
    out = np.empty(n, dtype=np.int64)
    out[order] = grade_sorted
    return out.tolist()


def _composite_grades(codes, vals, indicators, k=10):
    """복합종합 종합(AT) 기준 Natural Breaks(Jenks) 등급 리스트(codes 순서)."""
    if len(codes) == 0:
        return []
    _z, _t, _s, AT = _composite_at(codes, vals, indicators)
    return _jenks_grades(AT, k)


def _composite_cached_values(codes, vals, indicators):
    """복합종합 시트 수식셀의 '계산된 값'을 파이썬에서 산출 → {셀주소: 값문자열}.
    엑셀 수식(Z=(V-AVG)/STDEV.P, T=Z×부호+50, 부문=ΣT×가중치, 종합=Σ부문, 등급)과
    동일 규칙으로 계산해 셀에 캐시로 박는다(수식은 그대로 유지, 열자마자 값 표시).
    ※ Natural Break(AW)는 수식이 아니라 정적 등급값이라 여기서 다루지 않는다(write_composite가 기록)."""
    n = len(codes)
    if n == 0:
        return {}
    Zarr, Tarr, sec_sum, AT = _composite_at(codes, vals, indicators)
    vmap = {}
    for i, (iid, hdr, sec, sign, w) in enumerate(indicators):
        vcol = 4 + i * 3
        Zl = get_column_letter(vcol + 1); Tl = get_column_letter(vcol + 2)
        z, t = Zarr[i], Tarr[i]
        for r in range(n):
            _put_cache(vmap, f'{Zl}{3 + r}', z[r]); _put_cache(vmap, f'{Tl}{3 + r}', t[r])
    base = 4 + len(indicators) * 3
    for s_i, sec in enumerate(SECTORS):
        wcol = base + s_i * 2
        ANl = get_column_letter(wcol); AOl = get_column_letter(wcol + 1)
        ssum = sec_sum[sec]
        for r in range(n):
            _put_cache(vmap, f'{ANl}{3 + r}', ssum[r]); _put_cache(vmap, f'{AOl}{3 + r}', ssum[r])
    total_col = base + len(SECTORS) * 2
    ATl = get_column_letter(total_col); AUl = get_column_letter(total_col + 1); AVl = get_column_letter(total_col + 2)
    mx = float(AT.max()); mn = float(AT.min())
    for r in range(n):
        _put_cache(vmap, f'{ATl}{3 + r}', AT[r])
        rank = 1 + int(np.sum(AT > AT[r]))                       # RANK(…,0) 내림차순(동점 동일)
        vmap[f'{AUl}{3 + r}'] = str(min(10, int(np.ceil(rank / (n / 10.0)))))          # 등분위
        av = max(1, min(10, int(np.ceil((mx - AT[r]) / ((mx - mn) / 10.0))))) if mx > mn else 1
        vmap[f'{AVl}{3 + r}'] = str(av)                                                # 등간격
    return vmap


def _inject_cached(xml, valmap):
    """워크시트 XML의 수식셀(<f>있음)에 캐시값 <v>를 주입.
    문자열 결과(O/X 등)는 셀에 t="str"를 붙여야 엑셀이 읽는다."""
    import re as _re

    def _is_num(s):
        try:
            float(s); return True
        except (TypeError, ValueError):
            return False

    def repl(m):
        body = m.group(0); ref = m.group(1)
        v = valmap.get(ref)
        if v is None or '<f>' not in body:
            return body
        body = body.replace('<v/>', '<v></v>')
        if _is_num(v):
            vtag = f'<v>{v}</v>'
        else:                                        # 문자열 결과 → t="str" 필요
            esc = str(v).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            vtag = f'<v>{esc}</v>'
            gt = body.find('>')
            open_tag = body[:gt + 1]
            if ' t=' not in open_tag:
                body = open_tag[:-1] + ' t="str">' + body[gt + 1:]
        if '<v></v>' in body:                        # openpyxl가 남긴 빈 값칸을 채움
            return body.replace('<v></v>', vtag)
        if '<v>' in body:                            # 이미 값칸이 있으면 교체
            return _re.sub(r'<v>.*?</v>', vtag, body, count=1, flags=_re.DOTALL)
        return body[:-4] + vtag + '</c>'             # 끝의 </c> 앞에 <v> 삽입
    return _re.sub(r'<c r="([A-Z]+\d+)"[^>]*>.*?</c>', repl, xml, flags=_re.DOTALL)


def save_wb(wb, target):
    """wb에 `_tmpl_cache`(시트명→캐시맵)가 있으면 저장 후 해당 시트 수식셀에 값 주입.
    없으면 그냥 저장. target: 파일경로 또는 write() 가능한 객체."""
    import io as _io
    import re as _re
    import zipfile as _zip
    cache = getattr(wb, '_tmpl_cache', None)
    if not cache:
        wb.save(target); return
    tmp = _io.BytesIO(); wb.save(tmp); tmp.seek(0)
    zin = _zip.ZipFile(tmp, 'r')
    wbxml = zin.read('xl/workbook.xml').decode('utf-8')
    rels = zin.read('xl/_rels/workbook.xml.rels').decode('utf-8')
    rid_tgt = {}
    for tag in _re.findall(r'<Relationship\b[^>]*/>', rels):
        _id = _re.search(r'Id="([^"]+)"', tag); _t = _re.search(r'Target="([^"]+)"', tag)
        if _id and _t:
            rid_tgt[_id.group(1)] = _t.group(1)
    title_path = {}
    for tag in _re.findall(r'<sheet\b[^>]*/>', wbxml):
        nm = _re.search(r'name="([^"]+)"', tag); rid = _re.search(r'r:id="([^"]+)"', tag)
        if nm and rid:
            tgt = rid_tgt.get(rid.group(1), '')
            if tgt and not tgt.startswith('/'):
                tgt = 'xl/' + tgt
            title_path[nm.group(1)] = tgt.lstrip('/')
    # compresslevel=1: 원시 대용량 시트 재압축 CPU를 크게 줄임(파일은 조금 커지지만 배치 빌드가 빨라짐)
    out = _io.BytesIO(); zout = _zip.ZipFile(out, 'w', _zip.ZIP_DEFLATED, compresslevel=1)
    for item in zin.infolist():
        data = zin.read(item.filename)
        for title, path in title_path.items():
            if path == item.filename and cache.get(title):
                data = _inject_cached(data.decode('utf-8'), cache[title]).encode('utf-8')
                break
        zout.writestr(item, data)
    zout.close(); zin.close(); out.seek(0)
    if hasattr(target, 'write'):
        target.write(out.getvalue())
    else:
        with open(target, 'wb') as fh:
            fh.write(out.getvalue())


def build_composite_workbook(raw_sub=None, admin_path=None, indicators=TEMPLATE_INDICATORS,
                             values=None):
    """계산방법 + 복합쇠퇴진단 종합 2시트 wb (집계구 단위).
    indicators = TEMPLATE_INDICATORS(기본) 또는 indicators_from_cfg(cfg)로 만든
    ③설정 반영 목록. 값·계산방법·종합 셋 다 같은 목록으로 산출해 E열 참조가 일치한다.
    · values 주어지면(앱 scores에서 뽑은 base+custom+recipe 전체 지표값) 그걸 사용.
    · 없으면 raw_sub에서 기본지표만 계산(standalone·검증용)."""
    dong_names, _ = load_admin_names(admin_path or DEFAULT_ADMIN_PATH)
    ids = [i[0] for i in indicators]
    vals = values.reindex(columns=ids) if values is not None else compute_values(raw_sub, ids, key='집계구')
    vals = vals[vals.index.astype(str).str.len() == 14]      # 집계구(14자리)만
    codes = sorted(vals.index.astype(str))
    vals = vals.loc[codes]
    wb = openpyxl.Workbook()
    write_calc_method(wb.active, indicators)
    wb.active.title = CALC_SHEET
    ws2 = wb.create_sheet('2 복합쇠퇴진단 종합')
    grades = _composite_grades(codes, vals, indicators)
    write_composite(ws2, codes, vals, dong_names, indicators, grades=grades)
    # 수식셀 계산값을 캐시로 박아 열자마자 값이 보이게(수식은 유지) → save_wb에서 주입
    wb._tmpl_cache = {'2 복합쇠퇴진단 종합': _composite_cached_values(codes, vals, indicators)}
    return wb
