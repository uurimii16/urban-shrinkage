# -*- coding: utf-8 -*-
"""
verify_jeonju.py — Stage 1 셀단위 검증
=======================================
골든본 '(수정)' 시트의 원시 long 을 입력으로 엔진을 돌려,
같은 워크북의 '전주시 복합쇠퇴지수(행정동)' 시트 계산값(정답)을 재현하는지
행정동(34) × 지표(12) × {값·Z·T} + 부문·종합 을 셀단위로 대조합니다.

실행법 (Windows PowerShell / cmd):
    cd C:\\Users\\PC\\Desktop\\쇠퇴진단엔진
    set PYTHONUTF8=1
    python verify_jeonju.py
필요: pandas, numpy, openpyxl.
"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
import numpy as np
import pandas as pd
import openpyxl
import config as C
import decline_engine as E

GOLDEN = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', 'golden.xlsx')

# ── 원시 long 추출 스펙: (시트, 연도열, 집계구열, CODE열, 값열, 접두어) 0-based ──
RAW_SPECS = {
 'to_in': ('법적DATA_1.인구총괄(총인구)(수정)',            0, 1, 3, 4, 'to_in'),
 'to_fa': ('복합DATA_2.산업경제-총괄사업체수(수정)',        0, 1, 2, 3, 'to_fa'),
 'cp_bem':('복합DATA_2.산업경제-종사자수(수정)',            0, 1, 2, 5, 'cp_bem'),
 'in_age':('복합 DATA_1.인문사회(수정)',                   0, 1, 2, 3, 'in_age'),
 'ho_ar': ('복합DATA_3.물리환경-소형주택비율(수정)',        0, 1, 2, 5, 'ho_ar'),
 'ho_yr': ('복합DATA_3.물리환경-주택건축물비율(수정)',      0, 1, 2, 4, 'ho_yr'),
}

# ── 골든 정답(복합쇠퇴지수(행정동)) 열 위치: 지표 → (값,Z,T) 컬럼문자 ──
ANS_COLS = {
 '인구변화율':      ('D', 'E', 'F'),
 '노년부양비':      ('H', 'I', 'J'),
 '경제활동인구비율': ('L', 'M', 'N'),
 '소멸위험지수':    ('P', 'Q', 'R'),
 '총사업체수증감률': ('T', 'U', 'V'),
 '총종사자수증감률': ('X', 'Y', 'Z'),
 '제조업증감률':    ('AB', 'AC', 'AD'),
 '고차산업증감률':  ('AF', 'AG', 'AH'),
 '도소매증감률':    ('AJ', 'AK', 'AL'),
 '음식숙박증감률':  ('AN', 'AO', 'AP'),
 '노후건축물비율':  ('AR', 'AS', 'AT'),
 # 소형주택비율(AV,AW,AX): 2026-07-06 기본지표에서 제외 → 골든 대조 대상 아님.
}
ANS_SECTOR = {'인문사회': 'BB', '산업경제': 'BE', '물리환경': 'BH', '종합': 'BJ'}
ANS_ROWS = range(3, 37)      # 34개 행정동
ANS_CODE_COL = 'A'


def _to_num(x):
    if x is None: return np.nan
    if isinstance(x, (int, float)): return float(x)
    s = str(x).strip()
    if s in ('', 'N/A', 'NA', 'na', 'n/a', '#N/A'): return np.nan
    try: return float(s)
    except: return np.nan


def load_raw_from_golden(wb):
    """(수정) 시트에서 6개 원시 long 추출 → dict[str, DataFrame]."""
    raw = {}
    for key, (sh, ciy, cig, cic, civ, pref) in RAW_SPECS.items():
        ws = wb[sh]
        maxc = max(ciy, cig, cic, civ) + 1
        rows = []
        for r in ws.iter_rows(min_row=2, max_col=maxc, values_only=True):
            code = r[cic]
            if code is None or not str(code).startswith(pref):
                continue
            yr, gu, val = r[ciy], r[cig], r[civ]
            if yr is None or gu is None:
                continue
            try: yr = int(yr)
            except: continue
            gu = str(gu).strip().replace('.0', '') if str(gu).endswith('.0') else str(gu).strip()
            rows.append((yr, gu, str(code).strip(), _to_num(val)))
        df = pd.DataFrame(rows, columns=['연도', '집계구', 'CODE', '값'])
        df['행정동코드'] = df['집계구'].str[:8]
        raw[key] = df
    return raw


def load_golden_answers(wb):
    """복합쇠퇴지수(행정동) 계산값(정답) → DataFrame(index=행정동코드)."""
    ws = wb['전주시 복합쇠퇴지수(행정동)']
    ci = openpyxl.utils.column_index_from_string
    recs = {}
    for r in ANS_ROWS:
        code = ws.cell(row=r, column=ci(ANS_CODE_COL)).value
        if code is None:
            continue
        code = str(code).strip()
        rec = {}
        for ind, (cv, cz, ct) in ANS_COLS.items():
            rec[(ind, '값')] = _to_num(ws.cell(row=r, column=ci(cv)).value)
            rec[(ind, 'Z')]  = _to_num(ws.cell(row=r, column=ci(cz)).value)
            rec[(ind, 'T')]  = _to_num(ws.cell(row=r, column=ci(ct)).value)
        for sec, col in ANS_SECTOR.items():
            rec[('부문', sec)] = _to_num(ws.cell(row=r, column=ci(col)).value)
        recs[code] = rec
    df = pd.DataFrame(recs).T
    df.columns = pd.MultiIndex.from_tuples(df.columns)
    df.index.name = '행정동코드'
    return df


def compare(engine_scores, engine_comp, golden, abs_tol=1e-6, rel_tol=1e-4):
    """
    엔진값 vs 골든값 셀단위 대조.
    허용오차: |diff|<=abs_tol  또는  상대오차<=rel_tol(0.01%).
    반환: (요약 DataFrame, 불일치 목록 DataFrame)
    """
    idx = golden.index
    rows_summary = []
    mism = []

    def check(label, eng_series, gold_series):
        eng = eng_series.reindex(idx)
        n_ok = 0; n_tot = 0
        for code in idx:
            e = eng.get(code, np.nan); g = gold_series.get(code, np.nan)
            n_tot += 1
            if pd.isna(e) and pd.isna(g):
                n_ok += 1; continue
            if pd.isna(e) or pd.isna(g):
                mism.append((label, code, e, g, np.nan)); continue
            diff = abs(e - g)
            ok = (diff <= abs_tol) or (diff <= rel_tol * max(abs(e), abs(g), 1e-9))
            if ok: n_ok += 1
            else:  mism.append((label, code, e, g, diff))
        rows_summary.append((label, n_ok, n_tot, 100.0 * n_ok / n_tot))

    # 지표별 값/Z/T
    for ind in C.IND_IDS:
        check(f"{ind}|값", engine_scores[(ind, '값')], golden[(ind, '값')])
        check(f"{ind}|Z",  engine_scores[(ind, 'Z')],  golden[(ind, 'Z')])
        check(f"{ind}|T",  engine_scores[(ind, 'T')],  golden[(ind, 'T')])
    # 부문·종합
    for sec in C.SECTORS:
        check(f"부문|{sec}", engine_comp[sec], golden[('부문', sec)])
    check("부문|종합", engine_comp['종합'], golden[('부문', '종합')])

    summary = pd.DataFrame(rows_summary, columns=['항목', '일치', '전체', '일치율(%)'])
    mismdf = pd.DataFrame(mism, columns=['항목', '행정동코드', '엔진값', '골든값', '오차'])
    return summary, mismdf


def main():
    print("골든본 로드 중 (data_only, read_only)…")
    wb = openpyxl.load_workbook(GOLDEN, read_only=True, data_only=True)
    raw = load_raw_from_golden(wb)
    golden = load_golden_answers(wb)
    wb.close()
    print(f"  원시 long 추출: " + ", ".join(f"{k}={len(v)}행" for k, v in raw.items()))
    print(f"  정답 행정동 수: {len(golden)}")

    print("\n엔진 실행 중…")
    scores, comp, grades, stats = E.run(raw, grade_method='jenks')

    summary, mism = compare(scores, comp, golden)

    pd.set_option('display.unicode.east_asian_width', True)
    pd.set_option('display.width', 200)
    print("\n" + "=" * 64)
    print(" Stage 1 검증 결과 — 지표별 일치율 (허용오차 abs 1e-6 / rel 0.01%)")
    print("=" * 64)
    print(summary.to_string(index=False))

    # 지표 셀(값/Z/T)과 부문/종합을 분리 판정.
    #  · 지표 값/Z/T : 엔진 정확성 → 100% 여야 함.
    #  · 부문/종합   : 2026-07-06 소형주택비율(물리환경, 가중치 0.150)을 기본지표에서
    #                 제외했으므로 12지표 골든의 '물리환경'·'종합'과는 다른 게 정상.
    is_comp = summary['항목'].str.startswith('부문|')
    cell = summary[~is_comp]
    cell_ok, cell_tot = int(cell['일치'].sum()), int(cell['전체'].sum())
    print(f"\n  ▶ 지표 값/Z/T 일치율: {cell_ok}/{cell_tot} = {100.0*cell_ok/cell_tot:.4f}%"
          f"  {'✅ 완전 재현' if cell_ok == cell_tot else '❌ 엔진 회귀!'}")

    comp_mism = mism[mism['항목'].str.startswith('부문|')]
    cell_mism = mism[~mism['항목'].str.startswith('부문|')]
    print("  ▶ 부문/종합: 소형주택비율 제외로 '물리환경'·'종합'은 12지표 골든과 다름이 정상"
          f" (불일치 {len(comp_mism)}건 — 예상 68=물리환경34+종합34).")

    if len(cell_mism):
        print("\n" + "-" * 64)
        print(f" ⚠ 지표 셀 불일치: {len(cell_mism)}건 (엔진 회귀 — 조사 필요, 상위 40건)")
        print("-" * 64)
        print(cell_mism.head(40).to_string(index=False))
    else:
        print("\n  ✔ 지표 값/Z/T 불일치 0건 — 엔진이 골든본 지표계산을 완전 재현했습니다.")

    # 통계행(평균·표준편차) 참고 출력
    print("\n[참고] 엔진 표준화 통계 (지표별 평균 / 모표준편차)")
    st = pd.DataFrame(stats).T[['mean', 'std']]
    print(st.to_string())

    # 결과 저장
    out = os.path.join(os.path.dirname(os.path.abspath(__file__)), '검증결과_전주.xlsx')
    with pd.ExcelWriter(out, engine='openpyxl') as w:
        summary.to_excel(w, sheet_name='일치율요약', index=False)
        mism.to_excel(w, sheet_name='불일치목록', index=False)
        eng_flat = scores.copy(); eng_flat.columns = [f"{a}_{b}" for a, b in eng_flat.columns]
        eng_flat.join(comp).to_excel(w, sheet_name='엔진산출값')
    print(f"\n결과 저장: {out}")


if __name__ == '__main__':
    main()
