# -*- coding: utf-8 -*-
"""
sheet_builder.py — 원시 long dict → 검토용 DATA/피벗/통합 xlsx 시트 생성
=====================================================================
엔진 입력 스키마(raw dict)를 그대로 받아 법적/복합 DATA 개별 시트와
행정동 피벗, 최종 진단 시트를 한 워크북에 구성한다.
"""
from __future__ import annotations

import math
from typing import Iterable

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

import config as C
import decline_engine as E
import export


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
SUB_FILL = PatternFill("solid", fgColor="D9E1F2")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
HEADER_FONT = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)


RAW_SHEET_DEFS = {
    "legal_population": ("법적DATA_1.인구총괄(수정)", "to_in", "법적인구"),
    "legal_business": ("법적DATA_2.총괄사업체수(수정)", "to_fa", "법적산업"),
    "legal_physical": ("법적DATA_3.건축연도별주택(수정)", "ho_yr", "법적물리"),
    "complex_human": ("복합DATA_1.인문사회(수정)", "in_age", "복합인문"),
    "complex_business_total": ("복합DATA_2.총괄사업체수(수정)", "to_fa", "복합산업-사업체"),
    "complex_business_worker": ("복합DATA_2.종사자수(수정)", "cp_bem", "복합산업-종사자"),
    "complex_small_house": ("복합DATA_3.소형주택비율(수정)", "ho_ar", "복합물리-소형"),
    "complex_old_house": ("복합DATA_3.주택건축물비율(수정)", "ho_yr", "복합물리-노후"),
}

PIVOT_DEFS = {
    "pivot_human": ("복합DATA_1.인문사회(행정동)", "in_age"),
    "pivot_business_total": ("복합DATA_2.사업체수(행정동)", "to_fa"),
    "pivot_business_worker": ("복합DATA_2.종사자수(행정동)", "cp_bem"),
    "pivot_small_house": ("복합DATA_3.소형주택(행정동)", "ho_ar"),
    "pivot_old_house": ("복합DATA_3.노후건축물(행정동)", "ho_yr"),
}

DEFAULT_OPTIONS = {
    "legal_population": True,
    "legal_business": True,
    "legal_physical": True,
    "complex_human": True,
    "complex_business_total": True,
    "complex_business_worker": True,
    "complex_small_house": True,
    "complex_old_house": True,
    "pivot_human": True,
    "pivot_business_total": True,
    "pivot_business_worker": True,
    "pivot_small_house": True,
    "pivot_old_house": True,
    "final_legal_dong": True,
    "final_legal_jgu": True,
    "final_complex_dong": True,
    "final_complex_jgu": True,
    "summary": True,
}


def all_years(raw: dict[str, pd.DataFrame]) -> list[int]:
    years = set()
    for df in raw.values():
        if "연도" in df.columns:
            years.update(pd.to_numeric(df["연도"], errors="coerce").dropna().astype(int).tolist())
    return sorted(years)


def filter_raw_years(raw: dict[str, pd.DataFrame], years: Iterable[int] | None):
    if not years:
        return raw
    selected = {int(y) for y in years}
    out = {}
    for key, df in raw.items():
        if df is None or "연도" not in df.columns:
            out[key] = pd.DataFrame(columns=["연도", "집계구", "CODE", "값", "행정동코드"])
        else:
            out[key] = df[df["연도"].astype(int).isin(selected)].copy()
    return out


def normalize_options(options: dict | None) -> dict:
    merged = dict(DEFAULT_OPTIONS)
    if options:
        merged.update(options)
    return merged


def _safe_title(wb, title: str) -> str:
    base = title[:31]
    if base not in wb.sheetnames:
        return base
    i = 2
    while True:
        suffix = f"_{i}"
        cand = (base[: 31 - len(suffix)] + suffix)
        if cand not in wb.sheetnames:
            return cand
        i += 1


def _sheet(wb, title: str):
    return wb.create_sheet(_safe_title(wb, title))


def _num(v):
    if v is None:
        return None
    try:
        f = float(v)
    except Exception:
        return None
    if math.isnan(f) or math.isinf(f):
        return None
    return round(f, 6)


def _style_header(ws, row: int = 1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _append_note(ws, text: str):
    ws.append([text])
    for cell in ws[ws.max_row]:
        cell.fill = WARN_FILL
        cell.font = BOLD


def _write_summary(ws, raw, selected_years, options):
    ws.append(["구분", "값"])
    _style_header(ws)
    ws.append(["선택 연도", ", ".join(str(y) for y in selected_years) if selected_years else "전체"])
    for key, df in raw.items():
        years = sorted(pd.to_numeric(df["연도"], errors="coerce").dropna().astype(int).unique().tolist())
        ws.append([f"{key} 행 수", len(df)])
        ws.append([f"{key} 연도", ", ".join(map(str, years))])
    enabled = [k for k, v in options.items() if v]
    ws.append(["생성 옵션", ", ".join(enabled)])
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 90


def _write_long_data(ws, df: pd.DataFrame, title: str, code_label_map=None):
    ws.append(["연도", "집계구", "연도집계구", "CODE", "값", "항목명", "행정동코드"])
    _style_header(ws)
    if df is None or df.empty:
        _append_note(ws, f"{title}: 입력 데이터 없음")
        return
    d = df.copy()
    d["연도"] = pd.to_numeric(d["연도"], errors="coerce")
    d["값"] = pd.to_numeric(d["값"], errors="coerce")
    d = d.sort_values(["연도", "집계구", "CODE"], kind="stable")
    dong = d["행정동코드"] if "행정동코드" in d.columns else d["집계구"].astype(str).str[:8]
    for yr, gu, code, val, dcode in zip(d["연도"], d["집계구"], d["CODE"], d["값"], dong):
        if pd.isna(yr):
            continue
        gu = str(gu).strip()
        code = str(code).strip()
        label = (code_label_map or {}).get(code, export.code_name(code))
        ws.append([int(yr), gu, f"{int(yr)}{gu}", code, _num(val), label, str(dcode)])
    ws.freeze_panes = "A2"


def _pivot_by_code(df: pd.DataFrame, code_filter=None, key="행정동코드"):
    if df is None or df.empty:
        return pd.DataFrame()
    d = df.copy()
    if code_filter is not None:
        d = d[d["CODE"].astype(str).str[-3:].astype(int).isin(code_filter)]
    d["값"] = pd.to_numeric(d["값"], errors="coerce")
    return (d.groupby([key, "연도"])["값"].sum(min_count=1).unstack("연도").sort_index())


def _write_pivot(ws, df: pd.DataFrame, label: str, key="행정동코드"):
    piv = _pivot_by_code(df, key=key)
    if piv.empty:
        ws.append([key])
        _style_header(ws)
        _append_note(ws, f"{label}: 입력 데이터 없음")
        return
    piv["총합계"] = piv.sum(axis=1, min_count=1)
    ws.append([key] + [str(c) for c in piv.columns])
    _style_header(ws)
    for unit, row in piv.iterrows():
        ws.append([str(unit)] + [_num(v) for v in row.values])
    ws.freeze_panes = "B2"


def _write_pivot_formula(ws, df: pd.DataFrame, label: str, source_sheet_title: str,
                         key="행정동코드", key_col="G"):
    """원시 long 시트를 SUMIFS로 집계하는 피벗.
    key_col = 원시 시트에서 집계 기준이 되는 열(문자). 행정동=G(행정동코드), 집계구=B(집계구)."""
    piv = _pivot_by_code(df, key=key)
    if piv.empty:
        ws.append([key])
        _style_header(ws)
        _append_note(ws, f"{label}: 입력 데이터 없음")
        return
    years = list(piv.columns)
    ws.append([key] + [str(c) for c in years] + ["총합계"])
    _style_header(ws)
    src = f"'{source_sheet_title}'"
    for unit in piv.index:
        row_no = ws.max_row + 1
        ws.cell(row=row_no, column=1, value=str(unit))
        for i, yr in enumerate(years, start=2):
            col = _excel_col(i)
            ws.cell(
                row=row_no,
                column=i,
                value=f"=SUMIFS({src}!$E:$E,{src}!${key_col}:${key_col},$A{row_no},{src}!$A:$A,{col}$1)",
            )
        ws.cell(row=row_no, column=len(years) + 2, value=f"=SUM(B{row_no}:{_excel_col(len(years)+1)}{row_no})")
    ws.freeze_panes = "B2"


def _write_indicator_values(ws, scores, key="행정동코드", indicator_ids=None, label_map=None,
                            formula_mode=False, legal_ref_sheet=None):
    indicator_ids = indicator_ids or C.IND_IDS
    label_map = label_map or C.INDLABEL
    ws.append([key] + [label_map.get(ind, ind) for ind in indicator_ids])
    _style_header(ws)
    for unit in scores.index:
        row_no = ws.max_row + 1
        ws.cell(row=row_no, column=1, value=str(unit))
        for pos, ind in enumerate(indicator_ids, start=2):
            if formula_mode and ind == "인구변화율" and legal_ref_sheet:
                ws.cell(row=row_no, column=pos, value=f"=IFERROR(VLOOKUP($A{row_no},'{legal_ref_sheet}'!$A:$L,3,FALSE),\"\")")
            else:
                value = _num(scores.at[unit, (ind, "값")]) if (ind, "값") in scores.columns else None
                ws.cell(row=row_no, column=pos, value=value)
    ws.freeze_panes = "B2"


def _write_final_legal(ws, legal_df, name_map, key_label):
    export._write_legal(ws, legal_df, name_map, key_label)


def _write_final_complex(ws, res, name_map, key_label, n_classes, method,
                         indicator_ids=None, label_map=None):
    scores, comp, grades = res
    ind_grades = export.indicator_grades(scores, n_classes, method, indicator_ids)
    export._write_composite(ws, scores, comp, grades, ind_grades, name_map, key_label,
                            indicator_ids, label_map)


def _excel_col(n):
    from openpyxl.utils import get_column_letter
    return get_column_letter(n)


def _write_final_complex_formula(ws, value_sheet_name, scores, name_map, key_label,
                                 indicator_ids, label_map, sector_of, weight, sign_map):
    hdr1 = [key_label, "명칭"]
    hdr2 = ["", ""]
    for ind in indicator_ids:
        hdr1 += [label_map.get(ind, ind), "", "", ""]
        hdr2 += ["값", "Z", "T", "가중T"]
    hdr1 += ["인문사회", "산업경제", "물리환경", "종합"]
    hdr2 += ["", "", "", ""]
    ws.append(hdr1)
    ws.append(hdr2)
    _style_header(ws, 1)
    for cell in ws[2]:
        cell.fill = SUB_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    n = len(scores.index)
    data_start = 3
    value_ref_sheet = f"'{value_sheet_name}'"
    score_cols = {}
    for pos, ind in enumerate(indicator_ids):
        base_col = 3 + pos * 4
        score_cols[ind] = {
            "value": _excel_col(base_col),
            "z": _excel_col(base_col + 1),
            "t": _excel_col(base_col + 2),
            "wt": _excel_col(base_col + 3),
        }
    sector_cols = {sec: _excel_col(3 + len(indicator_ids) * 4 + i) for i, sec in enumerate(C.SECTORS)}
    total_col = _excel_col(3 + len(indicator_ids) * 4 + len(C.SECTORS))

    for r_offset, unit in enumerate(scores.index, start=0):
        row_no = data_start + r_offset
        ws.cell(row=row_no, column=1, value=str(unit))
        ws.cell(row=row_no, column=2, value=((name_map or {}).get(unit) or (name_map or {}).get(str(unit)[:8], "")))
        for pos, ind in enumerate(indicator_ids):
            value_col = _excel_col(2 + pos + 1)
            value_cell = score_cols[ind]["value"] + str(row_no)
            z_cell = score_cols[ind]["z"] + str(row_no)
            t_cell = score_cols[ind]["t"] + str(row_no)
            wt_cell = score_cols[ind]["wt"] + str(row_no)
            value_range = f"{value_ref_sheet}!${value_col}${data_start}:${value_col}${data_start + n - 1}"
            raw_value = f"IFERROR(VLOOKUP($A{row_no},{value_ref_sheet}!$A:${_excel_col(2 + len(indicator_ids))},{pos + 2},FALSE),\"\")"
            sign = sign_map.get(ind, C.SIGN.get(ind, 10))
            ws[value_cell] = f"={raw_value}"
            ws[z_cell] = f"=IFERROR(({value_cell}-AVERAGE({value_range}))/STDEV.P({value_range}),0)"
            ws[t_cell] = f"={z_cell}*{sign}+50"
            ws[wt_cell] = f"={t_cell}*{float(weight.get(ind, 0.0))}"
        for sec in C.SECTORS:
            refs = [score_cols[ind]["wt"] + str(row_no) for ind in indicator_ids if sector_of.get(ind) == sec]
            ws[sector_cols[sec] + str(row_no)] = "=" + ("+".join(refs) if refs else "0")
        ws[total_col + str(row_no)] = "=" + "+".join(sector_cols[sec] + str(row_no) for sec in C.SECTORS)
    ws.freeze_panes = "C3"


def build_integrated_workbook(
    raw: dict[str, pd.DataFrame],
    *,
    selected_years: Iterable[int] | None = None,
    sheet_options: dict | None = None,
    name_map: dict | None = None,
    dong_res=None,
    jgu_res=None,
    legal_dong=None,
    legal_jgu=None,
    n_classes: int = 10,
    method: str = "jenks",
    indicator_ids=None,
    label_map=None,
    sector_of=None,
    weight=None,
    sign_map=None,
    code_label_map=None,
    formula_mode: bool = False,
    pivot_level: str = "dong",
    final_only: bool = False,
):
    """원시 DATA 시트 + 피벗 + 최종 진단 시트를 포함하는 통합 Workbook 생성.
    pivot_level = 'dong'(행정동, 기본) | 'jgu'(집계구) | 'both'(둘 다). 복합/피벗 집계 단위.
    final_only = True 면 중간 DATA/피벗/요약을 모두 생략하고 최종 진단 4개 시트만
                 (법적·복합 × 행정동·집계구) 자기완결형 '값'으로 출력(유림_17시 3~6 대응)."""
    opts = normalize_options(sheet_options)
    years = list(selected_years) if selected_years else all_years(raw)
    raw_f = filter_raw_years(raw, years)
    indicator_ids = indicator_ids or C.IND_IDS
    label_map = label_map or C.INDLABEL
    sector_of = sector_of or C.SECTOR_OF
    weight = weight or C.WEIGHT
    sign_map = sign_map or C.SIGN

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── 최종 4개 시트만 (중간과정 생략, 값·자기완결형) ──
    if final_only:
        if legal_dong is not None:
            _write_final_legal(_sheet(wb, "전주시 법적쇠퇴진단(행정동)"), legal_dong, name_map, "행정동코드")
        if legal_jgu is not None:
            _write_final_legal(_sheet(wb, "전주시 법적쇠퇴진단(집계구)"), legal_jgu, name_map, "집계구")
        if dong_res is not None:
            _write_final_complex(_sheet(wb, "전주시 복합쇠퇴지수(행정동)"), dong_res, name_map, "행정동코드",
                                 n_classes, method, indicator_ids, label_map)
        if jgu_res is not None:
            _write_final_complex(_sheet(wb, "전주시 복합쇠퇴지수(집계구)"), jgu_res, name_map, "집계구",
                                 n_classes, method, indicator_ids, label_map)
        if not wb.sheetnames:  # 방어: 아무 결과도 없으면 빈 시트 하나
            wb.create_sheet("결과없음")
        return wb

    if opts.get("summary", True):
        _write_summary(_sheet(wb, "생성요약"), raw_f, years, opts)

    source_sheet_by_raw = {}
    for opt_key, (title, raw_key, label) in RAW_SHEET_DEFS.items():
        if opts.get(opt_key):
            ws = _sheet(wb, title)
            _write_long_data(ws, raw_f.get(raw_key), label, code_label_map=code_label_map)
            source_sheet_by_raw[raw_key] = ws.title

    # 피벗 집계 단위: 행정동(G열 기준) / 집계구(B열 기준) / 둘 다
    level_specs = []
    if pivot_level in ("dong", "both"):
        level_specs.append(("행정동코드", "G", "(행정동)"))
    if pivot_level in ("jgu", "both"):
        level_specs.append(("집계구", "B", "(집계구)"))
    if not level_specs:
        level_specs = [("행정동코드", "G", "(행정동)")]
    for opt_key, (title, raw_key) in PIVOT_DEFS.items():
        if opts.get(opt_key):
            base = title.replace("(행정동)", "")
            for key_name, key_col, suffix in level_specs:
                t = base + suffix
                ws = _sheet(wb, t)
                if formula_mode and raw_key in source_sheet_by_raw:
                    _write_pivot_formula(ws, raw_f.get(raw_key), t, source_sheet_by_raw[raw_key],
                                         key=key_name, key_col=key_col)
                else:
                    _write_pivot(ws, raw_f.get(raw_key), t, key=key_name)

    legal_dong_sheet = None
    legal_jgu_sheet = None
    if opts.get("final_legal_dong") and legal_dong is not None:
        legal_dong_sheet = _sheet(wb, "전주시 법적쇠퇴진단(행정동)")
        _write_final_legal(legal_dong_sheet, legal_dong, name_map, "행정동코드")
    if opts.get("final_legal_jgu") and legal_jgu is not None:
        legal_jgu_sheet = _sheet(wb, "전주시 법적쇠퇴진단(집계구)")
        _write_final_legal(legal_jgu_sheet, legal_jgu, name_map, "집계구")

    value_sheet_dong = None
    value_sheet_jgu = None
    if any(opts.get(k) for k in ("final_complex_dong", "final_complex_jgu")):
        # 최종표와 같은 워크북에서 검토하기 쉽도록 12지표 파생값 원표도 추가한다.
        if dong_res is not None:
            value_sheet_dong = _sheet(wb, "복합지표값(행정동)")
            _write_indicator_values(value_sheet_dong, dong_res[0], key="행정동코드",
                                    indicator_ids=indicator_ids, label_map=label_map,
                                    formula_mode=formula_mode,
                                    legal_ref_sheet=legal_dong_sheet.title if legal_dong_sheet else None)
        if jgu_res is not None:
            value_sheet_jgu = _sheet(wb, "복합지표값(집계구)")
            _write_indicator_values(value_sheet_jgu, jgu_res[0], key="집계구",
                                    indicator_ids=indicator_ids, label_map=label_map,
                                    formula_mode=formula_mode,
                                    legal_ref_sheet=legal_jgu_sheet.title if legal_jgu_sheet else None)
    if opts.get("final_complex_dong") and dong_res is not None:
        ws = _sheet(wb, "전주시 복합쇠퇴지수(행정동)")
        if formula_mode and value_sheet_dong is not None:
            _write_final_complex_formula(ws, value_sheet_dong.title, dong_res[0], name_map, "행정동코드",
                                         indicator_ids, label_map, sector_of, weight, sign_map)
        else:
            _write_final_complex(ws, dong_res, name_map, "행정동코드", n_classes, method,
                                 indicator_ids, label_map)
    if opts.get("final_complex_jgu") and jgu_res is not None:
        ws = _sheet(wb, "전주시 복합쇠퇴지수(집계구)")
        if formula_mode and value_sheet_jgu is not None:
            _write_final_complex_formula(ws, value_sheet_jgu.title, jgu_res[0], name_map, "집계구",
                                         indicator_ids, label_map, sector_of, weight, sign_map)
        else:
            _write_final_complex(ws, jgu_res, name_map, "집계구", n_classes, method,
                                 indicator_ids, label_map)

    return wb
